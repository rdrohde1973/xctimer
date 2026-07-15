"""Schools + athletes (roster) — Phases 1-2 (handoff §8).

- School CRUD + bib blocks (admins).
- Roster: athlete CRUD, auto-bib assignment within the school's block.
- AI document import (Excel/CSV/PDF/Word -> Claude normalize -> preview -> commit).
- Google Sheet sync (share link -> CSV -> normalize).
- Bib list + Avery sticker PDFs (with QR).
- Bib check (scan QR / type bib -> athlete lookup), district-scoped.

Access: super_admin (any), district_admin (own district), coach (own schools).
Timers get no roster access.
"""
import io
import os

from markupsafe import escape
from flask import (Blueprint, request, redirect, g, abort, jsonify, Response)

from . import db, ai, pdfs, demo
from .auth import login_required, role_required
from .tenancy import (active_district_id, require_district, scoped_district_or_403,
                      all_districts)
from .ui import shell

bp = Blueprint("schools", __name__)


def _districts_for_switcher():
    return all_districts() if g.principal.is_super else None


def _can_access_school(school):
    p = g.principal
    if not p or p.role == "timer" or p.meet_scope:
        return False
    if p.is_super:
        return True
    if p.district_id != school["district_id"]:
        return False
    if p.role == "district_admin":
        return True
    if p.role == "coach":
        return school["id"] in p.school_ids()
    return False


def _load_school_or_403(sid):
    conn = db.connect()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (sid,)).fetchone()
    conn.close()
    if not s:
        abort(404)
    if not _can_access_school(s):
        abort(403)
    return s


def _visible_schools():
    """Schools the current principal may see, newest-context-first."""
    p = g.principal
    did = active_district_id()
    conn = db.connect()
    if p.role == "coach":
        rows = conn.execute(
            "SELECT s.*, d.name AS dname FROM schools s JOIN districts d ON d.id=s.district_id "
            "JOIN user_schools us ON us.school_id=s.id WHERE us.user_id=? ORDER BY s.name",
            (p.id,),
        ).fetchall()
    elif p.is_super and did is None:
        rows = conn.execute(
            "SELECT s.*, d.name AS dname FROM schools s JOIN districts d ON d.id=s.district_id "
            "ORDER BY d.name, s.name"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT s.*, NULL AS dname FROM schools s WHERE s.district_id=? ORDER BY s.name",
            (did,),
        ).fetchall()
    conn.close()
    return rows


def _slug(name):
    import re
    return re.sub(r"[^a-z0-9]+", "-", (name or "").lower()).strip("-") or "school"


def _save_logo(file_storage, prefix):
    """Save an uploaded school logo (optimized PNG) -> '/static/logos/<name>'."""
    if not file_storage or not file_storage.filename:
        return None
    import io
    import secrets
    from PIL import Image
    try:
        im = Image.open(io.BytesIO(file_storage.read())).convert("RGBA")
    except Exception:  # noqa: BLE001 — not a valid image
        return None
    if im.width > 400:
        im = im.resize((400, round(400 * im.height / im.width)), Image.LANCZOS)
    d = os.path.join(os.path.dirname(__file__), "static", "logos")
    os.makedirs(d, exist_ok=True)
    name = f"{_slug(prefix)}-{secrets.token_hex(4)}.png"
    im.save(os.path.join(d, name), format="PNG", optimize=True)
    return f"/static/logos/{name}"


BIB_BLOCK = 200          # slots auto-allocated per school (roomy enough for any team)
BIB_BASE = 100           # first school's block starts here (100–299)


def _alloc_block(conn, district_id):
    """Next free (start, end) block for a new school — sequential, no gaps."""
    top = conn.execute("SELECT COALESCE(MAX(bib_end), 0) FROM schools WHERE district_id=?",
                       (district_id,)).fetchone()[0]
    lo = (top + 1) if top else BIB_BASE
    return lo, lo + BIB_BLOCK - 1


def _next_bib(conn, school):
    """Auto-assign the next bib: fill the school's block, and if it's full, take the
    next free number in a shared overflow zone above every block. Always unique
    across the district, so scans/heat-sheet matching are never ambiguous. Lazily
    allocates a block to a school that doesn't have one yet."""
    row = conn.execute("SELECT bib_start, bib_end, district_id FROM schools WHERE id=?",
                       (school["id"],)).fetchone()
    lo, hi, did = row["bib_start"], row["bib_end"], row["district_id"]
    if not (lo and hi):
        lo, hi = _alloc_block(conn, did)
        conn.execute("UPDATE schools SET bib_start=?, bib_end=? WHERE id=?", (lo, hi, school["id"]))
    used = {r[0] for r in conn.execute(
        "SELECT a.bib FROM athletes a JOIN schools s ON s.id=a.school_id "
        "WHERE s.district_id=? AND a.bib IS NOT NULL", (did,)).fetchall()}
    for b in range(lo, hi + 1):          # tidy: inside the school's own block
        if b not in used:
            return b
    top = conn.execute("SELECT COALESCE(MAX(bib_end), 0) FROM schools WHERE district_id=?",
                       (did,)).fetchone()[0]
    b = top + 1                          # overflow zone above all blocks — still unique
    while b in used:
        b += 1
    return b


# ------------------------------- school list -------------------------------
@bp.get("/schools")
@login_required
def list_schools():
    p = g.principal
    if p.role == "timer" or p.meet_scope:
        abort(403)
    # A coach has one school — "Roster" should open it directly, not a 1-row list.
    if p.role == "coach":
        ids = list(p.school_ids())
        if len(ids) == 1:
            return redirect(f"/schools/{ids[0]}")
    did = active_district_id()
    rows = _visible_schools()
    show_d = p.is_super and did is None

    hdr = ("<tr><th>School</th>" + ("<th>District</th>" if show_d else "")
           + "<th>Athletes</th><th>Bib block</th><th></th></tr>")
    conn = db.connect()
    trs = []
    for s in rows:
        n = conn.execute("SELECT COUNT(*) FROM athletes WHERE school_id=?", (s["id"],)).fetchone()[0]
        bib = f'{s["bib_start"]}&ndash;{s["bib_end"]}' if s["bib_start"] else '<span class="muted">—</span>'
        dcol = f'<td>{escape(s["dname"])}</td>' if show_d else ""
        actions = f'<a class="btn ghost" href="/schools/{s["id"]}">Open roster</a>'
        if p.is_admin:
            actions += (f' <form class="inline" method="post" action="/schools/{s["id"]}/delete" '
                        f'onsubmit="return confirm(\'Delete {escape(s["name"])} and its roster?\')">'
                        f'<button class="danger" type="submit">Delete</button></form>')
        logo = (f'<img src="{escape(s["logo_path"])}" alt="" '
                f'style="height:28px;width:28px;object-fit:contain;vertical-align:middle;margin-right:.5rem">'
                if s["logo_path"] else "")
        trs.append(f'<tr><td>{logo}<b>{escape(s["name"])}</b></td>{dcol}<td>{n}</td>'
                   f'<td>{bib}</td><td style="text-align:right">{actions}</td></tr>')
    conn.close()
    table = (f'<div class="card"><table>{hdr}{"".join(trs)}</table></div>'
             if rows else '<div class="card muted">No schools yet.</div>')

    form = ""
    if p.is_admin:
        if p.is_super and did is None:
            form = '<p class="muted">Pick a district in the header to add a school.</p>'
        else:
            form = """
<div class="card"><h2>Add a school</h2>
<form method="post" action="/schools" enctype="multipart/form-data">
  <label>Name</label><input name="name" required>
  <label>Logo (optional)</label><input name="logo" type="file" accept="image/*">
  <button type="submit" style="margin-top:1rem">Add school</button>
  <span class="muted">Bib numbers are assigned automatically.</span>
</form></div>"""

    heading = "Roster" if p.role == "coach" else "Schools"
    body = f"<h1>{heading}</h1><p class='sub'>Schools, bib blocks, and rosters.</p>{table}{form}"
    return shell(p, body, active="schools",
                 active_district=did, districts=_districts_for_switcher())


@bp.post("/schools")
@role_required("super_admin", "district_admin")
def create_school():
    did = scoped_district_or_403()
    name = (request.form.get("name") or "").strip()
    if not name:
        abort(400)
    logo_path = _save_logo(request.files.get("logo"), name)
    conn = db.connect()
    lo, hi = _alloc_block(conn, did)   # auto-assigned bib block (no manual ranges)
    conn.execute(
        "INSERT INTO schools (district_id, name, bib_start, bib_end, logo_path) VALUES (?,?,?,?,?)",
        (did, name, lo, hi, logo_path),
    )
    conn.commit()
    conn.close()
    return redirect("/schools")


@bp.post("/schools/<int:sid>/edit")
@role_required("super_admin", "district_admin")
def edit_school(sid):
    s = _load_school_or_403(sid)
    # Bib numbers are NOT editable here — they're managed by auto-assignment.
    name = (request.form.get("name") or "").strip() or s["name"]
    logo_path = _save_logo(request.files.get("logo"), name) or s["logo_path"]
    conn = db.connect()
    conn.execute("UPDATE schools SET name=?, logo_path=? WHERE id=?", (name, logo_path, sid))
    conn.commit()
    conn.close()
    return redirect(f"/schools/{sid}")


@bp.post("/schools/<int:sid>/delete")
@role_required("super_admin", "district_admin")
def delete_school(sid):
    s = _load_school_or_403(sid)
    conn = db.connect()
    try:
        # Full FK-safe cascade (outage postmortem: a bare parent DELETE raises a
        # constraint error once the school has competed). Results keep their name
        # snapshots, so past meets stay readable after the school is gone.
        conn.execute("DELETE FROM athlete_waivers WHERE athlete_id IN "
                     "(SELECT id FROM athletes WHERE school_id=?)", (sid,))
        conn.execute("UPDATE entries SET runner_id=NULL WHERE runner_id IN "
                     "(SELECT id FROM athletes WHERE school_id=?)", (sid,))
        conn.execute("UPDATE entries SET school_id=NULL WHERE school_id=?", (sid,))
        conn.execute("UPDATE meets SET host_school_id=NULL WHERE host_school_id=?", (sid,))
        conn.execute("DELETE FROM meet_schools WHERE school_id=?", (sid,))
        conn.execute("DELETE FROM athletes WHERE school_id=?", (sid,))
        conn.execute("DELETE FROM user_schools WHERE school_id=?", (sid,))
        conn.execute("DELETE FROM schools WHERE id=?", (sid,))
        conn.commit()
    finally:
        conn.close()
    return redirect("/schools")


# ------------------------------- roster -------------------------------
@bp.get("/schools/<int:sid>")
@login_required
def roster(sid):
    s = _load_school_or_403(sid)
    sport = request.args.get("sport", "all")
    grad_view = request.args.get("show") == "grad"

    where = ["school_id=?", "active=0" if grad_view else "active=1"]
    params = [sid]
    if sport == "xc":
        where.append("does_xc=1")
    elif sport == "track":
        where.append("does_track=1")
    conn = db.connect()
    ath = conn.execute(
        f"SELECT * FROM athletes WHERE {' AND '.join(where)} "
        f"ORDER BY bib IS NULL, bib, name", tuple(params)).fetchall()
    grad_count = conn.execute("SELECT COUNT(*) FROM athletes WHERE school_id=? AND active=0",
                              (sid,)).fetchone()[0]
    conn.close()

    mode = demo.mode_for(g.principal)
    ro = g.principal.is_demo
    trs = []
    for a in ath:
        nm = demo.display(a["name"], mode)
        if grad_view:
            act = "" if ro else (
                f'<form class="inline" method="post" action="/athletes/{a["id"]}/restore">'
                f'<button class="ghost" type="submit">Restore</button></form>')
            trs.append(
                f'<tr style="opacity:.7"><td>{"" if a["bib"] is None or mode=="anon" else a["bib"]}</td>'
                f'<td><b><a href="/athletes/{a["id"]}/progress">{escape(nm)}</a></b> 📈</td>'
                f'<td>{"" if a["grade"] is None else a["grade"]}</td>'
                f'<td>{a["gender"] or ""}</td>'
                f'<td style="text-align:right">{act}</td></tr>')
            continue
        dis = "disabled" if ro else ""
        xc = "checked" if a["does_xc"] else ""
        tr = "checked" if a["does_track"] else ""
        aid = a["id"]
        acts = [f'<a class="ic" href="/athletes/{aid}/progress" title="Stats &amp; progress">📈</a>',
                f'<button class="ic" onclick="openCard({aid})" title="Info">ⓘ</button>']
        if not ro:
            acts.append(f'<button class="ic edit" onclick="openEdit({aid})" title="Edit name / grade / sex">✏️</button>')
            acts.append(
                f'<form class="inline" method="post" action="/athletes/{aid}/delete" '
                f'onsubmit="return confirm(\'Remove {escape(nm)}?\')">'
                f'<button class="ic del" type="submit" title="Delete">✕</button></form>')
        trs.append(
            f'<tr><td>{"" if a["bib"] is None or mode=="anon" else a["bib"]}</td>'
            f'<td><b><a href="#" onclick="openCard({aid});return false">{escape(nm)}</a></b></td>'
            f'<td>{"" if a["grade"] is None else a["grade"]}</td>'
            f'<td>{a["gender"] or ""}</td>'
            f'<td style="text-align:center"><input type="checkbox" style="width:auto" {xc} {dis} '
            f'onchange="tog({aid},\'xc\',this)"></td>'
            f'<td style="text-align:center"><input type="checkbox" style="width:auto" {tr} {dis} '
            f'onchange="tog({aid},\'track\',this)"></td>'
            f'<td><div class="rowacts">{"".join(acts)}</div></td></tr>')

    if grad_view:
        head = ('<tr><th>Bib</th><th>Name</th><th>Gr</th><th>Sex</th><th></th></tr>'
                if ath else "")
        empty = "No graduated athletes."
    else:
        head = ('<tr><th>Bib</th><th>Name</th><th>Gr</th><th>Sex</th>'
                '<th style="text-align:center">XC</th><th style="text-align:center">Track</th><th></th></tr>'
                if ath else "")
        empty = "No athletes here yet — add or import below."
    table = (f'<div class="card"><table>{head}{"".join(trs)}</table></div>'
             if ath else f'<div class="card muted">{empty}</div>')

    # Sport filter + graduated toggle
    def _f(label, val):
        on = "background:var(--panel2);color:var(--fg)" if sport == val and not grad_view else ""
        return (f'<a class="btn ghost" style="{on}" '
                f'href="/schools/{sid}?sport={val}">{label}</a>')
    grad_link = (f'<a class="btn ghost" href="/schools/{sid}?show=grad" '
                 f'style="{"background:var(--panel2);color:var(--fg)" if grad_view else ""}">'
                 f'🎓 Graduated ({grad_count})</a>' if grad_count or grad_view else "")
    filt = (f'<div class="row" style="margin:.2rem 0 1rem;gap:.4rem">'
            f'{_f("All", "all")}{_f("🏃 XC", "xc")}{_f("🎽 Track", "track")}'
            f'<span style="flex:1"></span>{grad_link}</div>')

    bib_hint = (f'{s["bib_start"]}–{s["bib_end"]}' if s["bib_start"]
                else "no block set (bibs left blank)")
    logo_img = (f'<img src="{escape(s["logo_path"])}" alt="" style="height:42px;width:42px;'
                f'object-fit:contain;vertical-align:middle;margin-right:.6rem;background:#fff;'
                f'border-radius:8px;padding:3px"> ' if s["logo_path"] else "")
    body = f"""
<p class="muted"><a href="/schools">← Schools</a></p>
<h1>{logo_img}{escape(s['name'])}</h1>"""
    body += f"""
<p class="sub">{len(ath)} athletes · bib block {bib_hint}</p>

<div class="row">
  <a class="btn ghost" href="/schools/{sid}/bibs.pdf">Bib list (PDF)</a>
  <a class="btn ghost" href="/schools/{sid}/stickers.pdf?template=5160">Avery 5160 stickers</a>
  <a class="btn ghost" href="/schools/{sid}/stickers.pdf?template=5163">Avery 5163 stickers</a>
</div>
{filt}
{table}
<div id="cardModal" class="cardmodal" style="display:none" onclick="if(event.target===this)closeCard()">
  <div class="cardbox"><button class="cardx" onclick="closeCard()">✕</button>
  <div id="cardBody"></div></div>
</div>
<style>
.cardmodal{{position:fixed;inset:0;background:rgba(0,0,0,.55);display:flex;
  align-items:flex-start;justify-content:center;padding:4vh 1rem;z-index:50;overflow:auto}}
.cardbox{{background:var(--panel);border:1px solid var(--line);border-radius:14px;
  padding:1.4rem;max-width:520px;width:100%;position:relative}}
.cardx{{position:absolute;top:.6rem;right:.6rem;background:transparent;color:var(--mut);
  border:1px solid var(--line);border-radius:8px;padding:.2rem .55rem}}
.crow{{display:flex;justify-content:space-between;gap:1rem;padding:.4rem 0;border-bottom:1px solid var(--line)}}
.crow .k{{color:var(--mut);font-size:.82rem}}
.cbadge{{padding:.12rem .55rem;border-radius:999px;font-size:.75rem}}
.cbadge.ok{{background:rgba(63,191,127,.18);color:var(--ok)}}
.cbadge.no{{background:rgba(240,98,91,.16);color:var(--err)}}
.cbadge.warn{{background:rgba(240,178,75,.16);color:var(--warn)}}
.rowacts{{display:flex;gap:.35rem;justify-content:flex-end;align-items:center}}
.rowacts form{{margin:0}}
.ic{{background:var(--panel2);color:var(--fg);border:1px solid var(--line);border-radius:8px;
  padding:.32rem .52rem;font-size:.9rem;line-height:1;cursor:pointer;text-decoration:none;
  display:inline-flex;align-items:center}}
.ic:hover{{background:var(--line);text-decoration:none}}
.ic.edit{{color:#6bb0f7}}
.ic.del{{color:var(--err)}}
</style>
<script>
async function openCard(aid, edit){{
  const m=document.getElementById('cardModal'), b=document.getElementById('cardBody');
  b.innerHTML='<p class="muted">Loading…</p>'; m.style.display='flex';
  try{{ const r=await fetch('/athletes/'+aid+'/card'); b.innerHTML=await r.text();
    if(edit && document.getElementById('cardEdit')) editInfo(); }}
  catch(e){{ b.innerHTML='<p class="msg err">Could not load.</p>'; }}
}}
function closeCard(){{ document.getElementById('cardModal').style.display='none'; }}
async function openEdit(aid){{
  const m=document.getElementById('cardModal'), b=document.getElementById('cardBody');
  b.innerHTML='<p class="muted">Loading…</p>'; m.style.display='flex';
  try{{ const r=await fetch('/athletes/'+aid+'/editcard'); b.innerHTML=await r.text(); }}
  catch(e){{ b.innerHTML='<p class="msg err">Could not load.</p>'; }}
}}
async function saveCore(aid){{
  const f={{name:document.getElementById('e_name').value,
           grade:document.getElementById('e_grade').value,
           age:document.getElementById('e_age').value,
           gender:document.getElementById('e_gender').value}};
  if(!f.name.trim()){{ alert('Name is required.'); return; }}
  try{{ await jpost('/athletes/'+aid+'/edit', f); location.reload(); }}
  catch(e){{ alert(e.message); }}
}}
function editInfo(){{ document.getElementById('cardView').style.display='none';
  document.getElementById('cardEdit').style.display='block'; }}
async function saveInfo(aid){{
  const f={{}}; ['dob','email','phone','parent_name','parent_email','parent_phone',
    'emergency_name','emergency_phone','physical_date'].forEach(k=>{{
      const el=document.getElementById('f_'+k); if(el) f[k]=el.value; }});
  try{{ await jpost('/athletes/'+aid+'/info', f); openCard(aid); }}
  catch(e){{ alert(e.message); }}
}}
async function sendWaiver(aid){{
  if(!confirm('Email the waiver link to the parent on file?'))return;
  try{{ const j=await jpost('/athletes/'+aid+'/waiver/send',{{}});
    alert(j.sent ? ('Waiver link sent to '+j.to) : ('Link created but email is off. Link: '+j.url));
    openCard(aid); }}
  catch(e){{ alert(e.message); }}
}}
{"" if (ro or grad_view) else '''
async function tog(aid, sport, el){
  try{ await jpost('/athletes/'+aid+'/sports', {sport, on: el.checked}); }
  catch(e){ alert(e.message); el.checked = !el.checked; }
}'''}
</script>"""
    if not ro:
        body += f"""

<div class="card"><h2>Add athlete</h2>
<form method="post" action="/schools/{sid}/athletes">
  <div class="row">
    <div><label>Name</label><input name="name" required></div>
    <div style="max-width:110px"><label>Bib</label><input name="bib" type="number" placeholder="auto"></div>
    <div style="max-width:90px"><label>Grade</label><input name="grade" type="number"></div>
    <div style="max-width:80px"><label>Age</label><input name="age" type="number" placeholder="road"></div>
    <div style="max-width:110px"><label>Sex</label>
      <select name="gender"><option value="">—</option><option>M</option><option>F</option></select></div>
  </div>
  <div style="display:flex;gap:1.2rem;margin-top:.7rem">
    <label style="display:flex;gap:.4rem;align-items:center"><input type="checkbox" name="does_xc"
      value="1" checked style="width:auto"> 🏃 XC</label>
    <label style="display:flex;gap:.4rem;align-items:center"><input type="checkbox" name="does_track"
      value="1" checked style="width:auto"> 🎽 Track</label>
  </div>
  <button type="submit" style="margin-top:1rem">Add</button>
  <span class="muted">Leave bib blank to auto-assign the next free bib in the block.</span>
</form></div>

<div class="card"><h2>Import roster</h2>
<p class="muted">Upload Excel/CSV/PDF/Word, or paste a Google Sheet link. Claude
normalizes names, then you confirm before anything is saved.</p>
<p><a class="btn ghost" href="/roster-template.xlsx">⬇ Download roster template (Excel)</a>
 <button type="button" class="btn ghost"
   onclick="document.getElementById('gfmodal').style.display='flex'">📝 Set up a Google Form sign-up</button></p>
<p class="muted">Fill the template with athletes plus contact, parent &amp; emergency info and upload it back —
 or collect it all with a Google Form (see the guide).</p>

<div id="gfmodal" style="display:none;position:fixed;inset:0;z-index:50;background:rgba(3,10,20,.7);
  align-items:flex-start;justify-content:center;padding:4vh 1rem;overflow:auto"
  onclick="if(event.target===this)this.style.display='none'">
  <div class="card" style="max-width:640px;width:100%">
    <div class="row" style="justify-content:space-between;align-items:center">
      <h2 style="margin:0">📝 Collect sign-ups with a Google Form</h2>
      <button class="ghost" onclick="document.getElementById('gfmodal').style.display='none'">✕</button>
    </div>
    <p class="muted">Build a form once, share the link with parents, and every response lands in a
      Google Sheet you can import here in one click. Name the questions to match these columns.</p>
    <ol style="line-height:1.6;padding-left:1.2rem">
      <li>Go to <b>forms.google.com</b> → <b>Blank form</b>. Title it e.g. "Lehi XC / Track Sign-up".</li>
      <li>Add one question per column (the question title becomes the column header):
        <ul style="margin:.3rem 0;padding-left:1.1rem">
          <li><b>Name</b> — Short answer</li>
          <li><b>Grade</b> — Multiple choice: 6, 7, 8, 9</li>
          <li><b>Gender</b> — Multiple choice: M, F</li>
          <li><b>Cross Country</b> — Multiple choice: Yes, No</li>
          <li><b>Track</b> — Multiple choice: Yes, No</li>
          <li><b>Date of Birth</b> — Date</li>
          <li><b>Parent/Guardian Name</b> — Short answer</li>
          <li><b>Parent Email</b> · <b>Parent Phone</b> — Short answer</li>
          <li><b>Emergency Contact</b> · <b>Emergency Phone</b> — Short answer</li>
        </ul>
      </li>
      <li>Open the <b>Responses</b> tab → <b>Link to Sheets</b> → create a new spreadsheet.</li>
      <li>In that Sheet: <b>Share</b> → <b>Anyone with the link</b> → <b>Viewer</b>, and copy the link.</li>
      <li>Back here, paste it into <b>Google Sheet share link</b> → <b>Pull sheet</b> → review → <b>Import</b>.</li>
    </ol>
    <p class="muted">Tip: you can also open the Sheet → <b>File → Download → .xlsx</b> and upload that file instead.
      Extra Google Form columns (like a timestamp) are ignored automatically.</p>
    <button onclick="document.getElementById('gfmodal').style.display='none'" style="margin-top:.4rem">Got it</button>
  </div>
</div>
<div class="row">
  <div>
    <label>File</label>
    <input type="file" id="impfile" accept=".xlsx,.xls,.csv,.tsv,.txt,.pdf,.docx">
    <button type="button" onclick="uploadImport()" style="margin-top:.6rem">Parse file</button>
  </div>
  <div>
    <label>Google Sheet share link</label>
    <input id="sheeturl" placeholder="https://docs.google.com/spreadsheets/d/..."
      value="{escape(s['sheet_url'] or '')}">
    <button type="button" onclick="sheetImport()" style="margin-top:.6rem">Pull sheet</button>
  </div>
</div>
<div id="preview"></div>
</div>

<script>
let PARSED = [];
function renderPreview(rows){{
  PARSED = rows || [];
  if(!PARSED.length){{ document.getElementById('preview').innerHTML =
    '<p class="msg err">No athletes found.</p>'; return; }}
  const yn=v=> v===true?'✓':(v===false?'–':'');
  let h = '<h2>'+PARSED.length+' found — review &amp; import</h2><table>'
    +'<tr><th>Name</th><th>Gr</th><th>Sex</th><th>XC</th><th>Track</th></tr>';
  for(const a of PARSED) h += '<tr><td>'+esc(a.name)+'</td><td>'+esc(a.grade??'')
    +'</td><td>'+esc(a.gender??'')+'</td><td>'+yn(a.does_xc)+'</td><td>'+yn(a.does_track)+'</td></tr>';
  h += '</table><button type="button" onclick="commitImport()" style="margin-top:1rem">'
    +'Import '+PARSED.length+' athletes</button>';
  document.getElementById('preview').innerHTML = h;
}}
async function uploadImport(){{
  const f = document.getElementById('impfile').files[0];
  if(!f){{ alert('Choose a file'); return; }}
  document.getElementById('preview').innerHTML = '<p class="muted">Parsing…</p>';
  const fd = new FormData(); fd.append('file', f);
  const r = await fetch('/schools/{sid}/import/parse', {{method:'POST', body:fd}});
  const j = await r.json();
  if(!r.ok){{ document.getElementById('preview').innerHTML =
    '<p class="msg err">'+esc(j.error||'Parse failed')+'</p>'; return; }}
  renderPreview(j.athletes);
}}
async function sheetImport(){{
  const url = document.getElementById('sheeturl').value.trim();
  if(!url){{ alert('Paste a link'); return; }}
  document.getElementById('preview').innerHTML = '<p class="muted">Fetching…</p>';
  try{{ const j = await jpost('/schools/{sid}/import/sheet', {{url}}); renderPreview(j.athletes); }}
  catch(e){{ document.getElementById('preview').innerHTML =
    '<p class="msg err">'+esc(e.message)+'</p>'; }}
}}
async function commitImport(){{
  try{{ const j = await jpost('/schools/{sid}/import/commit', {{athletes: PARSED}});
    location.href = '/schools/{sid}'; }}
  catch(e){{ alert(e.message); }}
}}
</script>

<div class="card" style="border:1px solid #b5451f">
<h2>🧹 End of season — clear roster</h2>
<p class="muted">When the season's over, permanently delete this school's athletes and all of
their personal details — contact info, parent &amp; emergency contacts, date of birth,
physical dates, and signed waivers. <b>Meet results and times are kept</b> for records and
history. This can't be undone, so export anything you want to keep first.</p>
<form method="post" action="/schools/{sid}/end-season"
  onsubmit="return confirm('Delete ALL athletes and their personal data for this school? Results are kept. This CANNOT be undone.')">
  <label>Type <b>DELETE</b> to confirm</label>
  <input name="confirm" placeholder="DELETE" autocomplete="off" autocapitalize="characters"
    style="max-width:170px">
  <button class="danger" type="submit" style="margin-top:.6rem">Clear roster &amp; personal data</button>
</form></div>
"""
    if g.principal.is_admin and not ro:
        body += f"""
<div class="card"><h2>Edit school</h2>
<form method="post" action="/schools/{sid}/edit" enctype="multipart/form-data">
  <label>Name</label><input name="name" value="{escape(s['name'])}">
  <label>Replace logo (optional)</label><input name="logo" type="file" accept="image/*">
  <button type="submit" style="margin-top:.8rem">Save changes</button>
</form></div>"""
    return shell(g.principal, body, active="schools", err=request.args.get("err"),
                 active_district=active_district_id(), districts=_districts_for_switcher())


_TEMPLATE_COLS = ["Name", "Grade", "Gender", "Cross Country", "Track", "Date of Birth",
                  "Athlete Email", "Athlete Phone", "Parent/Guardian Name", "Parent Email",
                  "Parent Phone", "Emergency Contact", "Emergency Phone"]
_TEMPLATE_SAMPLES = [
    ["Alex Rivers", 7, "M", "Yes", "Yes", "2013-04-18", "", "", "Jordan Rivers",
     "jordan.rivers@example.com", "555-0142", "Jordan Rivers", "555-0142"],
    ["Sam Brooks", 8, "F", "No", "Yes", "2012-09-05", "sam.brooks@example.com", "",
     "Taylor Brooks", "taylor.brooks@example.com", "555-0199", "Casey Brooks", "555-0177"],
]


def _roster_template_xlsx():
    """A fill-in Excel roster template: header row, two example rows, and an
    Instructions sheet. Columns match what the importer reads back in."""
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(_TEMPLATE_COLS)
    for c in ws[1]:
        c.font = Font(bold=True)
    for row in _TEMPLATE_SAMPLES:
        ws.append(row)
    widths = [20, 7, 8, 13, 8, 13, 24, 14, 22, 26, 14, 22, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    tips = wb.create_sheet("Instructions")
    for i, line in enumerate([
            "XCTimer — roster template",
            "",
            "1. Replace the two example rows on the 'Roster' tab with your athletes.",
            "2. Keep the header row exactly as-is.",
            "3. Name = First Last.  Grade = a number (6, 7, 8, 9).  Gender = M or F.",
            "4. Cross Country / Track = Yes or No (which sports the athlete is doing).",
            "5. Date of Birth = MM/DD/YYYY (or YYYY-MM-DD).",
            "6. Contact / parent / emergency columns are optional — fill what you have.",
            "7. Bibs are assigned automatically on import (you don't enter them here).",
            "8. Save the file, then upload it on the school's Import roster card.",
            "",
            "You can also upload your own spreadsheet — the importer reads matching",
            "column headers — but this template is the easiest way to capture everything."],
            start=1):
        tips[f"A{i}"] = line
    tips["A1"].font = Font(bold=True, size=14)
    tips.column_dimensions["A"].width = 78
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@bp.get("/roster-template.xlsx")
@login_required
def roster_template():
    return Response(_roster_template_xlsx(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": 'attachment; filename="xctimer-roster-template.xlsx"'})


@bp.post("/schools/<int:sid>/athletes")
@login_required
def add_athlete(sid):
    s = _load_school_or_403(sid)
    name = (request.form.get("name") or "").strip()
    if not name:
        abort(400)
    grade = (request.form.get("grade") or "").strip()
    grade = int(grade) if grade.isdigit() else None
    age = (request.form.get("age") or "").strip()
    age = int(age) if age.isdigit() else None
    gender = (request.form.get("gender") or "").strip().upper() or None
    if gender not in ("M", "F", None):
        gender = None
    bib_raw = (request.form.get("bib") or "").strip()
    dx = 1 if request.form.get("does_xc") else 0
    dt = 1 if request.form.get("does_track") else 0
    conn = db.connect()
    bib = int(bib_raw) if bib_raw.isdigit() else _next_bib(conn, s)
    # Manual bib must be unique across the district (stickers/scans key on it).
    dup = conn.execute(
        "SELECT a.name, sc.name AS sname FROM athletes a JOIN schools sc ON sc.id=a.school_id "
        "WHERE a.bib=? AND sc.district_id=? AND a.active=1", (bib, s["district_id"])).fetchone()
    if dup:
        conn.close()
        return redirect(f"/schools/{sid}?err=Bib+{bib}+is+already+assigned+to+"
                        f"{dup['name'].replace(' ', '+')}+({dup['sname'].replace(' ', '+')})")
    conn.execute(
        "INSERT INTO athletes (school_id, bib, name, grade, age, gender, does_xc, does_track) "
        "VALUES (?,?,?,?,?,?,?,?)",
        (sid, bib, name, grade, age, gender, dx, dt),
    )
    conn.commit()
    conn.close()
    return redirect(f"/schools/{sid}")


@bp.post("/athletes/<int:aid>/sports")
@login_required
def athlete_sports(aid):
    """Toggle one sport membership (XC / Track) for an athlete from the roster."""
    conn = db.connect()
    a = conn.execute("SELECT * FROM athletes WHERE id=?", (aid,)).fetchone()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone() if a else None
    conn.close()
    if not a:
        abort(404)
    if not _can_access_school(s) or g.principal.is_demo:
        abort(403)
    data = request.get_json(silent=True) or {}
    col = {"xc": "does_xc", "track": "does_track"}.get(data.get("sport"))
    if not col:
        return jsonify(error="bad sport"), 400
    conn = db.connect()
    conn.execute(f"UPDATE athletes SET {col}=? WHERE id=?", (1 if data.get("on") else 0, aid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/athletes/<int:aid>/restore")
@login_required
def restore_athlete(aid):
    """Un-graduate an athlete (back onto the active roster)."""
    conn = db.connect()
    a = conn.execute("SELECT * FROM athletes WHERE id=?", (aid,)).fetchone()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone() if a else None
    conn.close()
    if not a:
        abort(404)
    if not _can_access_school(s) or g.principal.is_demo:
        abort(403)
    conn = db.connect()
    conn.execute("UPDATE athletes SET active=1 WHERE id=?", (aid,))
    conn.commit()
    conn.close()
    return redirect(f"/schools/{a['school_id']}?show=grad")


@bp.post("/schools/<int:sid>/end-season")
@login_required
def end_season(sid):
    """End-of-season cleanup (data minimization): permanently delete this school's
    athletes and their personal data. Meet results/times are kept — they carry name
    snapshots, so records and history survive. Irreversible; requires typed confirm."""
    s = _load_school_or_403(sid)
    if g.principal.is_demo:
        abort(403)
    if (request.form.get("confirm") or "").strip().upper() != "DELETE":
        return redirect(f"/schools/{sid}?err=Type+DELETE+to+confirm+the+roster+cleanup")
    conn = db.connect()
    ids = [r[0] for r in conn.execute(
        "SELECT id FROM athletes WHERE school_id=?", (sid,)).fetchall()]
    n = len(ids)
    if ids:
        qm = ",".join("?" * len(ids))
        # waiver records (PII) reference athletes -> delete first
        conn.execute(f"DELETE FROM athlete_waivers WHERE athlete_id IN ({qm})", ids)
        # detach meet entries from the athlete so results (with snapshot names) survive
        conn.execute(f"UPDATE entries SET runner_id=NULL WHERE runner_id IN ({qm})", ids)
        conn.execute(f"DELETE FROM athletes WHERE id IN ({qm})", ids)
    conn.commit()
    conn.close()
    return redirect(f"/schools/{sid}?msg=Cleared+{n}+athlete(s)+and+their+personal+data.+Results+kept.")


@bp.post("/athletes/<int:aid>/delete")
@login_required
def delete_athlete(aid):
    conn = db.connect()
    try:
        a = conn.execute("SELECT * FROM athletes WHERE id=?", (aid,)).fetchone()
        if not a:
            abort(404)
        s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone()
        if not _can_access_school(s):
            abort(403)
        # An athlete who has competed is referenced by meet entries + waivers (FKs). A bare
        # DELETE hits a FK constraint and, left unclosed, leaks a WRITE-LOCKED connection —
        # which cascaded into a site-wide "database is locked" outage. Detach entries (their
        # results keep name snapshots) and drop waivers first, then delete. try/finally
        # guarantees the connection is released even if anything above raises.
        conn.execute("DELETE FROM athlete_waivers WHERE athlete_id=?", (aid,))
        conn.execute("UPDATE entries SET runner_id=NULL WHERE runner_id=?", (aid,))
        conn.execute("DELETE FROM athletes WHERE id=?", (aid,))
        conn.commit()
    finally:
        conn.close()
    return redirect(f"/schools/{a['school_id']}")


# ------------------------------- athlete card (popup) -------------------------------
def _phys_badge(pd):
    from datetime import date, datetime
    if not pd:
        return '<span class="cbadge no">none on file</span>', ""
    try:
        d = datetime.strptime(str(pd)[:10], "%Y-%m-%d").date()
    except ValueError:
        return '<span class="cbadge warn">on file</span>', str(pd)
    days = (date.today() - d).days
    cls = "no" if days > 365 else "ok"
    lbl = "expired" if days > 365 else "current"
    return f'<span class="cbadge {cls}">{lbl}</span>', d.isoformat()


def _card_fragment(a, w, ro):
    aid = a["id"]
    pb, pd = _phys_badge(a["physical_date"])
    if w and w["status"] == "signed":
        wb = (f'<span class="cbadge ok">✅ signed</span> '
              f'<span class="muted">{escape(w["signer_name"] or "")} · '
              f'{escape((w["signed_at"] or "")[:10])}</span> '
              f'<a href="/waiver/{w["id"]}/cert.pdf" target="_blank">certificate ↗</a>')
        btn = "" if ro else f'<button class="ghost" onclick="sendWaiver({aid})">Resend</button>'
    elif w and w["status"] == "pending":
        wb = (f'<span class="cbadge warn">⏳ pending</span> '
              f'<span class="muted">sent {escape((w["created_at"] or "")[:10])} '
              f'to {escape(w["sent_to"] or "")}</span>')
        btn = "" if ro else f'<button class="ghost" onclick="sendWaiver({aid})">Resend</button>'
    else:
        wb = '<span class="cbadge no">not sent</span>'
        btn = "" if ro else f'<button onclick="sendWaiver({aid})">Send waiver</button>'

    def row(k, v):
        return (f'<div class="crow"><span class="k">{k}</span>'
                f'<span>{escape(v) if v else "—"}</span></div>')

    meta = (f'{"" if a["grade"] is None else "Grade " + str(a["grade"]) + " · "}'
            f'{a["gender"] or ""}{" · bib " + str(a["bib"]) if a["bib"] else ""}')
    view = f"""
<div id="cardView">
  <h2 style="margin:.1em 0">{escape(a['name'])}</h2>
  <p class="sub" style="margin:.1em 0 1rem">{escape(a['sname'])} · {escape(meta)}
     · <a href="/athletes/{aid}/progress">📈 progress</a></p>
  <div class="crow"><span class="k">Waiver</span><span>{wb} {btn}</span></div>
  <div class="crow"><span class="k">Physical</span><span>{pb}{" · " + escape(pd) if pd else ""}</span></div>
  {row("Date of birth", a["dob"])}
  {row("Athlete email", a["email"])}
  {row("Athlete phone", a["phone"])}
  {row("Parent / guardian", a["parent_name"])}
  {row("Parent email", a["parent_email"])}
  {row("Parent phone", a["parent_phone"])}
  {row("Emergency contact", a["emergency_name"])}
  {row("Emergency phone", a["emergency_phone"])}
  <div class="crow"><span class="k">Medical forms</span>
     <span class="muted">— coming soon</span></div>
  {"" if ro else f'<button style="margin-top:1rem" onclick="editInfo()">✏️ Edit info</button>'}
</div>"""
    if ro:
        return view

    def inp(k, label, typ="text"):
        return (f'<label>{label}</label>'
                f'<input id="f_{k}" type="{typ}" value="{escape(a[k] or "")}">')
    edit = f"""
<div id="cardEdit" style="display:none">
  <h2 style="margin:.1em 0 1rem">Edit — {escape(a['name'])}</h2>
  <label>Date of birth</label>
  <input id="f_dob" type="date" value="{escape((a['dob'] or '')[:10])}">
  <label>Physical date</label>
  <input id="f_physical_date" type="date" value="{escape((a['physical_date'] or '')[:10])}">
  {inp("email", "Athlete email", "email")}
  {inp("phone", "Athlete phone", "tel")}
  {inp("parent_name", "Parent / guardian name")}
  {inp("parent_email", "Parent email", "email")}
  {inp("parent_phone", "Parent phone", "tel")}
  {inp("emergency_name", "Emergency contact name")}
  {inp("emergency_phone", "Emergency phone", "tel")}
  <div style="margin-top:1rem;display:flex;gap:.5rem">
    <button onclick="saveInfo({aid})">Save</button>
    <button class="ghost" onclick="openCard({aid})">Cancel</button>
  </div>
</div>"""
    return view + edit


@bp.get("/athletes/<int:aid>/card")
@login_required
def athlete_card(aid):
    conn = db.connect()
    a = conn.execute("SELECT a.*, s.name AS sname FROM athletes a JOIN schools s ON s.id=a.school_id "
                     "WHERE a.id=?", (aid,)).fetchone()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone() if a else None
    if not a:
        conn.close()
        abort(404)
    if not _can_access_school(s):
        conn.close()
        abort(403)
    w = conn.execute("SELECT * FROM athlete_waivers WHERE athlete_id=? ORDER BY id DESC LIMIT 1",
                     (aid,)).fetchone()
    conn.close()
    return _card_fragment(a, w, g.principal.is_demo)


@bp.get("/athletes/<int:aid>/editcard")
@login_required
def athlete_editcard(aid):
    """Quick edit of the core roster fields — name, grade, sex (the ✏️ button)."""
    conn = db.connect()
    a = conn.execute("SELECT a.*, s.name AS sname FROM athletes a JOIN schools s ON s.id=a.school_id "
                     "WHERE a.id=?", (aid,)).fetchone()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone() if a else None
    conn.close()
    if not a:
        abort(404)
    if not _can_access_school(s) or g.principal.is_demo:
        abort(403)
    grade = "" if a["grade"] is None else a["grade"]
    agev = "" if a["age"] is None else a["age"]
    gsel = lambda v: "selected" if (a["gender"] or "") == v else ""
    return f"""
<h2 style="margin:.1em 0 1rem">Edit athlete</h2>
<label>Name</label>
<input id="e_name" value="{escape(a['name'])}" autofocus>
<div class="row" style="margin-top:.4rem">
  <div style="max-width:120px"><label>Grade</label>
    <input id="e_grade" type="number" inputmode="numeric" value="{grade}"></div>
  <div style="max-width:110px"><label>Age</label>
    <input id="e_age" type="number" inputmode="numeric" value="{agev}"></div>
  <div style="max-width:140px"><label>Sex</label>
    <select id="e_gender">
      <option value="" {gsel("")}>—</option>
      <option value="M" {gsel("M")}>M</option>
      <option value="F" {gsel("F")}>F</option>
    </select></div>
</div>
<div style="margin-top:1rem;display:flex;gap:.5rem">
  <button onclick="saveCore({aid})">Save</button>
  <button class="ghost" onclick="closeCard()">Cancel</button>
</div>
<p class="muted" style="margin-top:.9rem">Contacts, physical &amp; waiver are on the ⓘ Info panel.</p>"""


@bp.post("/athletes/<int:aid>/edit")
@login_required
def athlete_edit(aid):
    conn = db.connect()
    a = conn.execute("SELECT * FROM athletes WHERE id=?", (aid,)).fetchone()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone() if a else None
    conn.close()
    if not a:
        abort(404)
    if not _can_access_school(s) or g.principal.is_demo:
        abort(403)
    d = request.get_json(silent=True) or {}
    name = (str(d.get("name") or "")).strip()
    if not name:
        return jsonify(error="Name is required"), 400
    gr = str(d.get("grade") or "").strip()
    grade = int(gr) if gr.isdigit() else None
    ag = str(d.get("age") or "").strip()
    age = int(ag) if ag.isdigit() else None
    gender = (str(d.get("gender") or "").strip().upper() or None)
    if gender not in ("M", "F", None):
        gender = None
    conn = db.connect()
    conn.execute("UPDATE athletes SET name=?, grade=?, age=?, gender=? WHERE id=?",
                 (name, grade, age, gender, aid))
    # Resync recorded snapshots so a post-race typo fix reaches results/exports:
    # track results link by entries.runner_id; XC finishers match bib + school.
    conn.execute("UPDATE results SET snap_name=? WHERE entry_id IN "
                 "(SELECT id FROM entries WHERE runner_id=?)", (name, aid))
    if a["bib"] is not None:
        conn.execute(
            "UPDATE finishers SET snap_name=?, snap_grade=?, snap_gender=? "
            "WHERE bib=? AND snap_school=?", (name, grade, gender, a["bib"], s["name"]))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/athletes/<int:aid>/info")
@login_required
def athlete_info(aid):
    conn = db.connect()
    a = conn.execute("SELECT * FROM athletes WHERE id=?", (aid,)).fetchone()
    s = conn.execute("SELECT * FROM schools WHERE id=?", (a["school_id"],)).fetchone() if a else None
    conn.close()
    if not a:
        abort(404)
    if not _can_access_school(s) or g.principal.is_demo:
        abort(403)
    d = request.get_json(silent=True) or {}
    fields = ("dob", "email", "phone", "parent_name", "parent_email", "parent_phone",
              "emergency_name", "emergency_phone", "physical_date")
    vals = [(str(d.get(k)).strip() or None) if d.get(k) is not None else None for k in fields]
    conn = db.connect()
    conn.execute(f"UPDATE athletes SET {', '.join(k + '=?' for k in fields)} WHERE id=?",
                 (*vals, aid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ------------------------------- import -------------------------------
@bp.post("/schools/<int:sid>/import/parse")
@login_required
def import_parse(sid):
    _load_school_or_403(sid)
    f = request.files.get("file")
    if not f:
        return jsonify(error="No file uploaded"), 400
    try:
        text = ai.extract_text(f.filename, f.read())
        athletes = ai.normalize_roster(text)
    except Exception as e:  # noqa: BLE001
        return jsonify(error=f"Could not parse: {e}"), 400
    return jsonify(athletes=athletes)


@bp.post("/schools/<int:sid>/import/sheet")
@login_required
def import_sheet(sid):
    _load_school_or_403(sid)
    url = (request.get_json(silent=True) or {}).get("url", "")
    try:
        text = ai.fetch_google_sheet_text(url)
        athletes = ai.normalize_roster(text)
    except Exception as e:  # noqa: BLE001
        return jsonify(error=f"Could not read sheet: {e}. Make sure the sheet is shared "
                             "as 'Anyone with the link — Viewer'."), 400
    # Remember the sheet per school so re-syncs are one click next time.
    conn = db.connect()
    conn.execute("UPDATE schools SET sheet_url=? WHERE id=?", (url.strip(), sid))
    conn.commit()
    conn.close()
    return jsonify(athletes=athletes)


@bp.post("/schools/<int:sid>/import/commit")
@login_required
def import_commit(sid):
    s = _load_school_or_403(sid)
    rows = (request.get_json(silent=True) or {}).get("athletes", [])
    if not isinstance(rows, list):
        return jsonify(error="bad payload"), 400
    added = 0
    conn = db.connect()
    for r in rows:
        name = str((r or {}).get("name", "")).strip()
        if not name:
            continue
        grade = r.get("grade")
        grade = int(grade) if isinstance(grade, int) or (isinstance(grade, str) and grade.isdigit()) else None
        age = r.get("age")
        age = int(age) if isinstance(age, int) or (isinstance(age, str) and age.isdigit()) else None
        gender = r.get("gender")
        gender = gender if gender in ("M", "F") else None
        bib = _next_bib(conn, s)
        cf = {k: (str(r.get(k)).strip() if r.get(k) else None) for k in
              ("dob", "email", "phone", "parent_name", "parent_email", "parent_phone",
               "emergency_name", "emergency_phone")}
        # Sports: honor Cross Country / Track columns; default to BOTH when neither given.
        dx, dt = r.get("does_xc"), r.get("does_track")
        if dx is None and dt is None:
            dx = dt = 1
        else:
            dx, dt = (1 if dx else 0), (1 if dt else 0)
        conn.execute(
            "INSERT INTO athletes (school_id, bib, name, grade, age, gender, does_xc, does_track, "
            "dob, email, phone, parent_name, parent_email, parent_phone, emergency_name, emergency_phone) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            (sid, bib, name, grade, age, gender, dx, dt, cf["dob"], cf["email"], cf["phone"],
             cf["parent_name"], cf["parent_email"], cf["parent_phone"], cf["emergency_name"],
             cf["emergency_phone"]),
        )
        added += 1
    conn.commit()
    conn.close()
    return jsonify(added=added)


# ------------------------------- PDFs -------------------------------
def _roster_rows(sid):
    conn = db.connect()
    rows = conn.execute(
        "SELECT bib, name, grade, gender FROM athletes WHERE school_id=? AND active=1 "
        "ORDER BY bib IS NULL, bib, name", (sid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


@bp.get("/schools/<int:sid>/bibs.pdf")
@login_required
def bibs_pdf(sid):
    s = _load_school_or_403(sid)
    pdf = pdfs.bib_list_pdf(s["name"], _roster_rows(sid))
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{s["name"]}-bibs.pdf"'})


@bp.get("/schools/<int:sid>/stickers.pdf")
@login_required
def stickers_pdf(sid):
    s = _load_school_or_403(sid)
    template = request.args.get("template", "5160")
    ath = _roster_rows(sid)
    # Fill the rest of the last sheet with blank stickers on the next open bibs.
    real = [a for a in ath if a["bib"] is not None]
    pp = pdfs.per_page(template)
    ath += pdfs.blank_fillers([a["bib"] for a in real], s["bib_start"], s["bib_end"],
                              (pp - len(real) % pp) % pp)
    # QR encodes just the bib number (no URL).
    pdf = pdfs.bib_stickers_pdf(s["name"], ath, template=template,
                                qr_prefix="", logo_path=s["logo_path"])
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="{s["name"]}-stickers.pdf"'})


# ------------------------------- bib check -------------------------------
@bp.get("/bibcheck")
@login_required
def bibcheck():
    """Type/scan a bib -> athlete lookup, scoped to the principal's district."""
    p = g.principal
    bib = (request.args.get("bib") or "").strip()
    result = None
    if bib.isdigit():
        did = active_district_id()
        conn = db.connect()
        if p.is_super and did is None:
            row = conn.execute(
                "SELECT a.*, s.name AS sname FROM athletes a JOIN schools s ON s.id=a.school_id "
                "WHERE a.bib=?", (int(bib),)).fetchone()
        else:
            # District-wide for everyone (incl. coaches): the host coach at the timing
            # tent needs to verify VISITING schools' stickers, not just their own.
            did = did if did is not None else p.district_id
            row = conn.execute(
                "SELECT a.*, s.name AS sname FROM athletes a JOIN schools s ON s.id=a.school_id "
                "WHERE a.bib=? AND s.district_id=?", (int(bib), did)).fetchone()
        conn.close()
        result = dict(row) if row else {}
    if request.args.get("format") == "json":
        return jsonify(result or {})

    body = """
<p class="muted"><a href="/meets">← Meets</a></p>
<h1>Bib check</h1>
<p class="sub">Scan a sticker QR or type a bib number. The box clears after each lookup,
so you can scan one after another.</p>
<div class="card">
  <label>Bib number</label>
  <input id="bib" type="number" inputmode="numeric" autofocus autocomplete="off"
    onkeydown="if(event.key==='Enter'){lookup();}">
  <button onclick="lookup()" style="margin-top:1rem">Look up</button>
</div>
<div id="result"></div>
<script>
async function lookup(){
  const el=document.getElementById('bib'); const v=el.value.trim();
  el.value=''; el.focus();                 // clear immediately, ready for the next scan
  if(!v) return;
  let a={};
  try{ a=await (await fetch('/bibcheck?format=json&bib='+encodeURIComponent(v))).json(); }catch(e){}
  const box=document.getElementById('result');
  if(a && a.bib){
    box.innerHTML='<div class="card"><div style="font-size:1.6rem;font-weight:700">#'+a.bib
      +' · '+esc(a.name||'')+'</div><p class="muted">'+esc(a.sname||'')+' · grade '
      +(a.grade||'—')+' · '+(a.gender||'—')+'</p></div>';
  }else{
    box.innerHTML='<div class="msg err">Bib '+esc(v)+' — no athlete with that bib in your scope.</div>';
  }
}
</script>"""
    return shell(p, body, active="meets", active_district=active_district_id(),
                 districts=_districts_for_switcher())

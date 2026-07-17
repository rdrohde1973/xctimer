"""Cross-country timing engine (handoff §8). Reference: ~/xc-district/xc_district.py.

Races + finishers, a tap-to-finish timing console, bib assignment with results
snapshotting, DQ, reorder (times stay in slots), combined results across races by
gender, MileSplit-style team scoring, xlsx export, and a public results page.
"""
import io
import json
import time
from collections import defaultdict
from datetime import datetime, timezone

from markupsafe import escape
from flask import Blueprint, request, redirect, g, abort, jsonify, Response

from . import db, demo
from .auth import login_required
from .tenancy import active_district_id, all_districts
from .ui import shell, BRAND_HTML, CSS, HEAD_EXTRA
from .meets import (load_meet, can_view_meet, can_setup_meet, can_record_meet)

bp = Blueprint("xc", __name__)


# ------------------------------- helpers -------------------------------
def _now():
    return datetime.now(timezone.utc)


def _iso(dt):
    return dt.isoformat()


def _parse(s):
    if not s:
        return None
    dt = datetime.fromisoformat(s)
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def _ms(dt):
    return int(dt.timestamp() * 1000)


def fmt_time(sec):
    if sec is None:
        return ""
    m = int(sec // 60)
    s = sec - 60 * m
    return f"{m}:{s:04.1f}"


def fmt_hms(sec):
    """Console/finisher clock: H:MM:SS.mmm (matches the old app's timer)."""
    if sec is None:
        return ""
    sec = max(0.0, sec)
    h = int(sec // 3600)
    m = int((sec % 3600) // 60)
    s = sec - 3600 * h - 60 * m
    return f"{h}:{m:02d}:{s:06.3f}"


def _race_or_403(rid, check):
    conn = db.connect()
    r = conn.execute("SELECT * FROM races WHERE id=?", (rid,)).fetchone()
    conn.close()
    if not r:
        abort(404)
    m = load_meet(r["meet_id"])
    if not check(m):
        abort(403)
    return r, m


def _athlete_for_bib(conn, meet_id, bib):
    # Per-meet bib numbering: resolve the scanned bib via the meet's bib map.
    from .meets import athlete_by_meet_bib
    return athlete_by_meet_bib(conn, meet_id, bib)


def _participant_for_bib(conn, meet_id, bib):
    return conn.execute(
        "SELECT name, age, gender, club, city FROM participants WHERE meet_id=? AND bib=? LIMIT 1",
        (meet_id, bib)).fetchone()


def _snap_for_bib(conn, m, bib):
    """Snapshot tuple (name, grade, gender, school, age) for a scanned bib.
    Community road events resolve against participants; everything else the roster."""
    if _is_org(m):
        p = _participant_for_bib(conn, m["id"], bib)
        if not p:
            return (None, None, None, None, None)
        return (p["name"], None, p["gender"], p["club"] or p["city"], p["age"])
    a = _athlete_for_bib(conn, m["id"], bib)
    if not a:
        return (None, None, None, None, None)
    return (a["name"], a["grade"], a["gender"], a["sname"], a["age"])


CAPTURE_MODES = [("scan", "Scan at finish"),
                 ("tap", "Tap then scan"),
                 ("tapselect", "Tap then select")]
CAPTURE_VALUES = {v for v, _ in CAPTURE_MODES}


def _is_org(m):
    """True if this meet is a community road event owned by an organizer (no district)."""
    return "organizer_id" in m.keys() and m["organizer_id"] is not None


def _bracket_chips(brackets, muted_when_empty=""):
    """Pill chips for a list of {label,...} brackets; muted note when empty."""
    if not brackets:
        return f'<span class="muted" style="font-size:.85rem">{muted_when_empty}</span>' if muted_when_empty else ''
    chips = "".join(
        f'<span style="display:inline-block;background:#eef3f9;color:#12385f;'
        f'border:1px solid #d5dde6;border-radius:999px;padding:.15rem .6rem;'
        f'margin:.15rem .25rem 0 0;font-size:.82rem;font-weight:600">'
        f'{escape(b.get("label", ""))}</span>' for b in brackets)
    return f'<div style="margin:.2rem 0">{chips}</div>'


def _brackets_to_text(brackets):
    return ", ".join(b.get("label", "") for b in (brackets or []))


def setup_section(m, setup):
    """Setup card(s): XC/track heats table, or the Road events + age-group layout."""
    import json as _json
    conn = db.connect()
    races = conn.execute("SELECT * FROM races WHERE meet_id=? ORDER BY id", (m["id"],)).fetchall()
    counts = {r[0]: r[1] for r in conn.execute(
        "SELECT r.id, COUNT(f.id) FROM races r LEFT JOIN finishers f ON f.race_id=r.id "
        "WHERE r.meet_id=? GROUP BY r.id", (m["id"],)).fetchall()}
    conn.close()

    is_road = m["sport"] == "road"
    rename_js = (
        '<script>function renameHeat(id,cur){var n=prompt("Name",cur);if(!n)return;'
        'var f=document.createElement("form");f.method="post";f.action="/races/"+id+"/rename";'
        'var i=document.createElement("input");i.name="name";i.value=n;f.appendChild(i);'
        'document.body.appendChild(f);f.submit();}</script>')

    if is_road:
        return _road_setup_section(m, setup, races, counts, _json, rename_js)

    # ---- XC / track: classic heats table ----
    rows = []
    for r in races:
        status = "ended" if r["stop_time"] else ("running" if r["start_time"] else "not started")
        act = (f'<a class="btn" href="/races/{r["id"]}/console">⏱ Time</a> '
               f'<a class="btn ghost" href="/races/{r["id"]}/camera">📷 Camera</a>')
        if setup:
            nm = _json.dumps(r["name"])
            act += (
                f" <button class='ghost' onclick='renameHeat({r['id']}, {escape(nm)})'>✎</button>"
                f" <form class='inline' method='post' action='/races/{r['id']}/delete' "
                f"onsubmit=\"return confirm('Delete heat?')\"><button class='danger'>✕</button></form>")
        rows.append(
            f'<tr><td><b>{escape(r["name"])}</b></td><td>{r["capture_mode"]}</td>'
            f'<td>{status}</td><td>{counts.get(r["id"], 0)}</td>'
            f'<td style="text-align:right">{act}</td></tr>')
    tbl = (f'<table><tr><th>Heat</th><th>Mode</th><th>Status</th><th>Finishers</th><th></th></tr>'
           f'{"".join(rows)}</table>' if races else '<p class="muted">No heats yet.</p>')

    ts_toggle = ""
    add = ""
    if setup:
        chk = "checked" if m["team_scoring"] else ""
        ts_toggle = (
            f'<form method="post" action="/meets/{m["id"]}/scoring" style="margin-bottom:.6rem">'
            f'<label style="display:flex;gap:.5rem;align-items:center">'
            f'<input type="checkbox" name="team_scoring" style="width:auto" {chk} onchange="this.form.submit()"> '
            f'<b>Team scoring</b> <span class="muted">— adds team scores (top 5 per school) to results</span></label></form>')
        opts = "".join(f'<option value="{v}">{escape(lbl)}</option>' for v, lbl in CAPTURE_MODES)
        add = (
            f'<form method="post" action="/meets/{m["id"]}/races" class="row" style="margin-top:.8rem">'
            f'<div><input name="name" placeholder="Heat name (e.g. Girls)"></div>'
            f'<div style="max-width:200px"><select name="capture_mode">{opts}</select></div>'
            f'<div style="display:flex;align-items:flex-end"><button type="submit">+ Add heat</button></div>'
            f'</form>{rename_js}')
    return f'<div class="card"><h2>Heats</h2>{ts_toggle}{tbl}{add}</div>'


def _road_setup_section(m, setup, races, counts, _json, rename_js):
    """Road setup: meet-wide default age groups + per-event list with per-event
    age-group override. Athlete→event assignment lives on the dedicated Assign tab."""
    try:
        s = _json.loads(m["settings_json"] or "{}")
    except (ValueError, TypeError):
        s = {}
    default_brackets = s.get("age_brackets") or []
    default_text = s.get("age_brackets_text") or _brackets_to_text(default_brackets)

    org = _is_org(m)
    conn = db.connect()
    ent = conn.execute("SELECT race_id, COUNT(*) AS n FROM race_entries WHERE meet_id=? "
                       "GROUP BY race_id", (m["id"],)).fetchall()
    conn.close()
    assigned_count = {e["race_id"]: e["n"] for e in ent}

    # --- meet-wide default age groups ---
    dflt_editor = ""
    if setup:
        dflt_editor = (
            f'<form method="post" action="/meets/{m["id"]}/age-groups" style="margin-top:.4rem">'
            f'<label class="muted">One group per line or comma-separated. '
            f'Examples: <code>10 &amp; Under, 11-14, 15-19, 20-29, 30-39, 40+</code></label>'
            f'<textarea name="brackets" rows="2" style="width:100%;margin:.3rem 0" '
            f'placeholder="10 &amp; Under, 11-14, 15-19, 20-29, 30+">{escape(default_text)}</textarea>'
            f'<button type="submit">Save default age groups</button></form>')
    default_card = (
        f'<div class="card"><h2>Default age groups</h2>'
        f'<p class="muted" style="margin-top:0">Applied to every event unless the event sets its own. '
        f'Road results place athletes individually by gender × age group — each athlete needs an '
        f'<b>Age</b> on their roster entry.</p>'
        f'{_bracket_chips(default_brackets, "No default set yet.")}{dflt_editor}</div>')

    # --- events ---
    ev_blocks = []
    for r in races:
        status = "ended" if r["stop_time"] else ("running" if r["start_time"] else "not started")
        try:
            own = _json.loads(r["age_brackets"]) if ("age_brackets" in r.keys() and r["age_brackets"]) else None
        except (ValueError, TypeError):
            own = None
        if own:
            groups_line = ('<b>Own age groups:</b> ' + str(_bracket_chips(own)))
        else:
            groups_line = ('<span class="muted">Uses default: </span>'
                           + (str(_bracket_chips(default_brackets)) if default_brackets
                              else '<span class="muted">none set</span>'))
        nassigned = assigned_count.get(r["id"], 0)
        act = (f'<a class="btn" href="/races/{r["id"]}/console">⏱ Time</a> '
               f'<a class="btn ghost" href="/races/{r["id"]}/camera">📷 Camera</a>')
        override = ""
        if setup:
            nm = _json.dumps(r["name"])
            act += (
                f" <button class='ghost' onclick='resetRace({r['id']})'>↺ Reset</button>"
                f" <button class='ghost' onclick='renameHeat({r['id']}, {escape(nm)})'>✎ Rename</button>"
                f" <form class='inline' method='post' action='/races/{r['id']}/delete' "
                f"onsubmit=\"return confirm('Delete this {'race' if org else 'event'} and its results?')\">"
                f"<button class='danger'>✕</button></form>")
            override = (
                f'<details style="margin-top:.5rem"><summary class="muted" style="cursor:pointer">'
                f'Override age groups for this event</summary>'
                f'<form method="post" action="/races/{r["id"]}/age-groups" style="margin-top:.4rem">'
                f'<textarea name="brackets" rows="2" style="width:100%;margin:.3rem 0" '
                f'placeholder="Leave blank to use the meet default">{escape(_brackets_to_text(own))}</textarea>'
                f'<button type="submit">Save</button> '
                f'<span class="muted">Blank = use the default above.</span></form></details>')
        # Community events register participants (no race-entry "assigned" count).
        assigned_txt = "" if org else f" · {nassigned} assigned"
        ev_blocks.append(
            f'<div class="card" style="padding:.8rem 1rem">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;gap:.6rem;flex-wrap:wrap">'
            f'<div><b>{escape(r["name"])}</b> <span class="muted">· {r["capture_mode"]} · {status}'
            f'{assigned_txt} · {counts.get(r["id"], 0)} finishers</span></div>'
            f'<div style="text-align:right">{act}</div></div>'
            f'<div style="margin-top:.3rem">{groups_line}</div>{override}</div>')
    # Community (organizer) events call their sub-races "Races"; district-road calls them "Events".
    noun, noun_s = ("Races", "race") if org else ("Events", "event")
    events_html = "".join(ev_blocks) or f'<p class="muted">No {noun.lower()} yet — add one below.</p>'
    assign_link = (f'<p style="margin:.2rem 0 0"><a class="btn ghost" href="/meets/{m["id"]}/road-assign">'
                   f'🧭 Assign athletes to events →</a></p>' if (setup and races and not org) else '')

    add = ""
    if setup:
        opts = "".join(f'<option value="{v}">{escape(lbl)}</option>' for v, lbl in CAPTURE_MODES)
        add = (
            f'<form method="post" action="/meets/{m["id"]}/races" class="row" style="margin-top:.8rem">'
            f'<div><input name="name" placeholder="{noun_s.capitalize()} name (e.g. 5K, 10K, Fun Run)" required></div>'
            f'<div style="max-width:200px"><select name="capture_mode">{opts}</select></div>'
            f'<div style="display:flex;align-items:flex-end"><button type="submit">+ Add {noun_s}</button></div>'
            f'</form>')
    reset_js = ('<script>async function resetRace(id){'
                'if(!confirm("Reset this race? This clears its clock and ALL recorded finishers."))return;'
                'try{await jpost("/races/"+id+"/reset",{});location.reload();}'
                'catch(e){alert(e.message);}}</script>') if setup else ""
    events_card = f'<div class="card"><h2>{noun}</h2>{events_html}{assign_link}{add}{rename_js}{reset_js}</div>'
    if not org:
        return default_card + events_card

    # Community event: branding + public registration + bib stickers.
    import os as _os
    base = _os.environ.get("XC_PUBLIC_URL", request.host_url.rstrip("/"))
    logo = s.get("event_logo")
    reg_url = f'{base}/register/{m["public_token"]}'
    fee_cents = s.get("fee_cents") or 0
    fee_val = f"{fee_cents / 100:.2f}" if fee_cents else ""
    logo_thumb = (f'<img src="{escape(logo)}" style="max-height:64px;background:#fff;'
                  f'border-radius:8px;padding:4px">' if logo else '<span class="muted">No logo yet</span>')
    event_card = ""
    if setup:
        reg_open = bool(s.get("reg_open"))
        if reg_open:
            import base64 as _b64
            import qrcode as _qr
            _buf = io.BytesIO()
            _qr.make(reg_url).save(_buf, format="PNG")
            _uri = "data:image/png;base64," + _b64.b64encode(_buf.getvalue()).decode()
            link_html = (f'<a href="{escape(reg_url)}" target="_blank">{escape(reg_url)}</a>'
                         f'<br><img src="{_uri}" width="150" height="150" '
                         f'style="background:#fff;padding:8px;border-radius:10px;display:block;margin-top:.4rem">')
        else:
            link_html = '<span class="muted">Turn on “Registration open” to share the link.</span>'
        event_card = (
            '<div class="card"><h2>Event page &amp; registration</h2>'
            f'<div style="display:flex;gap:1rem;align-items:center;flex-wrap:wrap">{logo_thumb}'
            f'<form method="post" action="/meets/{m["id"]}/logo" enctype="multipart/form-data">'
            '<input type="file" name="logo" accept="image/*"> <button type="submit">Upload logo</button>'
            '<div class="muted" style="font-size:.8rem">Shows on the registration page and bib stickers.</div>'
            '</form></div>'
            f'<form method="post" action="/meets/{m["id"]}/event-settings" style="margin-top:.8rem">'
            f'<label style="display:flex;gap:.5rem;align-items:center"><input type="checkbox" name="reg_open" '
            f'style="width:auto" {"checked" if reg_open else ""}> <b>Registration open</b> '
            '<span class="muted">— let runners sign up at the public link</span></label>'
            '<label style="margin-top:.5rem">Registration page text (welcome / instructions)</label>'
            f'<textarea name="reg_text" rows="3" style="width:100%">{escape(s.get("reg_text") or "")}</textarea>'
            '<label style="margin-top:.4rem">Entry fee (optional, $)</label>'
            f'<input name="fee" value="{fee_val}" placeholder="0.00" style="max-width:140px">'
            '<div class="muted" style="font-size:.8rem;margin:.2rem 0">Shown to runners and tracked as '
            'owed; online payment is coming later — collect at packet pickup for now.</div>'
            '<button type="submit" style="margin-top:.4rem">Save</button></form>'
            f'<div style="margin-top:.8rem"><b>Public registration link</b><br>{link_html}</div>'
            '</div>')
        # Bibs / print / camera — separate from registration.
        pc = db.connect()
        nparts = pc.execute("SELECT COUNT(*) FROM participants WHERE meet_id=?", (m["id"],)).fetchone()[0]
        pc.close()
        pnote = (f'<span class="muted">{nparts} participant(s) registered.</span>' if nparts else
                 '<span class="muted">No participants yet — import or register runners on the '
                 f'<a href="/meets/{m["id"]}/participants">Participants</a> tab, then these print with bibs.</span>')
        event_card += (
            '<div class="card"><h2>Bibs, print &amp; camera</h2>'
            f'<p style="margin-top:0">{pnote}</p>'
            f'<div style="margin-top:.4rem"><b>Bibs to print</b> '
            f'<a class="btn" href="/meets/{m["id"]}/participants/tags.pdf">🏁 Bib tags (camera-readable)</a> '
            '<span class="muted">big ArUco tag + number + name, one per runner — '
            '<b>print on matte paper</b> for reliable camera reads.</span></div>'
            f'<div style="margin-top:.8rem"><b>Finish-line camera</b> '
            f'<a class="btn" href="/meets/{m["id"]}/camera">📷 Whole-event camera</a> '
            f'<span class="muted">— one camera for all races; each runner is routed to their own race.</span></div>'
            '</div>')
    return event_card + default_card + events_card


@bp.post("/meets/<int:mid>/age-groups")
@login_required
def save_age_brackets(mid):
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    text = request.form.get("brackets") or ""
    parsed = _parse_brackets(text)
    conn = db.connect()
    row = conn.execute("SELECT settings_json FROM meets WHERE id=?", (mid,)).fetchone()
    try:
        s = json.loads((row["settings_json"] if row else None) or "{}")
    except (ValueError, TypeError):
        s = {}
    s["age_brackets"] = parsed
    s["age_brackets_text"] = text
    conn.execute("UPDATE meets SET settings_json=? WHERE id=?", (json.dumps(s), mid))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}")


@bp.post("/races/<int:rid>/assign")
@login_required
def assign_athlete(rid):
    """Assign/unassign a road athlete to this event. Exactly one event per meet:
    turning an athlete ON clears any prior assignment first."""
    r, m = _race_or_403(rid, can_setup_meet)
    data = request.get_json(silent=True) or {}
    try:
        aid = int(data.get("athlete_id"))
    except (TypeError, ValueError):
        return jsonify(error="bad athlete"), 400
    on = bool(data.get("on"))
    conn = db.connect()
    if on:
        conn.execute("DELETE FROM race_entries WHERE meet_id=? AND athlete_id=?", (m["id"], aid))
        conn.execute("INSERT INTO race_entries (meet_id, race_id, athlete_id) VALUES (?,?,?)",
                     (m["id"], rid, aid))
    else:
        conn.execute("DELETE FROM race_entries WHERE race_id=? AND athlete_id=?", (rid, aid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


def _road_unassigned_warn(conn, m, rid, bib):
    """Warn text if a scanned bib isn't registered/assigned to this distance."""
    if m["sport"] != "road" or bib is None:
        return None
    if _is_org(m):   # community event: participants are born into one distance
        p = conn.execute(
            "SELECT p.race_id, rc.name FROM participants p LEFT JOIN races rc ON rc.id=p.race_id "
            "WHERE p.meet_id=? AND p.bib=?", (m["id"], bib)).fetchone()
        if p is None:
            return f"Bib {bib} isn't registered for this event"
        if p["race_id"] is not None and p["race_id"] != rid:
            return f"Bib {bib} is registered for {p['name']}, not this distance"
        return None
    if not conn.execute("SELECT 1 FROM race_entries WHERE meet_id=? LIMIT 1", (m["id"],)).fetchone():
        return None  # no assignments made for this meet → nothing to enforce
    row = conn.execute(
        "SELECT re.race_id, rc.name FROM race_entries re "
        "JOIN athletes a ON a.id=re.athlete_id "
        "LEFT JOIN races rc ON rc.id=re.race_id "
        "WHERE re.meet_id=? AND a.bib=?", (m["id"], bib)).fetchone()
    if row is None:
        return f"Bib {bib} isn't assigned to any event"
    if row["race_id"] != rid:
        return f"Bib {bib} is assigned to {row['name']}, not this event"
    return None


@bp.post("/races/<int:rid>/age-groups")
@login_required
def save_event_brackets(rid):
    r, m = _race_or_403(rid, can_setup_meet)
    parsed = _parse_brackets(request.form.get("brackets") or "")
    conn = db.connect()
    # Empty = clear the override so the event falls back to the meet default.
    conn.execute("UPDATE races SET age_brackets=? WHERE id=?",
                 (json.dumps(parsed) if parsed else None, rid))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{m['id']}")


def _norm_event(s):
    import re
    return re.sub(r"\s+", "", (s or "").strip().lower())


def _match_event(label, races):
    """Loosely match a roster event label to one of the meet's races.
    Exact (normalized) first; else a substring match either direction. race_id or None."""
    n = _norm_event(label)
    if not n:
        return None
    norms = [(r["id"], _norm_event(r["name"])) for r in races]
    for rid, rn in norms:                       # exact normalized
        if rn == n:
            return rid
    for rid, rn in norms:                        # roster label inside event name ("5k" in "5krun")
        if rn and n in rn:
            return rid
    for rid, rn in norms:                        # event name inside roster label
        if rn and rn in n:
            return rid
    return None


@bp.get("/meets/<int:mid>/road-assign")
@login_required
def road_assign(mid):
    m = load_meet(mid)
    if not can_view_meet(m) or m["sport"] != "road":
        abort(403)
    editable = can_setup_meet(m)
    conn = db.connect()
    races = conn.execute("SELECT * FROM races WHERE meet_id=? ORDER BY id", (mid,)).fetchall()
    roster = conn.execute(
        "SELECT a.id, a.name, a.bib, a.age, a.road_event FROM athletes a "
        "JOIN meet_schools ms ON ms.school_id=a.school_id "
        "WHERE ms.meet_id=? AND a.active=1 AND a.does_road=1 ORDER BY a.name", (mid,)).fetchall()
    ent = {e["athlete_id"]: e["race_id"]
           for e in conn.execute("SELECT athlete_id, race_id FROM race_entries WHERE meet_id=?",
                                 (mid,)).fetchall()}
    conn.close()

    ev_opts = "".join(f'<option value="{r["id"]}">{escape(r["name"])}</option>' for r in races)
    by_event = {r["id"]: [] for r in races}
    unassigned = []
    for a in roster:
        rid = ent.get(a["id"])
        if rid in by_event:
            by_event[rid].append(a)
        else:
            unassigned.append(a)

    def arow(a, assigned_rid=None):
        bib = f'#{a["bib"]}' if a["bib"] is not None else '<span class="muted">no bib</span>'
        age = f'age {a["age"]}' if a["age"] is not None else '<span class="muted">no age</span>'
        rlab = (f' <span class="muted">· roster: {escape(a["road_event"])}</span>'
                if a["road_event"] else "")
        if not editable:
            return f'<div class="arow"><span>{escape(a["name"])}</span><span class="muted">{bib} · {age}</span></div>'
        if assigned_rid is not None:
            btn = (f'<button class="ghost" onclick="unassign({assigned_rid},{a["id"]})">✕ remove</button>')
        else:
            sel = (f'<select onchange="assignTo(this.value,{a["id"]})">'
                   f'<option value="">— assign to —</option>{ev_opts}</select>')
            btn = sel
        return (f'<div class="arow" data-name="{escape((a["name"] or "").lower())}">'
                f'<span><b>{escape(a["name"])}</b> <span class="muted">{bib} · {age}{rlab}</span></span>'
                f'<span>{btn}</span></div>')

    ev_cards = []
    for r in races:
        rows = "".join(arow(a, r["id"]) for a in by_event.get(r["id"], []))
        ev_cards.append(
            f'<div class="card"><h3 style="margin:.1rem 0 .5rem">{escape(r["name"])} '
            f'<span class="muted">— {len(by_event.get(r["id"], []))} assigned</span></h3>'
            f'{rows or "<p class=muted>None yet.</p>"}</div>')
    un_rows = "".join(arow(a) for a in unassigned)
    un_card = (
        f'<div class="card"><h3 style="margin:.1rem 0 .5rem">Unassigned '
        f'<span class="muted">— {len(unassigned)}</span></h3>'
        f'<input id="asearch" placeholder="Search name…" oninput="filt()" '
        f'style="width:100%;margin-bottom:.5rem" {"" if editable else "disabled"}>'
        f'<div id="unlist">{un_rows or "<p class=muted>Everyone is assigned. 🎉</p>"}</div></div>')

    auto = ""
    if editable and races:
        auto = (f'<form method="post" action="/meets/{mid}/road-assign/auto" style="margin:.2rem 0 1rem">'
                f'<button type="submit">⚡ Auto-assign from roster</button> '
                f'<span class="muted">Fills unassigned runners by matching their roster event '
                f'to an event here (loose match).</span></form>')
    msg = request.args.get("msg", "")
    msg_html = f'<div class="card" style="border-color:var(--ok)">{escape(msg)}</div>' if msg else ""

    body = (
        f'<p class="muted"><a href="/meets">← Meets</a></p><h1>{escape(m["name"])} — Assign</h1>'
        f'{_xc_tabs(mid, "assign", road=True)}'
        f'<style>.arow{{display:flex;justify-content:space-between;align-items:center;gap:.6rem;'
        f'padding:.35rem .2rem;border-bottom:1px solid var(--line);flex-wrap:wrap}}'
        f'.arow:last-child{{border-bottom:0}}.arow select,.arow button{{font-size:.9rem}}</style>'
        f'{msg_html}{auto}'
        f'<p class="muted">Each runner does exactly one event — assigning moves them off any other.</p>'
        f'{"".join(ev_cards)}{un_card}'
        '<script>'
        'async function assignTo(rid, aid){ if(!rid) return;'
        '  try{ await jpost("/races/"+rid+"/assign", {athlete_id:aid, on:true}); location.reload(); }'
        '  catch(e){ alert(e.message); } }'
        'async function unassign(rid, aid){'
        '  try{ await jpost("/races/"+rid+"/assign", {athlete_id:aid, on:false}); location.reload(); }'
        '  catch(e){ alert(e.message); } }'
        'function filt(){ var q=document.getElementById("asearch").value.toLowerCase();'
        '  document.querySelectorAll("#unlist .arow").forEach(function(r){'
        '    r.style.display=(!q || (r.getAttribute("data-name")||"").indexOf(q)>=0)?"":"none"; }); }'
        '</script>')
    return shell(g.principal, body, active="meets")


@bp.post("/meets/<int:mid>/road-assign/auto")
@login_required
def road_assign_auto(mid):
    m = load_meet(mid)
    if not can_setup_meet(m) or m["sport"] != "road":
        abort(403)
    conn = db.connect()
    races = conn.execute("SELECT * FROM races WHERE meet_id=? ORDER BY id", (mid,)).fetchall()
    already = {e["athlete_id"] for e in
               conn.execute("SELECT athlete_id FROM race_entries WHERE meet_id=?", (mid,)).fetchall()}
    roster = conn.execute(
        "SELECT a.id, a.road_event FROM athletes a "
        "JOIN meet_schools ms ON ms.school_id=a.school_id "
        "WHERE ms.meet_id=? AND a.active=1 AND a.does_road=1", (mid,)).fetchall()
    assigned, unmatched = 0, 0
    for a in roster:
        if a["id"] in already:
            continue  # never override a manual assignment
        if not a["road_event"]:
            continue
        rid = _match_event(a["road_event"], races)
        if rid is None:
            unmatched += 1
            continue
        conn.execute("INSERT OR IGNORE INTO race_entries (meet_id, race_id, athlete_id) VALUES (?,?,?)",
                     (mid, rid, a["id"]))
        assigned += 1
    conn.commit()
    conn.close()
    msg = f"Auto-assigned {assigned} runner(s)."
    if unmatched:
        msg += f" {unmatched} had a roster event with no matching event here."
    return redirect(f"/meets/{mid}/road-assign?msg={msg.replace(' ', '+')}")


# ------------------------------- races -------------------------------
@bp.post("/meets/<int:mid>/races")
@login_required
def create_race(mid):
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    name = (request.form.get("name") or "").strip() or "Heat"
    mode = request.form.get("capture_mode")
    mode = mode if mode in CAPTURE_VALUES else "tap"
    conn = db.connect()
    conn.execute("INSERT INTO races (meet_id, name, capture_mode) VALUES (?,?,?)",
                 (mid, name, mode))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}")


@bp.post("/races/<int:rid>/rename")
@login_required
def rename_race(rid):
    r, m = _race_or_403(rid, can_setup_meet)
    name = (request.form.get("name") or "").strip()
    mode = request.form.get("capture_mode")
    conn = db.connect()
    if name:
        conn.execute("UPDATE races SET name=? WHERE id=?", (name, rid))
    if mode in CAPTURE_VALUES:
        conn.execute("UPDATE races SET capture_mode=? WHERE id=?", (mode, rid))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{m['id']}")


@bp.post("/races/<int:rid>/delete")
@login_required
def delete_race(rid):
    r, m = _race_or_403(rid, can_setup_meet)
    conn = db.connect()
    conn.execute("DELETE FROM finishers WHERE race_id=?", (rid,))
    conn.execute("DELETE FROM race_entries WHERE race_id=?", (rid,))
    conn.execute("UPDATE participants SET race_id=NULL WHERE race_id=?", (rid,))  # keep the registrant
    conn.execute("DELETE FROM races WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{m['id']}")


# ------------------------------- meet-day tabs -------------------------------
def _xc_tabs(mid, active, road=False, organizer=False):
    """Tab bar for XC/road meets. XC: Setup · Meet day · Results. District-road adds
    an Assign tab; community (organizer) events get a Participants tab instead."""
    def tab(href, label, key):
        on = "background:var(--panel2);color:var(--fg)" if active == key else "color:var(--mut)"
        return (f'<a href="{href}" style="padding:.4rem .9rem;border-radius:8px;'
                f'text-decoration:none;{on}">{label}</a>')
    if organizer:
        mid_tab = tab(f"/meets/{mid}/participants", "👥 Participants", "participants")
    elif road:
        mid_tab = tab(f"/meets/{mid}/road-assign", "🧭 Assign", "assign")
    else:
        mid_tab = ""
    return ('<div style="display:flex;gap:.3rem;margin:.4rem 0 1rem;border-bottom:1px solid var(--line);'
            'padding-bottom:.5rem;flex-wrap:wrap">'
            + tab(f"/meets/{mid}", "⚙️ Setup", "setup")
            + mid_tab
            + tab(f"/meets/{mid}/xc-day", "🏁 Meet day", "meetday")
            + tab(f"/meets/{mid}/results", "📊 Results", "results")
            + '</div>')


@bp.get("/meets/<int:mid>/xc-day")
@login_required
def xc_meet_day(mid):
    """Day-of view: run each heat's timing console + print stickers / bib lists."""
    m = load_meet(mid)
    if not can_view_meet(m) or m["sport"] not in ("xc", "road"):
        abort(403)
    conn = db.connect()
    races = conn.execute("SELECT * FROM races WHERE meet_id=? ORDER BY id", (mid,)).fetchall()
    counts = {r[0]: r[1] for r in conn.execute(
        "SELECT r.id, COUNT(f.id) FROM races r LEFT JOIN finishers f ON f.race_id=r.id "
        "WHERE r.meet_id=? GROUP BY r.id", (mid,)).fetchall()}
    conn.close()
    rows = []
    for r in races:
        status = "ended" if r["stop_time"] else ("running" if r["start_time"] else "not started")
        rows.append(
            f'<tr><td><b>{escape(r["name"])}</b></td><td>{r["capture_mode"]}</td>'
            f'<td>{status}</td><td>{counts.get(r["id"], 0)}</td>'
            f'<td style="text-align:right"><a class="btn" href="/races/{r["id"]}/console">⏱ Time</a> '
            f'<a class="btn ghost" href="/races/{r["id"]}/camera">📷 Camera</a></td></tr>')
    noun = "Races" if _is_org(m) else ("Events" if m["sport"] == "road" else "Heats")
    tbl = (f'<div class="card"><h2>{noun} — tap to time</h2><table><tr><th>{noun[:-1]}</th><th>Mode</th>'
           f'<th>Status</th><th>Finishers</th><th></th></tr>{"".join(rows)}</table></div>'
           if races else f'<div class="card muted">No {noun.lower()} yet — add them on the Setup tab.</div>')
    print_bar = (
        f'<div class="card"><b>Print:</b> '
        f'<a class="btn ghost" href="/meets/{mid}/stickers.pdf?template=5160">Stickers 5160</a> '
        f'<a class="btn ghost" href="/meets/{mid}/stickers.pdf?template=5163">Stickers 5163</a> '
        f'<a class="btn ghost" href="/meets/{mid}/biblist.pdf">Bib lists</a></div>')
    body = (f'<p class="muted"><a href="/meets">← Meets</a></p><h1>{escape(m["name"])}</h1>'
            f'{_xc_tabs(mid, "meetday", road=(m["sport"]=="road"), organizer=_is_org(m))}{tbl}{print_bar}')
    return shell(g.principal, body, active="meets")


# ------------------------------- timing console -------------------------------
CONSOLE_CSS = """
.tc-clock{font-size:3.2rem;font-weight:800;font-variant-numeric:tabular-nums;
  text-align:center;letter-spacing:-.02em;line-height:1}
.tc-clock.stopped{color:var(--err)}
.tc-btns{display:flex;gap:.6rem;justify-content:center;align-items:center;flex-wrap:wrap;margin:1rem 0 .2rem}
.tc-btns button,.tc-btns .btn{font-size:1.05rem;padding:.7rem 1.3rem}
.tc-btns #btn-start{background:var(--ok);color:#04101f;font-size:1.4rem;padding:1rem 2.6rem}
.tc-btns #btn-start:hover{background:#34a86e}
.tc-btns #btn-start:disabled{opacity:.4;cursor:default}
.tc-btns #btn-stop{background:var(--err);color:#fff}
.tc-btns #btn-stop:hover{background:#d9534f}
.tc-btns #btn-stop:disabled{opacity:.4;cursor:default}
.tc-status{text-align:center;font-weight:600;padding:.5rem;border-radius:8px;margin-top:.5rem}
.tc-status.wait{color:var(--warn)}
.tc-status.run{color:var(--ok)}
.tc-status.end{color:var(--err);background:rgba(240,98,91,.12)}
.tc-bibrow{display:flex;gap:.6rem;align-items:center;flex-wrap:wrap}
.tc-bibrow input{max-width:170px;font-size:1.1rem}
.tc-bibrow #slot{margin-left:.2rem}
#rows td{vertical-align:middle}
#rows .grip{cursor:grab;color:var(--dim);text-align:center;font-size:1.1rem;width:1.4rem}
#rows tr.drag-over td{border-top:2px solid var(--acc)}
"""


@bp.get("/races/<int:rid>/console")
@login_required
def console(rid):
    r, m = _race_or_403(rid, can_record_meet)
    body = f"""
<style>{CONSOLE_CSS}</style>
<p class="muted"><a href="/meets/{m['id']}">← {escape(m['name'])}</a></p>
<h1>{escape(r['name'])} <span class="muted" style="font-weight:400">· {escape(m['name'])}</span></h1>
<div class="card">
  <div id="clock" class="tc-clock">0:00:00.000</div>
  <div class="tc-btns">
    <button id="btn-start" onclick="startRace()">🚦 Start</button>
    <button id="btn-stop" onclick="stopRace()">⏹ Stop</button>
    <button class="ghost" onclick="resetRace()">🔄 Reset</button>
  </div>
  <div id="status" class="tc-status wait">Not started.</div>
</div>
<div class="card">
  <div class="tc-bibrow">
    <strong>Bib #</strong>
    <input id="bib" placeholder="scan/type" autocomplete="off" autocapitalize="off"
      onkeydown="if(event.key==='Enter')recordBib()">
    <button onclick="recordBib()">✔ <span id="verb">Assign</span></button>
    <span id="slot" class="muted"></span>
  </div>
  <div id="help" class="muted" style="margin-top:.5rem"></div>
</div>
<div class="card"><h2>Finishers (<span id="count">0</span>)</h2>
  <table><thead><tr><th></th><th>#</th><th>Bib</th><th>Name</th><th>School</th><th>Time</th><th></th></tr></thead>
  <tbody id="rows"></tbody></table>
</div>
<script>
const RID={rid}, MID={m['id']};
let OFFSET=0, START=null, STOPMS=null, STOPPED=false, STARTED=false, MODE='tap', FIN=[], OPEN=0;
let dragId=null, dragging=false;
function nowms(){{ return Date.now()+OFFSET; }}
function fmt(sec){{ if(sec==null)return''; sec=Math.max(0,sec);
  const h=Math.floor(sec/3600), m=Math.floor((sec%3600)/60), s=sec-3600*h-60*m;
  return h+':'+String(m).padStart(2,'0')+':'+s.toFixed(3).padStart(6,'0'); }}
async function load(){{
  const s=await jget('/races/'+RID+'/state');
  OFFSET=s.server_ms-Date.now(); START=s.start_ms; STOPMS=s.stop_ms;
  STOPPED=s.stopped; STARTED=s.started; MODE=s.capture_mode; FIN=s.finishers; OPEN=s.open;
  syncUI(); render();
}}
function syncUI(){{
  const scan = MODE==='scan';
  document.getElementById('btn-start').disabled = STARTED && !STOPPED;
  document.getElementById('btn-stop').disabled = !STARTED || STOPPED;
  document.getElementById('verb').textContent = scan?'Record':'Assign';
  document.getElementById('help').textContent = scan
    ? 'Scan mode: each bib records with the current race time.'
    : 'Tap mode: taps come from the phone; scan bibs here to fill open slots in order.';
  const st=document.getElementById('status');
  if(!STARTED){{ st.className='tc-status wait'; st.textContent='Not started.'; }}
  else if(STOPPED){{ st.className='tc-status end';
    st.textContent = scan?'🏁 Race ended.':'🏁 Race ended — keep scanning bibs to fill open slots.'; }}
  else {{ st.className='tc-status run'; st.textContent='🟢 Running.'; }}
  const slot=document.getElementById('slot');
  if(scan){{ slot.textContent=''; }}
  else {{ const nx=FIN.find(f=>f.bib==null);
    slot.textContent = OPEN? (OPEN+' open — next #'+nx.seq+' @ '+fmt(nx.elapsed)) : 'no open slots'; }}
}}
function tick(){{
  const c=document.getElementById('clock');
  if(!START){{ c.textContent='0:00:00.000'; c.classList.remove('stopped'); return; }}
  const end=(STOPPED&&STOPMS)?STOPMS:nowms(); let e=(end-START)/1000; if(e<0)e=0;
  c.textContent=fmt(e); c.classList.toggle('stopped',STOPPED);
}}
function render(){{
  document.getElementById('count').textContent=FIN.length;
  if(!FIN.length){{ document.getElementById('rows').innerHTML=
    '<tr><td colspan=7 class="muted">No finishers yet.</td></tr>'; return; }}
  let h='';
  FIN.forEach(f=>{{
    const nm = f.name? esc(f.name) : (f.bib?'':'<span class=muted>—</span>');
    const name = f.dq? ('<s>'+nm+' (DQ)</s>') : nm;
    h+='<tr draggable="true" data-id="'+f.id+'" ondragstart="dstart(event,'+f.id+')"'
     +' ondragover="dover(event,this)" ondragleave="this.classList.remove(\\'drag-over\\')"'
     +' ondrop="ddrop(event,'+f.id+')" ondragend="dend()">'
     +'<td class="grip">⠿</td><td>'+f.seq+'</td>'
     +'<td><input value="'+(f.bib??'')+'" style="width:64px" onchange="setBib('+f.id+',this.value)"></td>'
     +'<td>'+name+'</td><td>'+esc(f.school||'')+'</td>'
     +'<td style="font-variant-numeric:tabular-nums">'+fmt(f.elapsed)+'</td>'
     +'<td style="text-align:right;white-space:nowrap">'
     +'<button class="ghost" onclick="dq('+f.id+')">'+(f.dq?'un-DQ':'DQ')+'</button> '
     +'<button class="danger" onclick="del('+f.id+')">✕</button></td></tr>';
  }});
  document.getElementById('rows').innerHTML=h;
}}
async function startRace(){{
  const body={{}};
  if(STOPPED&&FIN.length){{ if(!confirm('Race ended with '+FIN.length+' finisher(s). Restarting CLEARS them. Continue?'))return; body.clear=true; }}
  try{{ await jpost('/races/'+RID+'/start',body); }}catch(e){{ alert(e.message); }} load(); }}
async function stopRace(){{ if(!confirm('Stop the race clock?'))return; await jpost('/races/'+RID+'/stop',{{}}); load(); }}
async function resetRace(){{ if(!confirm('Reset clears the clock and all finishers. Continue?'))return;
  await jpost('/races/'+RID+'/reset',{{}}); load(); }}
async function recordBib(){{ const el=document.getElementById('bib'); const v=el.value.trim(); if(!v)return;
  try{{ const j=await jpost('/races/'+RID+'/finish',{{bib:v}}); if(j&&j.warn) alert('⚠ '+j.warn); el.value=''; el.focus(); load(); }}
  catch(e){{ alert(e.message); el.select(); }} }}
async function setBib(id,v){{ try{{ const j=await jpost('/finishers/'+id+'/bib',{{bib:v}}); if(j&&j.warn) alert('⚠ '+j.warn); }}catch(e){{ alert(e.message); }} load(); }}
async function dq(id){{ await jpost('/finishers/'+id+'/dq',{{}}); load(); }}
async function del(id){{ if(!confirm('Delete finisher?'))return; await jpost('/finishers/'+id+'/delete',{{}}); load(); }}
function dstart(e,id){{ dragging=true; dragId=id; e.dataTransfer.effectAllowed='move'; }}
function dover(e,tr){{ e.preventDefault(); tr.classList.add('drag-over'); }}
async function ddrop(e,overId){{ e.preventDefault();
  document.querySelectorAll('#rows tr').forEach(t=>t.classList.remove('drag-over'));
  if(dragId==null||dragId===overId){{ dend(); return; }}
  const order=FIN.map(f=>f.id); const from=order.indexOf(dragId), to=order.indexOf(overId);
  order.splice(to,0,order.splice(from,1)[0]);
  dend(); await jpost('/races/'+RID+'/reorder',{{order}}); load(); }}
function dend(){{ dragging=false; dragId=null; }}
setInterval(tick,75);
setInterval(()=>{{ if(!dragging)load(); }},2000);
load();
</script>
"""
    return shell(g.principal, body, active="meets")


@bp.get("/races/<int:rid>/eligible")
@login_required
def race_eligible(rid):
    """Runners who can still be selected for this race (tap-then-select mode):
    registered/rostered and not yet recorded with a bib in this race."""
    r, m = _race_or_403(rid, can_record_meet)
    conn = db.connect()
    used = {row[0] for row in conn.execute(
        "SELECT bib FROM finishers WHERE race_id=? AND bib IS NOT NULL", (rid,)).fetchall()}
    out = []
    if _is_org(m):
        rows = conn.execute(
            "SELECT bib, name, age, gender, club FROM participants "
            "WHERE meet_id=? AND (race_id=? OR race_id IS NULL) ORDER BY name", (m["id"], rid)).fetchall()
        for p in rows:
            if p["bib"] is None or p["bib"] in used:
                continue
            meta = " · ".join(x for x in (
                (f'age {p["age"]}' if p["age"] is not None else ""), (p["gender"] or ""),
                (p["club"] or "")) if x)
            out.append({"bib": p["bib"], "name": p["name"], "meta": meta})
    else:
        rows = conn.execute(
            "SELECT a.bib, a.name, a.grade, s.name AS sname FROM athletes a "
            "JOIN schools s ON s.id=a.school_id JOIN meet_schools ms ON ms.school_id=a.school_id "
            "WHERE ms.meet_id=? AND a.bib IS NOT NULL AND a.active=1 ORDER BY a.name", (m["id"],)).fetchall()
        for a in rows:
            if a["bib"] in used:
                continue
            meta = " · ".join(x for x in (
                (f'gr {a["grade"]}' if a["grade"] is not None else ""), (a["sname"] or "")) if x)
            out.append({"bib": a["bib"], "name": a["name"], "meta": meta})
    conn.close()
    return jsonify(runners=out)


@bp.get("/races/<int:rid>/state")
@login_required
def race_state(rid):
    r, m = _race_or_403(rid, can_record_meet)
    conn = db.connect()
    rows = conn.execute("SELECT * FROM finishers WHERE race_id=? ORDER BY seq", (rid,)).fetchall()
    conn.close()
    start = _parse(r["start_time"])
    stop = _parse(r["stop_time"])
    fins = [{
        "id": f["id"], "seq": f["seq"], "bib": f["bib"],
        "elapsed": f["elapsed_seconds"], "elapsed_str": fmt_hms(f["elapsed_seconds"]),
        "dq": bool(f["dq"]),
        "name": f["snap_name"], "school": f["snap_school"],
    } for f in rows]
    return jsonify(
        name=r["name"],
        capture_mode=r["capture_mode"],
        start_ms=_ms(start) if start else None,
        stop_ms=_ms(stop) if stop else None,
        started=bool(r["start_time"]),
        stopped=bool(r["stop_time"]),
        open=sum(1 for f in rows if f["bib"] is None),
        server_ms=_ms(_now()),
        finishers=fins,
    )


@bp.post("/races/<int:rid>/start")
@login_required
def race_start(rid):
    r, m = _race_or_403(rid, can_record_meet)
    conn = db.connect()
    # Restarting an ended race with finishers recorded is a data-loss trap: their
    # elapsed times belong to the old clock. Require an explicit clear (confirmed
    # client-side) before we hand out a fresh start time.
    if r["start_time"] and r["stop_time"]:
        n = conn.execute("SELECT COUNT(*) FROM finishers WHERE race_id=?", (rid,)).fetchone()[0]
        clear = bool((request.get_json(silent=True) or {}).get("clear"))
        if n and not clear:
            conn.close()
            return jsonify(error=f"Race ended with {n} finisher(s). Restarting clears them — "
                                 "confirm the restart (or use Reset).", needs_clear=True), 409
        if n:
            conn.execute("DELETE FROM finishers WHERE race_id=?", (rid,))
    conn.execute("UPDATE races SET start_time=?, stop_time=NULL WHERE id=?", (_iso(_now()), rid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/races/<int:rid>/stop")
@login_required
def race_stop(rid):
    r, m = _race_or_403(rid, can_record_meet)
    conn = db.connect()
    conn.execute("UPDATE races SET stop_time=? WHERE id=?", (_iso(_now()), rid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/races/<int:rid>/reset")
@login_required
def race_reset(rid):
    """Clear the clock and every finisher — a clean slate for the heat."""
    r, m = _race_or_403(rid, can_record_meet)
    conn = db.connect()
    conn.execute("DELETE FROM finishers WHERE race_id=?", (rid,))
    conn.execute("UPDATE races SET start_time=NULL, stop_time=NULL WHERE id=?", (rid,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/races/<int:rid>/untap")
@login_required
def race_untap(rid):
    """Undo the last tap. Won't drop a slot that already has a bib (tap mode)."""
    r, m = _race_or_403(rid, can_record_meet)
    conn = db.connect()
    last = conn.execute("SELECT * FROM finishers WHERE race_id=? ORDER BY seq DESC LIMIT 1",
                        (rid,)).fetchone()
    if not last:
        conn.close()
        return jsonify(ok=True, count=0)
    if last["bib"] is not None and r["capture_mode"] != "scan":
        conn.close()
        return jsonify(error="Last finisher already has a bib — remove it instead"), 400
    conn.execute("DELETE FROM finishers WHERE id=?", (last["id"],))
    conn.commit()
    cnt = conn.execute("SELECT COUNT(*) FROM finishers WHERE race_id=?", (rid,)).fetchone()[0]
    conn.close()
    return jsonify(ok=True, count=cnt)


@bp.post("/races/<int:rid>/finish")
@login_required
def race_finish(rid):
    """Assign a bib. Tap mode: fill the next open (bib-less) slot in order — works
    even after Stop. Scan mode: record a new finisher with the current race time."""
    r, m = _race_or_403(rid, can_record_meet)
    raw = (request.get_json(silent=True) or {}).get("bib")
    try:
        bib = int(str(raw).strip())
    except (TypeError, ValueError):
        return jsonify(error="Bib must be a number"), 400
    if bib <= 0:
        return jsonify(error="Enter a bib number"), 400
    conn = db.connect()
    dup = conn.execute("SELECT snap_name FROM finishers WHERE race_id=? AND bib=?",
                       (rid, bib)).fetchone()
    if dup:
        # Accidental double-scan at the line: silently discard, don't block the timer.
        conn.close()
        return jsonify(ok=True, duplicate=True)
    snap = _snap_for_bib(conn, m, bib)
    if r["capture_mode"] == "scan":
        start = _parse(r["start_time"])
        if not start or r["stop_time"]:
            conn.close()
            return jsonify(error="Race not running"), 400
        elapsed = (_now() - start).total_seconds()
        seq = (conn.execute("SELECT COALESCE(MAX(seq),0) FROM finishers WHERE race_id=?",
                            (rid,)).fetchone()[0]) + 1
        conn.execute(
            "INSERT INTO finishers (race_id, seq, finish_time, elapsed_seconds, bib, "
            "snap_name, snap_grade, snap_gender, snap_school, snap_age) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (rid, seq, _iso(_now()), elapsed, bib, *snap))
        remaining = 0
    else:
        slot = conn.execute("SELECT id FROM finishers WHERE race_id=? AND bib IS NULL "
                            "ORDER BY seq LIMIT 1", (rid,)).fetchone()
        if not slot:
            conn.close()
            return jsonify(error="No open slots — tap a finisher first"), 400
        conn.execute("UPDATE finishers SET bib=?, snap_name=?, snap_grade=?, snap_gender=?, "
                     "snap_school=?, snap_age=? WHERE id=?", (bib, *snap, slot["id"]))
        remaining = conn.execute("SELECT COUNT(*) FROM finishers WHERE race_id=? AND bib IS NULL",
                                 (rid,)).fetchone()[0]
    warn = _road_unassigned_warn(conn, m, rid, bib)
    conn.commit()
    conn.close()
    return jsonify(ok=True, bib=bib, name=snap[0], school=snap[3], remaining=remaining, warn=warn)


@bp.post("/races/<int:rid>/tap")
@login_required
def race_tap(rid):
    r, m = _race_or_403(rid, can_record_meet)
    start = _parse(r["start_time"])
    if not start:
        return jsonify(error="Race not started"), 400
    if r["stop_time"]:
        return jsonify(error="Race has ended — Reset to run again"), 400
    elapsed = (_now() - start).total_seconds()
    conn = db.connect()
    seq = (conn.execute("SELECT COALESCE(MAX(seq),0) FROM finishers WHERE race_id=?",
                        (rid,)).fetchone()[0]) + 1
    cur = conn.execute(
        "INSERT INTO finishers (race_id, seq, finish_time, elapsed_seconds) VALUES (?,?,?,?)",
        (rid, seq, _iso(_now()), elapsed))
    fid = cur.lastrowid
    conn.commit()
    conn.close()
    return jsonify(id=fid, seq=seq, bib=None, elapsed=elapsed, dq=False, name=None, school=None)


@bp.post("/finishers/<int:fid>/bib")
@login_required
def finisher_bib(fid):
    conn = db.connect()
    f = conn.execute("SELECT f.*, r.meet_id FROM finishers f JOIN races r ON r.id=f.race_id "
                     "WHERE f.id=?", (fid,)).fetchone()
    conn.close()
    if not f:
        abort(404)
    _, m = _race_or_403(f["race_id"], can_record_meet)
    raw = (request.get_json(silent=True) or {}).get("bib")
    conn = db.connect()
    warn = None
    if raw in (None, "", "0"):
        conn.execute("UPDATE finishers SET bib=NULL, snap_name=NULL, snap_grade=NULL, "
                     "snap_gender=NULL, snap_school=NULL, snap_age=NULL WHERE id=?", (fid,))
    else:
        try:
            bib = int(str(raw).strip())
        except (TypeError, ValueError):
            conn.close()
            return jsonify(error="Bib must be a number"), 400
        dup = conn.execute("SELECT seq, snap_name FROM finishers WHERE race_id=? AND bib=? AND id!=?",
                           (f["race_id"], bib, fid)).fetchone()
        if dup:
            conn.close()
            nm = f" ({dup['snap_name']})" if dup["snap_name"] else ""
            return jsonify(error=f"Bib {bib}{nm} is already on finisher #{dup['seq']}"), 400
        snap = _snap_for_bib(conn, m, bib)
        conn.execute(
            "UPDATE finishers SET bib=?, snap_name=?, snap_grade=?, snap_gender=?, snap_school=?, "
            "snap_age=? WHERE id=?", (bib, *snap, fid))
        warn = _road_unassigned_warn(conn, m, f["race_id"], bib)
    conn.commit()
    conn.close()
    return jsonify(ok=True, warn=warn)


@bp.post("/finishers/<int:fid>/dq")
@login_required
def finisher_dq(fid):
    conn = db.connect()
    f = conn.execute("SELECT * FROM finishers WHERE id=?", (fid,)).fetchone()
    conn.close()
    if not f:
        abort(404)
    _race_or_403(f["race_id"], can_record_meet)
    conn = db.connect()
    conn.execute("UPDATE finishers SET dq=? WHERE id=?", (0 if f["dq"] else 1, fid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/finishers/<int:fid>/delete")
@login_required
def finisher_delete(fid):
    conn = db.connect()
    f = conn.execute("SELECT * FROM finishers WHERE id=?", (fid,)).fetchone()
    conn.close()
    if not f:
        abort(404)
    _race_or_403(f["race_id"], can_record_meet)
    conn = db.connect()
    conn.execute("DELETE FROM finishers WHERE id=?", (fid,))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/races/<int:rid>/reorder")
@login_required
def race_reorder(rid):
    """Reorder runners while times stay in their slots (handoff §8)."""
    r, m = _race_or_403(rid, can_record_meet)
    order = (request.get_json(silent=True) or {}).get("order", [])
    conn = db.connect()
    rows = conn.execute("SELECT * FROM finishers WHERE race_id=? ORDER BY seq", (rid,)).fetchall()
    conn.close()
    if sorted(order) != sorted(f["id"] for f in rows):
        return jsonify(error="order mismatch"), 400
    # Fixed slots: seq + elapsed + finish_time stay; runner payload moves into them.
    slots = [(f["seq"], f["elapsed_seconds"], f["finish_time"]) for f in rows]
    by_id = {f["id"]: f for f in rows}
    conn = db.connect()
    for i, fid in enumerate(order):
        seq, elapsed, ftime = slots[i]
        f = by_id[fid]
        conn.execute(
            "UPDATE finishers SET seq=?, elapsed_seconds=?, finish_time=? WHERE id=?",
            (seq, elapsed, ftime, fid))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ------------------------------- scoring -------------------------------
def team_scores(runners):
    """MileSplit/Hy-Tek scoring. `runners` = time-sorted list of dicts with
    keys school, place-eligible. Drops teams <5, re-ranks, sums top 5, tracks
    6th/7th displacers. Returns ranked list of team dicts."""
    by_school = defaultdict(list)
    for r in runners:
        if r["school"]:
            by_school[r["school"]].append(r)
    complete = [s for s, rs in by_school.items() if len(rs) >= 5]
    scoring = [r for r in runners if r["school"] in complete]
    for i, r in enumerate(scoring):
        r["score_place"] = i + 1  # re-ranked among complete-team runners only
    teams = []
    for s in complete:
        rs = [r for r in scoring if r["school"] == s][:7]
        top5 = rs[:5]
        teams.append({
            "school": s,
            "score": sum(r["score_place"] for r in top5),
            "places": [r["score_place"] for r in top5],
            "sixth": rs[5]["score_place"] if len(rs) > 5 else None,
            "seventh": rs[6]["score_place"] if len(rs) > 6 else None,
        })
    # Ties break on the 6th runner, then the 7th (standard XC tie-break order).
    teams.sort(key=lambda t: (t["score"], t["sixth"] if t["sixth"] else 9999,
                              t["seventh"] if t["seventh"] else 9999))
    for i, t in enumerate(teams):
        t["rank"] = i + 1
    return teams


def _meet_finishers(mid):
    """All finishers across the meet's races, deduped: a bib that appears in more
    than one heat counts once, at its fastest time (bib-less taps are kept as-is)."""
    conn = db.connect()
    rows = conn.execute(
        "SELECT f.* FROM finishers f JOIN races r ON r.id=f.race_id WHERE r.meet_id=?", (mid,)
    ).fetchall()
    conn.close()
    out, best = [], {}
    for r in rows:
        d = dict(r)
        if d["bib"] is None or d["elapsed_seconds"] is None:
            out.append(d)
            continue
        cur = best.get(d["bib"])
        if cur is None or d["elapsed_seconds"] < cur["elapsed_seconds"]:
            best[d["bib"]] = d
    out.extend(best.values())
    return out


GENDERS = [("M", "Boys"), ("F", "Girls")]


# ------------------------------- road: gender × age-group -------------------------------
def _parse_brackets(text):
    """Parse a free-form age-group list into ordered {label,min,max} dicts.
    Accepts '10 & Under', '10U', '11-14', '15 to 19', '30+', '40 & over', plain '12'."""
    import re
    out = []
    for raw in re.split(r"[\n,;]", text or ""):
        t = raw.strip()
        if not t:
            continue
        low = t.lower().replace("–", "-").replace("—", "-")
        nums = re.findall(r"\d+", low)
        if not nums:
            continue
        n0 = int(nums[0])
        rng = re.search(r"(\d+)\s*(?:-|to)\s*(\d+)", low)
        if rng:
            a, b = int(rng.group(1)), int(rng.group(2))
            lo, hi = min(a, b), max(a, b)
        elif re.search(r"\+|over|older|\bup\b", low):
            lo, hi = n0, 200
        elif re.search(r"under|younger|u\b|u$", low):
            lo, hi = 0, n0
        else:
            lo = hi = n0
        out.append({"label": t, "min": lo, "max": hi})
    return out


def _road_brackets(m):
    """The meet's saved age brackets (ordered list of {label,min,max}), or []."""
    try:
        s = json.loads((m["settings_json"] if "settings_json" in m.keys() else None) or "{}")
    except (ValueError, TypeError):
        return []
    br = s.get("age_brackets") or []
    return [b for b in br if isinstance(b, dict) and "label" in b]


def _bracket_index_label(age, brackets):
    """(sort_index, label) for an age within the bracket list; unmatched → trailing 'No age'."""
    if age is not None:
        for i, b in enumerate(brackets):
            try:
                if int(b["min"]) <= age <= int(b["max"]):
                    return i, b["label"]
            except (ValueError, TypeError, KeyError):
                continue
    return len(brackets), "No age listed"


def _race_finishers(rid):
    """Finishers for ONE road event (not deduped across events), best time per bib."""
    conn = db.connect()
    rows = conn.execute("SELECT * FROM finishers WHERE race_id=? AND elapsed_seconds IS NOT NULL",
                        (rid,)).fetchall()
    conn.close()
    out, best = [], {}
    for r in rows:
        d = dict(r)
        if d["bib"] is None:
            out.append(d)
            continue
        cur = best.get(d["bib"])
        if cur is None or d["elapsed_seconds"] < cur["elapsed_seconds"]:
            best[d["bib"]] = d
    out.extend(best.values())
    return out


def _road_event_groups(fins, brackets):
    """Group one event's finishers by gender × age bracket, ranked within each group."""
    buckets = {}
    for f in fins:
        grank, gword = {"F": (0, "Women"), "M": (1, "Men")}.get(f["snap_gender"], (2, "Open"))
        if brackets:
            bi, blabel = _bracket_index_label(f["snap_age"], brackets)
        else:
            bi, blabel = 0, ""
        buckets.setdefault((grank, gword, bi, blabel), []).append(f)
    groups = []
    for key in sorted(buckets):
        _, gword, _, blabel = key
        place = 0
        indiv = []
        for f in sorted(buckets[key], key=lambda f: f["elapsed_seconds"]):
            if not f["dq"]:
                place += 1
                p = place
            else:
                p = None
            indiv.append({
                "place": p, "time": f["elapsed_seconds"], "bib": f["bib"],
                "name": f["snap_name"] or (f"Bib {f['bib']}" if f["bib"] else "—"),
                "school": f["snap_school"], "grade": None, "age": f["snap_age"],
                "gender": f["snap_gender"], "dq": bool(f["dq"]),
            })
        label = (gword + (" · " + blabel if blabel else "")).strip()
        groups.append((label, indiv))
    return groups


def build_road_results(m):
    """Event-aware: ordered list of (event_name, [(group_label, individuals[])]).
    Age groups per event = the event's own brackets, else the meet default.
    Individual placing only — no team scoring for road races."""
    mid = m["id"]
    default_brackets = _road_brackets(m)
    conn = db.connect()
    races = conn.execute("SELECT * FROM races WHERE meet_id=? ORDER BY id", (mid,)).fetchall()
    conn.close()
    events = []
    for r in races:
        try:
            own = json.loads(r["age_brackets"]) if ("age_brackets" in r.keys() and r["age_brackets"]) else None
        except (ValueError, TypeError):
            own = None
        brackets = own if own else default_brackets
        groups = _road_event_groups(_race_finishers(r["id"]), brackets)
        events.append((r["name"], groups))
    return events


def build_results(mid):
    """Return {gender_key: {'label','individuals','teams'}} plus 'overall'."""
    fins = [f for f in _meet_finishers(mid)
            if f["elapsed_seconds"] is not None]
    conn = db.connect()
    ts = conn.execute("SELECT team_scoring FROM meets WHERE id=?", (mid,)).fetchone()
    conn.close()
    team_on = bool(ts["team_scoring"]) if ts else True
    out = {}
    groups = {"M": [], "F": [], None: []}
    for f in fins:
        groups.get(f["snap_gender"], groups[None]).append(f)

    def rank_group(items):
        items = sorted(items, key=lambda f: f["elapsed_seconds"])
        indiv = []
        place = 0
        for f in items:
            if not f["dq"]:
                place += 1
                p = place
            else:
                p = None
            indiv.append({
                "place": p, "time": f["elapsed_seconds"], "bib": f["bib"],
                "name": f["snap_name"] or (f"Bib {f['bib']}" if f["bib"] else "—"),
                "school": f["snap_school"], "grade": f["snap_grade"],
                "gender": f["snap_gender"], "dq": bool(f["dq"]),
            })
        scoring_runners = [dict(i) for i in indiv if not i["dq"] and i["school"]]
        teams = team_scores(scoring_runners) if team_on else []
        return indiv, teams

    for key, label in GENDERS:
        indiv, teams = rank_group(groups[key])
        if indiv:
            out[key] = {"label": label, "individuals": indiv, "teams": teams}
    if groups[None]:
        indiv, teams = rank_group(groups[None])
        out["U"] = {"label": "Unspecified", "individuals": indiv, "teams": teams}
    return out


# ------------------------------- results (HTML) -------------------------------
def _results_inner(meet, results, name_mode=None):
    if not results:
        return '<div class="card muted">No results yet.</div>'
    html = []
    for key in ("M", "F", "U"):
        g_ = results.get(key)
        if not g_:
            continue
        rows = "".join(
            f'<tr><td>{"" if i["place"] is None else i["place"]}</td>'
            f'<td>{fmt_time(i["time"])}</td>'
            f'<td>{"" if i["bib"] is None or name_mode else i["bib"]}</td>'
            f'<td>{"<s>" if i["dq"] else ""}{escape(demo.display(i["name"], name_mode))}{" (DQ)" if i["dq"] else ""}{"</s>" if i["dq"] else ""}</td>'
            f'<td>{escape(i["school"] or "")}</td>'
            f'<td>{i["grade"] or ""}</td></tr>'
            for i in g_["individuals"])
        html.append(
            f'<div class="card"><h2>{g_["label"]}</h2>'
            f'<table><thead><tr><th>Pl</th><th>Time</th><th>Bib</th><th>Runner</th>'
            f'<th>School</th><th>Gr</th></tr></thead><tbody>{rows}</tbody></table></div>')
    # Team scores broken out by grade × gender (real jr-high XC scores per grade race).
    tparts = []
    for label, teams in _team_grade_gender_groups(meet["id"]):
        if not teams:
            continue
        team_rows = "".join(
            f'<tr><td>{t["rank"]}</td><td>{escape(t["school"])}</td>'
            f'<td><b>{t["score"]}</b></td>'
            f'<td class="muted">{" + ".join(str(p) for p in t["places"])}'
            f'{" (" + str(t["sixth"]) + ("," + str(t["seventh"]) if t["seventh"] else "") + ")" if t["sixth"] else ""}</td></tr>'
            for t in teams)
        tparts.append(
            f'<h3>{escape(label)} — Team scores</h3>'
            f'<table><thead><tr><th>Rank</th><th>School</th><th>Score</th>'
            f'<th>Top 5 (6th,7th)</th></tr></thead><tbody>{team_rows}</tbody></table>')
    if tparts:
        html.append(f'<div class="card"><h2>🏆 Team scores — by grade &amp; gender</h2>{"".join(tparts)}</div>')
    return "".join(html)


@bp.get("/meets/<int:mid>/results")
@login_required
def results_page(mid):
    m = load_meet(mid)
    if not can_view_meet(m):
        abort(403)
    if m["sport"] == "road":
        inner = _road_results_inner(m, build_road_results(m), name_mode=demo.mode_for(g.principal))
    else:
        inner = _results_inner(m, build_results(mid), name_mode=demo.mode_for(g.principal))
    import os
    import base64
    import qrcode
    base = os.environ.get("XC_PUBLIC_URL", request.host_url.rstrip("/"))
    url = f"{base}/r/{m['public_token']}"
    b = io.BytesIO()
    qrcode.make(url).save(b, format="PNG")
    qr_uri = "data:image/png;base64," + base64.b64encode(b.getvalue()).decode()
    qr_card = (f'<div class="card" style="display:flex;gap:1rem;align-items:center;flex-wrap:wrap">'
               f'<img src="{qr_uri}" width="128" height="128" '
               f'style="background:#fff;padding:8px;border-radius:10px">'
               f'<div><b>Public results</b><br>'
               f'<span class="muted">Scan to open the live public results page — share on the big screen '
               f'or a flyer.</span><br><a href="{url}" target="_blank">{escape(url)}</a></div></div>')
    body = (f'<p class="muted"><a href="/meets/{mid}">← {escape(m["name"])}</a></p>'
            f'<h1>{escape(m["name"])} — Results</h1>'
            f'{_xc_tabs(mid, "results", road=(m["sport"]=="road"), organizer=_is_org(m))}'
            f'<div class="row"><a class="btn ghost" href="/r/{m["public_token"]}" target="_blank">'
            f'Public page ↗</a> <a class="btn ghost" href="/meets/{mid}/results.xlsx">Export xlsx</a></div>'
            f'{qr_card}{inner}')
    return shell(g.principal, body, active="meets")


def _meet_by_token(token):
    conn = db.connect()
    m = conn.execute("SELECT * FROM meets WHERE public_token=?", (token,)).fetchone()
    conn.close()
    if not m:
        abort(404)
    return m


def _public_mask(m):
    # Per-meet setting wins: 'bib' -> bib-only, 'initials' -> r.rohd, 'full' -> names.
    pn = m["public_names"] if "public_names" in m.keys() else None
    if pn == "bib":
        return "bib"
    if pn == "initials":
        return "mask"
    if pn == "full":
        return None
    # Unset (older meets): fall back to the district-wide mask_public toggle.
    import json
    conn = db.connect()
    drow = conn.execute("SELECT settings_json FROM districts WHERE id=?",
                        (m["district_id"],)).fetchone()
    conn.close()
    try:
        masked = bool(json.loads((drow["settings_json"] if drow else None) or "{}").get("mask_public"))
    except (ValueError, TypeError):
        masked = False
    return "mask" if masked else None


PUB_CSS = """
*{box-sizing:border-box}
body{margin:0;font:15px/1.5 system-ui,-apple-system,Segoe UI,Roboto,sans-serif;background:#eef1f5;color:#1b2b3a}
.top{background:#12385f;color:#fff;padding:1rem 1.2rem;display:flex;gap:1rem;align-items:center;justify-content:space-between}
.pubfoot{text-align:center;padding:1.6rem 1rem 2.4rem;color:#8a97a5;font-size:.85rem}
.pubfoot .bx{color:#ea6a2d;font-weight:800}.pubfoot .bt{color:#12385f;font-weight:800}
.top .mt{font-size:1.4rem;font-weight:800;line-height:1.15}
.top .sub{opacity:.85;font-size:.85rem;margin-top:.25rem}
.top .right{display:flex;align-items:center;gap:1rem;flex-shrink:0}
.xls{background:#2e8b57;color:#fff;text-decoration:none;font-weight:700;padding:.55rem 1rem;border-radius:9px;white-space:nowrap}
.xls:hover{background:#287a4c}
.hostlogo{height:52px;width:auto;max-width:90px;object-fit:contain;background:#fff;border-radius:8px;padding:5px}
.qr{text-align:center}
.qr img{width:84px;height:84px;background:#fff;padding:6px;border-radius:8px;display:block}
.qr span{display:block;font-size:.66rem;opacity:.85;margin-top:.2rem}
main{max-width:900px;margin:0 auto;padding:1rem 1rem 3rem}
.tabs{display:flex;gap:.6rem;margin:.4rem 0 1.2rem}
.tab{flex:1;text-align:center;padding:.7rem;border-radius:10px;background:#fff;border:1px solid #d5dde6;font-weight:700;color:#33475b;cursor:pointer;font-size:.95rem}
.tab.on{background:#2f6db5;color:#fff;border-color:#2f6db5}
.sec{background:#fff;border:1px solid #d9e0e8;border-radius:12px;overflow:hidden;margin:0 0 1.1rem}
.sec h2{background:#12385f;color:#fff;margin:0;padding:.6rem 1rem;font-size:1.02rem}
table{width:100%;border-collapse:collapse;font-size:.92rem}
th{background:#f1f4f8;text-align:left;padding:.5rem .8rem;font-size:.7rem;text-transform:uppercase;letter-spacing:.04em;color:#5b6b7c}
td{padding:.55rem .8rem;border-top:1px solid #edf1f5}
tbody tr:nth-child(even) td{background:#f8fafb}
.pl{color:#2f6db5;font-weight:800;text-align:center;width:2.2rem}
.tm{font-weight:700;font-variant-numeric:tabular-nums;white-space:nowrap}
.mut{color:#7c8b9a;font-size:.85rem}
@media(max-width:600px){
  .top{flex-wrap:wrap;padding:.75rem .8rem;gap:.7rem}
  .top .mt{font-size:1.1rem}
  .top .right{gap:.7rem}
  .xls{padding:.5rem .8rem;font-size:.9rem}
  .qr img{width:58px;height:58px}
  main{padding:.6rem}
  .tabs{gap:.35rem;margin:.3rem 0 .9rem}
  .tab{padding:.6rem .25rem;font-size:.85rem}
  .sec{border-radius:10px;margin-bottom:.8rem}
  .sec h2{font-size:.92rem;padding:.5rem .7rem}
  th{padding:.35rem .45rem;font-size:.6rem}
  td{padding:.45rem .45rem;font-size:.86rem}
}
"""


def _pub_rows(individuals, mode, show_grade):
    out = []
    for i in individuals:
        nm = i["name"]
        if mode != "bib" and i["bib"] and (not nm or nm == f"Bib {i['bib']}"):
            disp = "Bib not found"
        else:
            disp = demo.public_ident(nm, i["bib"], mode)
        pl = "" if i["place"] is None else i["place"]
        dq = " (DQ)" if i["dq"] else ""
        grcell = f'<td>{i["grade"] or ""}</td>' if show_grade else ""
        out.append(
            f'<tr><td class="pl">{pl}</td>'
            f'<td>{escape(disp)}{dq}</td><td>{escape(i["school"] or "")}</td>'
            f'{grcell}<td class="tm">{fmt_hms(i["time"])}</td></tr>')
    return "".join(out)


def _pub_table(label, individuals, mode, show_grade):
    head = ('<tr><th>Pl</th><th>Name</th><th>School</th>'
            + ('<th>Gr</th>' if show_grade else '') + '<th>Time</th></tr>')
    return (f'<div class="sec"><h2>{escape(label)}</h2><table><thead>{head}</thead>'
            f'<tbody>{_pub_rows(individuals, mode, show_grade)}</tbody></table></div>')


def _grade_gender_groups(mid):
    """Rank finishers within each grade×gender group (for the 'By Grade' tab)."""
    fins = [f for f in _meet_finishers(mid) if f["elapsed_seconds"] is not None]
    buckets = {}
    for f in fins:
        buckets.setdefault((f["snap_grade"], f["snap_gender"]), []).append(f)

    def sk(k):
        grade, gender = k
        return (grade if grade is not None else 999, {"F": 0, "M": 1}.get(gender, 2))

    groups = []
    for key in sorted(buckets, key=sk):
        grade, gender = key
        place = 0
        rows = []
        for f in sorted(buckets[key], key=lambda f: f["elapsed_seconds"]):
            if not f["dq"]:
                place += 1
                p = place
            else:
                p = None
            rows.append({"place": p, "time": f["elapsed_seconds"], "bib": f["bib"],
                         "name": f["snap_name"] or (f"Bib {f['bib']}" if f["bib"] else "—"),
                         "school": f["snap_school"], "grade": f["snap_grade"], "dq": bool(f["dq"])})
        gword = {"F": "Girls", "M": "Boys"}.get(gender, "Other")
        label = f"{grade}th Grade {gword}" if grade is not None else gword
        groups.append((label, rows))
    return groups


def _team_grade_gender_groups(mid):
    """Team scores computed WITHIN each grade×gender group, sorted grade then gender.
    Real jr-high XC runs grade-level races, so teams score per grade + gender —
    not lumped across all grades."""
    conn = db.connect()
    ts = conn.execute("SELECT team_scoring FROM meets WHERE id=?", (mid,)).fetchone()
    conn.close()
    team_on = bool(ts["team_scoring"]) if ts else True
    fins = [f for f in _meet_finishers(mid) if f["elapsed_seconds"] is not None]
    buckets = {}
    for f in fins:
        buckets.setdefault((f["snap_grade"], f["snap_gender"]), []).append(f)

    def sk(k):
        grade, gender = k
        return (grade if grade is not None else 999, {"F": 0, "M": 1}.get(gender, 2))

    groups = []
    for key in sorted(buckets, key=sk):
        grade, gender = key
        runners = [{"school": f["snap_school"]}
                   for f in sorted(buckets[key], key=lambda f: f["elapsed_seconds"])
                   if not f["dq"] and f["snap_school"]]
        teams = team_scores(runners) if team_on else []
        gword = {"F": "Girls", "M": "Boys"}.get(gender, "Other")
        label = f"{grade}th Grade {gword}" if grade is not None else gword
        groups.append((label, teams))
    return groups


def _host_logo_tag(m, cls="hostlogo"):
    """<img> for the meet's host-school logo (top corner of public results), or ''."""
    if not m["host_school_id"]:
        return ""
    conn = db.connect()
    s = conn.execute("SELECT logo_path FROM schools WHERE id=?", (m["host_school_id"],)).fetchone()
    conn.close()
    if s and s["logo_path"]:
        return f'<img class="{cls}" src="{escape(s["logo_path"])}" alt="">'
    return ""


_XLIVE_CACHE = {}      # mid -> (expires_monotonic, races_list)
_XLIVE_TTL = 1.0
XLIVE_HOLD = 30.0      # keep an ended race on the public board this many seconds


def public_live_xc(mid):
    """Public live feed for XC: races currently running (or ended within the last 30s).
    Name + clock only — the results table itself refreshes with finishers."""
    nowm = time.monotonic()
    hit = _XLIVE_CACHE.get(mid)
    if hit and hit[0] > nowm:
        heats = hit[1]
    else:
        heats = _live_races(mid)
        _XLIVE_CACHE[mid] = (nowm + _XLIVE_TTL, heats)
    return {"server_ms": _ms(_now()), "heats": heats}


def _live_races(mid):
    conn = db.connect()
    rows = conn.execute(
        "SELECT id, name, start_time, stop_time FROM races "
        "WHERE meet_id=? AND start_time IS NOT NULL ORDER BY start_time", (mid,)).fetchall()
    now = _now()
    heats = []
    for r in rows:
        stop = _parse(r["stop_time"])
        if stop and (now - stop).total_seconds() > XLIVE_HOLD:
            continue
        # Last 5 to cross (highest seq), returned oldest→newest so the ticker scrolls up.
        recent = conn.execute(
            "SELECT seq, elapsed_seconds, snap_name, snap_school FROM finishers "
            "WHERE race_id=? ORDER BY seq DESC LIMIT 5", (r["id"],)).fetchall()
        fin = [{"n": f["seq"], "elapsed": f["elapsed_seconds"],
                "who": f["snap_name"], "school": f["snap_school"]}
               for f in reversed(recent)]
        heats.append({"name": r["name"] or "Race", "start_ms": _ms(_parse(r["start_time"])),
                      "stop_ms": _ms(stop) if stop else None, "ended": bool(stop),
                      "finishers": fin})
    conn.close()
    return heats


def _public_xc(m, mode):
    mid = m["id"]
    results = build_results(mid)
    conn = db.connect()
    races = conn.execute("SELECT start_time, stop_time FROM races WHERE meet_id=?", (mid,)).fetchall()
    conn.close()
    if races and all(r["stop_time"] for r in races):
        status = "Final"
    elif any(r["start_time"] for r in races):
        status = "Live"
    else:
        status = ""

    overall = "".join(
        _pub_table(f"{lbl} Overall", results[key]["individuals"], mode, True)
        for key, lbl in (("F", "Girls"), ("M", "Boys"), ("U", "Other")) if results.get(key)
    ) or '<div class="sec"><h2>No results yet</h2></div>'

    grade = "".join(_pub_table(lbl, rows, mode, False) for lbl, rows in _grade_gender_groups(mid)) \
        or '<div class="sec"><h2>No results yet</h2></div>'

    team_parts = []
    for label, teams in _team_grade_gender_groups(mid):
        if not teams:
            continue
        trows = "".join(
            f'<tr><td class="pl">{t["rank"]}</td><td>{escape(t["school"])}</td>'
            f'<td class="tm">{t["score"]}</td><td class="mut">'
            f'{" + ".join(str(p) for p in t["places"])}'
            f'{" (" + str(t["sixth"]) + ((", " + str(t["seventh"])) if t["seventh"] else "") + ")" if t["sixth"] else ""}'
            f'</td></tr>' for t in teams)
        team_parts.append(
            f'<div class="sec"><h2>{escape(label)} — Team Scores</h2><table><thead>'
            f'<tr><th>Rank</th><th>School</th><th>Score</th><th>Top 5 (6th, 7th)</th></tr>'
            f'</thead><tbody>{trows}</tbody></table></div>')
    team = "".join(team_parts) or '<div class="sec"><h2>No complete teams yet (need 5+ per school in a grade)</h2></div>'

    sub = escape(m["date"] or "") + (f" · {status}" if status else "")
    return f"""<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width, initial-scale=1">
<title>{escape(m['name'])} — Results</title>{HEAD_EXTRA}<style>{PUB_CSS}
#livebox{{margin:0 0 1.1rem}}
.livecard{{background:#fff;border:2px solid #e8622a;border-radius:12px;padding:.85rem 1rem;margin:0 0 .8rem;box-shadow:0 0 0 4px rgba(232,98,42,.12)}}
.livehd{{font-weight:800;font-size:1.02rem;color:#e8622a;display:flex;align-items:center;gap:.5rem;margin-bottom:.15rem}}
.livecard.final{{border-color:#2e9e5b;box-shadow:0 0 0 4px rgba(46,158,91,.12)}}
.livehd.final{{color:#2e9e5b}}
.livedot{{width:.7rem;height:.7rem;border-radius:50%;background:#2e9e5b;animation:lblink 1s infinite}}
@keyframes lblink{{50%{{opacity:.2}}}}
.liveclock{{font-size:2.6rem;font-weight:800;font-variant-numeric:tabular-nums;text-align:center;margin:.1rem 0;color:#12385f;letter-spacing:.5px}}
.livescroll{{max-height:8.5rem;overflow-y:auto;border-top:1px solid #e6ebf1;margin-top:.5rem}}
.livescroll table{{width:100%;border-collapse:collapse}}
.livescroll td{{padding:.3rem .4rem;border-top:1px solid #eef2f6;font-size:.92rem}}
.livescroll .lp{{color:#2f6db5;font-weight:800;width:2rem;text-align:center}}
.livescroll .lt{{font-variant-numeric:tabular-nums;color:#5b6b7c;width:5.4rem;white-space:nowrap}}
.livescroll .lm{{color:#8a97a5}}
</style></head><body>
<div class="top">
  <div style="display:flex;align-items:center;gap:.8rem">{_host_logo_tag(m)}
    <div><div class="mt">{escape(m['name'])} — Combined</div><div class="sub">{sub}</div></div>
  </div>
</div>
<main>
  <div id="livebox"></div>
  <div class="tabs">
    <button class="tab on" id="t-overall" onclick="tab('overall')">📋 Overall</button>
    <button class="tab" id="t-grade" onclick="tab('grade')">🎽 Sorted</button>
    <button class="tab" id="t-team" onclick="tab('team')">🏆 Team</button>
  </div>
  <div id="v-overall">{overall}</div>
  <div id="v-grade" style="display:none">{grade}</div>
  <div id="v-team" style="display:none">{team}</div>
</main>
<footer class="pubfoot">Powered by {BRAND_HTML}</footer>
<script>
function tab(n){{
  ['overall','grade','team'].forEach(function(k){{
    document.getElementById('v-'+k).style.display = k===n?'':'none';
    document.getElementById('t-'+k).className = 'tab'+(k===n?' on':'');
  }});
  try{{ sessionStorage.setItem('xctab', n); }}catch(e){{}}
}}
try{{ const t=sessionStorage.getItem('xctab'); if(t) tab(t); }}catch(e){{}}
// ---- live 'now running' panel (race name + clock only) ----
const LTOKEN=location.pathname.replace(/^\\/r\\//,'').split('/')[0];
let LOFFSET=0, LTIMER=null;
function lfmt(sec){{ if(sec==null)return''; sec=Math.max(0,sec);
  const h=Math.floor(sec/3600), mm=Math.floor((sec%3600)/60), s=sec-3600*h-60*mm;
  return h+':'+String(mm).padStart(2,'0')+':'+s.toFixed(1).padStart(4,'0'); }}
function lesc(s){{ return String(s==null?'':s).replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c])); }}
function renderLive(heats){{
  const box=document.getElementById('livebox');
  if(!heats.length){{ box.innerHTML=''; return; }}
  let h='';
  heats.forEach(function(ht){{
    const hd = ht.ended
      ? '<div class="livehd final">✅ FINAL · '+lesc(ht.name)+'</div>'
      : '<div class="livehd"><span class="livedot"></span> LIVE · '+lesc(ht.name)+'</div>';
    const clk='<div class="liveclock" data-start="'+ht.start_ms+'"'
      +(ht.ended?(' data-stop="'+ht.stop_ms+'"'):'')+'>0:00:00.0</div>';
    let rows='';
    (ht.finishers||[]).forEach(function(f){{
      rows+='<tr><td class="lp">'+f.n+'</td><td class="lt">'+lfmt(f.elapsed)+'</td>'
        +'<td>'+(f.who?lesc(f.who):'<span class="lm">… crossing</span>')+'</td></tr>';
    }});
    const scroll = rows ? '<div class="livescroll"><table>'+rows+'</table></div>' : '';
    h+='<div class="livecard'+(ht.ended?' final':'')+'">'+hd+clk+scroll+'</div>';
  }});
  box.innerHTML=h;
  document.querySelectorAll('.livescroll').forEach(function(s){{ s.scrollTop=s.scrollHeight; }});
}}
function tickLive(){{
  const now=Date.now()+LOFFSET;
  document.querySelectorAll('.liveclock').forEach(function(c){{
    const st=parseInt(c.getAttribute('data-start'),10);
    const sp=parseInt(c.getAttribute('data-stop'),10);
    if(st) c.textContent=lfmt(((sp||now)-st)/1000);
  }});
}}
async function pollLive(){{
  if(LTIMER){{ clearTimeout(LTIMER); LTIMER=null; }}
  try{{
    const r=await fetch('/r/'+LTOKEN+'/live'); const d=await r.json();
    LOFFSET=d.server_ms-Date.now();
    window.__LIVE_ACTIVE=!!(d.heats&&d.heats.length);
    renderLive(d.heats||[]);
  }}catch(e){{}}
  const gap = document.hidden ? 15000 : (window.__LIVE_ACTIVE ? 2500 : 8000);
  LTIMER=setTimeout(pollLive, gap);
}}
document.addEventListener('visibilitychange',function(){{ if(!document.hidden) pollLive(); }});
setInterval(tickLive,100); pollLive();
// Live scoreboard: refresh every 20s while the page is visible (tab is remembered).
setInterval(function(){{ if(!document.hidden) location.reload(); }}, 20000);
</script>
</body></html>"""


# ------------------------------- road: results (HTML) -------------------------------
def _road_results_inner(m, events, name_mode=None):
    if not any(indiv for _, groups in events for _, indiv in groups):
        return '<div class="card muted">No results yet.</div>'
    html = []
    for event_name, groups in events:
        if not any(indiv for _, indiv in groups):
            continue
        parts = []
        for label, indiv in groups:
            if not indiv:
                continue
            rows = "".join(
                f'<tr><td>{"" if i["place"] is None else i["place"]}</td>'
                f'<td>{fmt_hms(i["time"])}</td>'
                f'<td>{"" if i["bib"] is None or name_mode else i["bib"]}</td>'
                f'<td>{"<s>" if i["dq"] else ""}{escape(demo.display(i["name"], name_mode))}'
                f'{" (DQ)" if i["dq"] else ""}{"</s>" if i["dq"] else ""}</td>'
                f'<td>{i["age"] if i["age"] is not None else ""}</td></tr>'
                for i in indiv)
            parts.append(
                f'<h3 style="margin:.9rem 0 .3rem">{escape(label)}</h3>'
                f'<table><thead><tr><th>Pl</th><th>Time</th><th>Bib</th><th>Name</th>'
                f'<th>Age</th></tr></thead><tbody>{rows}</tbody></table>')
        html.append(f'<div class="card"><h2>🛣 {escape(event_name)}</h2>{"".join(parts)}</div>')
    return "".join(html)


def _pub_road_table(label, individuals, mode):
    rows = []
    for i in individuals:
        nm = i["name"]
        if mode != "bib" and i["bib"] and (not nm or nm == f"Bib {i['bib']}"):
            disp = "Bib not found"
        else:
            disp = demo.public_ident(nm, i["bib"], mode)
        pl = "" if i["place"] is None else i["place"]
        dq = " (DQ)" if i["dq"] else ""
        rows.append(
            f'<tr><td class="pl">{pl}</td><td>{escape(disp)}{dq}</td>'
            f'<td>{i["age"] if i["age"] is not None else ""}</td>'
            f'<td class="tm">{fmt_hms(i["time"])}</td></tr>')
    head = '<tr><th>Pl</th><th>Name</th><th>Age</th><th>Time</th></tr>'
    return (f'<div class="sec"><h2>{escape(label)}</h2><table><thead>{head}</thead>'
            f'<tbody>{"".join(rows)}</tbody></table></div>')


def _public_road(m, mode):
    mid = m["id"]
    events = build_road_results(m)
    conn = db.connect()
    races = conn.execute("SELECT start_time, stop_time FROM races WHERE meet_id=?", (mid,)).fetchall()
    conn.close()
    if races and all(r["stop_time"] for r in races):
        status = "Final"
    elif any(r["start_time"] for r in races):
        status = "Live"
    else:
        status = ""
    ev_sections, ev_names = [], []
    for idx, (event_name, groups) in enumerate(events):
        tables = "".join(_pub_road_table(lbl, indiv, mode) for lbl, indiv in groups if indiv)
        if not tables:
            continue
        ev_names.append((idx, event_name))
        ev_sections.append(
            f'<div class="evsec" data-ev="{idx}">'
            f'<h2 class="evh">🛣 {escape(event_name)}</h2>{tables}</div>')
    body = "".join(ev_sections) or '<div class="sec"><h2>No results yet</h2></div>'
    # Sort/filter-by-event dropdown (only worth showing with more than one event).
    evfilter = ""
    if len(ev_names) > 1:
        opts = '<option value="">All events</option>' + "".join(
            f'<option value="{i}">{escape(nm)}</option>' for i, nm in ev_names)
        evfilter = (
            f'<div class="evbar"><label for="evsel">Event</label>'
            f'<select id="evsel" onchange="filterEv()">{opts}</select></div>')
    sub = escape(m["date"] or "") + (f" · {status}" if status else "")
    return f"""<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width, initial-scale=1">
<title>{escape(m['name'])} — Results</title>{HEAD_EXTRA}<style>{PUB_CSS}
#livebox{{margin:0 0 1.1rem}}
.livecard{{background:#fff;border:2px solid #e8622a;border-radius:12px;padding:.85rem 1rem;margin:0 0 .8rem;box-shadow:0 0 0 4px rgba(232,98,42,.12)}}
.livehd{{font-weight:800;font-size:1.02rem;color:#e8622a;display:flex;align-items:center;gap:.5rem;margin-bottom:.15rem}}
.livecard.final{{border-color:#2e9e5b;box-shadow:0 0 0 4px rgba(46,158,91,.12)}}
.livehd.final{{color:#2e9e5b}}
.livedot{{width:.7rem;height:.7rem;border-radius:50%;background:#2e9e5b;animation:lblink 1s infinite}}
@keyframes lblink{{50%{{opacity:.2}}}}
.liveclock{{font-size:2.6rem;font-weight:800;font-variant-numeric:tabular-nums;text-align:center;margin:.1rem 0;color:#12385f;letter-spacing:.5px}}
.livescroll{{max-height:8.5rem;overflow-y:auto;border-top:1px solid #e6ebf1;margin-top:.5rem}}
.livescroll table{{width:100%;border-collapse:collapse}}
.livescroll td{{padding:.3rem .4rem;border-top:1px solid #eef2f6;font-size:.92rem}}
.livescroll .lp{{color:#2f6db5;font-weight:800;width:2rem;text-align:center}}
.livescroll .lt{{font-variant-numeric:tabular-nums;color:#5b6b7c;width:5.4rem;white-space:nowrap}}
.livescroll .lm{{color:#8a97a5}}
.evbar{{display:flex;align-items:center;gap:.5rem;margin:0 0 1rem}}
.evbar label{{font-size:.72rem;text-transform:uppercase;letter-spacing:.04em;color:#5b6b7c;font-weight:700}}
.evbar select{{flex:1;max-width:340px;padding:.55rem .7rem;border:1px solid #d5dde6;border-radius:9px;
  background:#fff;color:#1b2b3a;font-size:1rem;font-weight:600}}
.evh{{margin:.2rem 0 .7rem;color:#12385f;font-size:1.15rem;font-weight:800}}
</style></head><body>
<div class="top">
  <div style="display:flex;align-items:center;gap:.8rem">{_host_logo_tag(m)}
    <div><div class="mt">🛣 {escape(m['name'])}</div><div class="sub">{sub}</div></div>
  </div>
</div>
<main>
  <div id="livebox"></div>
  {evfilter}
  {body}
</main>
<footer class="pubfoot">Powered by {BRAND_HTML}</footer>
<script>
const LTOKEN=location.pathname.replace(/^\\/r\\//,'').split('/')[0];
let LOFFSET=0, LTIMER=null;
function lfmt(sec){{ if(sec==null)return''; sec=Math.max(0,sec);
  const h=Math.floor(sec/3600), mm=Math.floor((sec%3600)/60), s=sec-3600*h-60*mm;
  return h+':'+String(mm).padStart(2,'0')+':'+s.toFixed(1).padStart(4,'0'); }}
function lesc(s){{ return String(s==null?'':s).replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c])); }}
function renderLive(heats){{
  const box=document.getElementById('livebox');
  if(!heats.length){{ box.innerHTML=''; return; }}
  let h='';
  heats.forEach(function(ht){{
    const hd = ht.ended
      ? '<div class="livehd final">✅ FINAL · '+lesc(ht.name)+'</div>'
      : '<div class="livehd"><span class="livedot"></span> LIVE · '+lesc(ht.name)+'</div>';
    const clk='<div class="liveclock" data-start="'+ht.start_ms+'"'
      +(ht.ended?(' data-stop="'+ht.stop_ms+'"'):'')+'>0:00:00.0</div>';
    let rows='';
    (ht.finishers||[]).forEach(function(f){{
      rows+='<tr><td class="lp">'+f.n+'</td><td class="lt">'+lfmt(f.elapsed)+'</td>'
        +'<td>'+(f.who?lesc(f.who):'<span class="lm">… crossing</span>')+'</td></tr>';
    }});
    const scroll = rows ? '<div class="livescroll"><table>'+rows+'</table></div>' : '';
    h+='<div class="livecard'+(ht.ended?' final':'')+'">'+hd+clk+scroll+'</div>';
  }});
  box.innerHTML=h;
  document.querySelectorAll('.livescroll').forEach(function(s){{ s.scrollTop=s.scrollHeight; }});
}}
function tickLive(){{
  const now=Date.now()+LOFFSET;
  document.querySelectorAll('.liveclock').forEach(function(c){{
    const st=parseInt(c.getAttribute('data-start'),10);
    const sp=parseInt(c.getAttribute('data-stop'),10);
    if(st) c.textContent=lfmt(((sp||now)-st)/1000);
  }});
}}
async function pollLive(){{
  if(LTIMER){{ clearTimeout(LTIMER); LTIMER=null; }}
  try{{
    const r=await fetch('/r/'+LTOKEN+'/live'); const d=await r.json();
    LOFFSET=d.server_ms-Date.now();
    window.__LIVE_ACTIVE=!!(d.heats&&d.heats.length);
    renderLive(d.heats||[]);
  }}catch(e){{}}
  const gap = document.hidden ? 15000 : (window.__LIVE_ACTIVE ? 2500 : 8000);
  LTIMER=setTimeout(pollLive, gap);
}}
document.addEventListener('visibilitychange',function(){{ if(!document.hidden) pollLive(); }});
setInterval(tickLive,100); pollLive();
// Sort/filter by event — remembered across the auto-refresh.
function filterEv(){{
  var sel=document.getElementById('evsel'); if(!sel) return;
  var v=sel.value;
  try{{ sessionStorage.setItem('roadEv', v); }}catch(e){{}}
  document.querySelectorAll('.evsec').forEach(function(s){{
    s.style.display=(!v || s.getAttribute('data-ev')===v)?'':'none'; }});
}}
(function(){{
  var sel=document.getElementById('evsel'); if(!sel) return;
  try{{ var v=sessionStorage.getItem('roadEv');
    if(v && sel.querySelector('option[value="'+v+'"]')) sel.value=v; }}catch(e){{}}
  filterEv();
}})();
setInterval(function(){{ if(!document.hidden && !window.__LIVE_ACTIVE) location.reload(); }}, 20000);
</script>
</body></html>"""


@bp.get("/r/<token>")
def public_results(token):
    m = _meet_by_token(token)
    mode = _public_mask(m)
    if m["sport"] == "road":
        return _public_road(m, mode)
    if m["sport"] == "track":
        from . import track  # lazy import avoids a circular import at module load
        inner = track.results_inner(m["id"], name_mode=mode)
        return f"""<!doctype html><html lang=en><head><meta charset=utf-8>
<meta name=viewport content="width=device-width, initial-scale=1">
<title>{escape(m['name'])} — Results · XCTimer</title><style>{CSS}
main{{max-width:960px;margin:0 auto;padding:1.4rem 1rem 4rem}}
.pubhdr{{display:flex;align-items:center;gap:.7rem;padding:1rem;border-bottom:1px solid var(--line)}}
.pubhdr .hostlogo{{height:48px;width:auto;max-width:80px;object-fit:contain;background:#fff;border-radius:8px;padding:4px}}
.pubfoot{{text-align:center;padding:1.6rem 1rem 2.4rem;color:#8a97a5;font-size:.85rem}}
.pubfoot .bx,.pubfoot .bt{{font-weight:800}}
.livecard{{border:2px solid #e8622a;box-shadow:0 0 0 4px rgba(232,98,42,.12)}}
.livehd{{font-weight:800;font-size:1.05rem;color:#e8622a;display:flex;align-items:center;gap:.5rem;margin-bottom:.2rem}}
.livecard.final{{border-color:#2e9e5b;box-shadow:0 0 0 4px rgba(46,158,91,.12)}}
.livehd.final{{color:#2e9e5b}}
.livedot{{width:.7rem;height:.7rem;border-radius:50%;background:#2e9e5b;animation:lblink 1s infinite}}
@keyframes lblink{{50%{{opacity:.2}}}}
.liveclock{{font-size:2.6rem;font-weight:800;font-variant-numeric:tabular-nums;text-align:center;margin:.1rem 0 .6rem;letter-spacing:.5px}}
.livescroll{{max-height:300px;overflow-y:auto;border-top:1px solid var(--line)}}
.livescroll table{{width:100%}}
.livescroll td{{padding:.35rem .4rem}}
</style></head><body>
<div class="pubhdr">{_host_logo_tag(m)}</div>
<main><h1>{escape(m['name'])}</h1>
<p class="sub">🎽 Track · {escape(m['date'] or '')}</p>
<div id="livebox"></div>
{inner}</main>
<footer class="pubfoot">Powered by {BRAND_HTML}</footer>
<script>
const TOKEN={json.dumps(token)};
let LOFFSET=0;
function lfmt(sec){{ if(sec==null)return''; sec=Math.max(0,sec);
  const h=Math.floor(sec/3600), m=Math.floor((sec%3600)/60), s=sec-3600*h-60*m;
  return h+':'+String(m).padStart(2,'0')+':'+s.toFixed(1).padStart(4,'0'); }}
function lesc(s){{ return String(s==null?'':s).replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c])); }}
let LTIMER=null;
async function pollLive(){{
  if(LTIMER){{ clearTimeout(LTIMER); LTIMER=null; }}   // single chain, never stacked
  try{{
    const r=await fetch('/r/'+TOKEN+'/live'); const d=await r.json();
    LOFFSET=d.server_ms-Date.now();
    window.__LIVE_ACTIVE=!!(d.heats&&d.heats.length);
    renderLive(d.heats||[]);
  }}catch(e){{}}
  // poll fast only while a race runs; back off when idle or the tab is hidden
  const gap = document.hidden ? 15000 : (window.__LIVE_ACTIVE ? 2500 : 8000);
  LTIMER=setTimeout(pollLive, gap);
}}
document.addEventListener('visibilitychange',function(){{ if(!document.hidden) pollLive(); }});
function renderLive(heats){{
  const box=document.getElementById('livebox');
  if(!heats.length){{ box.innerHTML=''; return; }}
  let h='';
  heats.forEach(function(ht){{
    let rows='';
    ht.finishers.forEach(function(f){{
      rows+='<tr><td style="width:2.4rem;color:#e8622a;font-weight:700">'+f.n+'</td>'
        +'<td style="font-variant-numeric:tabular-nums;width:6rem">'+lfmt(f.elapsed)+'</td>'
        +'<td>'+(f.who?lesc(f.who)+(f.school?' <span class=muted>· '+lesc(f.school)+'</span>':'')
                     :'<span class=muted>… crossing</span>')+'</td></tr>';
    }});
    if(!rows) rows='<tr><td class="muted">Waiting for the first finisher…</td></tr>';
    const hd = ht.ended
      ? '<div class="livehd final">✅ FINAL · '+lesc(ht.name)+'</div>'
      : '<div class="livehd"><span class="livedot"></span> LIVE · '+lesc(ht.name)+'</div>';
    const clk = '<div class="liveclock" data-start="'+ht.start_ms+'"'
      +(ht.ended?(' data-stop="'+ht.stop_ms+'"'):'')+'>0:00:00.0</div>';
    h+='<div class="card livecard'+(ht.ended?' final':'')+'">'+hd+clk
      +'<div class="livescroll"><table>'+rows+'</table></div></div>';
  }});
  box.innerHTML=h;
  document.querySelectorAll('.livescroll').forEach(function(s){{ s.scrollTop=s.scrollHeight; }});
}}
function tickLive(){{
  const now=Date.now()+LOFFSET;
  document.querySelectorAll('.liveclock').forEach(function(c){{
    const st=parseInt(c.getAttribute('data-start'),10);
    const sp=parseInt(c.getAttribute('data-stop'),10);   // frozen final time if ended
    if(st) c.textContent=lfmt(((sp||now)-st)/1000);
  }});
}}
setInterval(tickLive,100);
pollLive();
// Full-page refresh keeps the static results fresh — but pause it during a live race
// (the live panel updates itself) and never yank an active search/filter.
setInterval(function(){{
  if(window.__LIVE_ACTIVE) return;
  const q=document.getElementById('rsearch'), g=document.getElementById('rgender');
  if(!document.hidden && (!q||!q.value) && (!g||!g.value)) location.reload();
}}, 20000);
</script></body></html>"""
    return _public_xc(m, mode)


@bp.get("/r/<token>/live")
def public_live_json(token):
    """Live scoreboard feed for the public results page (currently-running track heats)."""
    m = _meet_by_token(token)
    if m["sport"] == "track":
        from . import track
        return jsonify(track.public_live(m["id"], name_mode=_public_mask(m)))
    return jsonify(public_live_xc(m["id"]))


@bp.get("/r/<token>/results.xlsx")
def public_results_xlsx(token):
    m = _meet_by_token(token)
    if m["sport"] == "track":
        from . import track
        fname = (m["name"] or "results").replace(" ", "_")
        return Response(track.track_workbook(m["id"], name_mode=_public_mask(m)),
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})
    return _xlsx_response(m, _public_mask(m))


# ------------------------------- xlsx export -------------------------------
def _results_workbook(mid, name_mode):
    """Build the XC results workbook (Boys/Girls/Unspecified tabs). Returns bytes."""
    import openpyxl
    results = build_results(mid)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    any_tab = False
    for title, key in (("Boys", "M"), ("Girls", "F"), ("Unspecified", "U")):
        g_ = results.get(key)
        if not g_:
            continue
        any_tab = True
        ws = wb.create_sheet(title[:31])
        ws.append(["Place", "Time", "Bib", "Runner", "School", "Grade", "Gender"])
        for i in g_["individuals"]:
            ws.append([i["place"], fmt_time(i["time"]), None if name_mode else i["bib"],
                       demo.public_ident(i["name"], i["bib"], name_mode), i["school"], i["grade"], i["gender"]])
        ws.append([])
        ws.append(["Team Rank", "School", "Score", "Top-5 places"])
        for t in g_["teams"]:
            ws.append([t["rank"], t["school"], t["score"],
                       " + ".join(str(p) for p in t["places"])])
    if not any_tab:
        wb.create_sheet("Results").append(["No results yet"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _road_workbook(m, name_mode):
    """Road results workbook: one sheet per event, sections by gender × age group."""
    import re
    import openpyxl
    events = build_road_results(m)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used = set()
    any_tab = False
    for event_name, groups in events:
        if not any(indiv for _, indiv in groups):
            continue
        any_tab = True
        base = (re.sub(r"[\[\]:*?/\\]", " ", event_name or "Event")[:31]).strip() or "Event"
        title, n = base, 2
        while title in used:
            title = (base[:27] + f" {n}")[:31]
            n += 1
        used.add(title)
        ws = wb.create_sheet(title)
        for label, indiv in groups:
            if not indiv:
                continue
            ws.append([label])
            ws.append(["Place", "Time", "Bib", "Name", "Age"])
            for i in indiv:
                ws.append([i["place"], fmt_hms(i["time"]), None if name_mode else i["bib"],
                           demo.public_ident(i["name"], i["bib"], name_mode), i["age"]])
            ws.append([])
    if not any_tab:
        wb.create_sheet("Results").append(["No results yet"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_response(m, name_mode):
    name = m["name"]
    fname = (name or "results").replace(" ", "_")
    data = _road_workbook(m, name_mode) if m["sport"] == "road" else _results_workbook(m["id"], name_mode)
    return Response(data,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})


@bp.get("/meets/<int:mid>/results.xlsx")
@login_required
def results_xlsx(mid):
    m = load_meet(mid)
    if not can_view_meet(m):
        abort(403)
    return _xlsx_response(m, demo.mode_for(g.principal))


# ------------------------------- camera timing (ArUco prototype) -------------------------------
_CAMERA_PAGE = """
<p class="muted"><a href="/meets/__MID__">← __MNAME__</a></p>
<h1>📷 Camera — __RNAME__</h1>
<div class="card">
  <div style="display:flex;gap:1rem;flex-wrap:wrap;align-items:center;justify-content:space-between">
    <div><b id="cstatus" class="muted">connecting…</b> · <span id="cclock" style="font-variant-numeric:tabular-nums">0:00:00</span>
      · <span id="ccount">0</span> finishers</div>
    <div style="display:flex;gap:.4rem;flex-wrap:wrap">
      <button id="mLine" onclick="setMode('line')">🏁 Finish line</button>
      <button id="mFrame" class="ghost" onclick="setMode('frame')">Whole frame</button>
      <button id="mChute" class="ghost" onclick="setMode('chute')">🚶 Chute scan</button>
      <button id="orbtn" class="ghost" onclick="flipOrient()">⇔ Horizontal</button>
      <button id="dirbtn" class="ghost" onclick="flipDir()">Cross ↓</button>
      <button class="ghost" onclick="selftest()">Self-test</button>
    </div>
  </div>
  <div id="races" style="display:none;flex-wrap:wrap;gap:.4rem;margin-top:.5rem"></div>
  <div id="wrap" style="position:relative;max-width:720px;margin-top:.6rem;line-height:0">
    <video id="v" playsinline muted autoplay
      style="width:100%;border-radius:12px;background:#000;display:block;min-height:200px"></video>
    <canvas id="c" style="position:absolute;inset:0;width:100%;height:100%;touch-action:none"></canvas>
    <button id="enablecam" onclick="startCamera()" style="position:absolute;left:50%;top:50%;
      transform:translate(-50%,-50%);font-size:1.15rem;padding:.9rem 1.6rem;border:0;border-radius:12px;
      background:#ea6a2d;color:#fff;font-weight:800;cursor:pointer;line-height:1.2">📷 Enable camera</button>
  </div>
  <p id="chint" class="muted" style="margin:.4rem 0 0"></p>
  <p class="muted" style="margin:.2rem 0 0">Each bib records once (repeats ignored). Start/stop the
  race from the timing console or phone as usual.</p>
</div>
<div class="card"><h2>Manual entry</h2>
  <div class="row"><div style="max-width:220px">
    <input id="mbib" inputmode="numeric" placeholder="bib #" onkeydown="if(event.key==='Enter')manual()"></div>
  <div style="display:flex;align-items:flex-end"><button onclick="manual()">Record</button></div></div>
</div>
<div class="card"><h2>Recent</h2><div id="clog" class="muted">Nothing yet.</div></div>
<script src="/static/vendor/cv.js"></script>
<script src="/static/vendor/aruco.js"></script>
<script>
const RID=__RID__;
const MEET=__MEET__;   // 1 = whole-event camera (route each bib to its own race)
const RECURL = MEET ? '/meets/'+RID+'/camera-record' : '/races/'+RID+'/finish';
const STATEURL = MEET ? '/meets/'+RID+'/camera-state' : '/races/'+RID+'/state';
// Bigger detection canvas on desktops (more pixels per tag); phones stay lean.
const DW=/Mobi|iPhone|Android.*Mobile/.test(navigator.userAgent)?640:960;
let SEEN=new Set(), DET=null, VID=null, CAN=null, CTX=null, RUNNING=false, LOGN=0, CAPMODE='';
let MODE=localStorage.getItem('camM'+RID)||'frame',   // default to Whole frame (most forgiving)
    ORIENT=localStorage.getItem('camO'+RID)||'h',   // 'h' horizontal line / 'v' vertical line
    LINEP=parseFloat(localStorage.getItem('camP'+RID)||'0.55'),
    DIR=parseInt(localStorage.getItem('camD'+RID)||'1',10),
    TRACKS={}, DRAG=false, FLASH={};
function saveCfg(){ localStorage.setItem('camM'+RID,MODE); localStorage.setItem('camO'+RID,ORIENT);
  localStorage.setItem('camP'+RID,String(LINEP)); localStorage.setItem('camD'+RID,String(DIR)); }
function setMode(m){ MODE=m; saveCfg(); updUi(); }
function flipDir(){ DIR=-DIR; saveCfg(); updUi(); }
function flipOrient(){ ORIENT=ORIENT==='h'?'v':'h'; DIR=1; saveCfg(); updUi(); }
function dirArrow(){ return ORIENT==='v' ? (DIR===1?'→':'←') : (DIR===1?'↓':'↑'); }
function updUi(){
  document.getElementById('mLine').className = MODE==='line'?'':'ghost';
  document.getElementById('mFrame').className = MODE==='frame'?'':'ghost';
  document.getElementById('mChute').className = MODE==='chute'?'':'ghost';
  document.getElementById('mChute').style.display = MEET?'none':'';   // chute fill is per-race
  document.getElementById('dirbtn').style.display = MODE==='line'?'':'none';
  document.getElementById('orbtn').style.display = MODE==='line'?'':'none';
  document.getElementById('orbtn').textContent = ORIENT==='h'?'⇔ Horizontal':'⇕ Vertical';
  document.getElementById('dirbtn').textContent = 'Cross '+dirArrow();
  var hint;
  if(MODE==='line') hint='Camera must be STILL (tripod). Drag the orange line onto the painted finish line — a runner records the moment their tag crosses it in the arrow ('+dirArrow()+') direction. Use ⇔/⇕ to match how runners cross the frame.';
  else if(MODE==='chute') hint='CHUTE SCAN: a helper taps each runner at the LINE (records time + order); this camera reads tags as they walk the single-file chute and fills each open place in order. Race must be in “Tap then scan”.';
  else hint='Records each tag the moment it is seen anywhere in frame — zoom tight on the last few meters before the line. OK handheld.';
  document.getElementById('chint').textContent = hint;
}
function toCanvas(e){ const r=CAN.getBoundingClientRect();
  return {x:(e.clientX-r.left)*CAN.width/r.width, y:(e.clientY-r.top)*CAN.height/r.height}; }
function log(t){ const el=document.getElementById('clog');
  if(!LOGN) el.innerHTML=''; LOGN++;
  el.innerHTML='<div>'+t+'</div>'+el.innerHTML; el.classList.remove('muted'); }
// Audible confirm on FIRST read (not duplicates). WebAudio must be unlocked by a gesture.
let AC=null;
function unlockAudio(){ try{ if(!AC) AC=new (window.AudioContext||window.webkitAudioContext)();
  if(AC.state==='suspended') AC.resume(); }catch(e){} }
document.addEventListener('touchend', unlockAudio, {once:false});
document.addEventListener('click', unlockAudio, {once:false});
function beep(){
  try{ if(!AC){ unlockAudio(); if(!AC) return; }
    const o=AC.createOscillator(), g=AC.createGain(), t=AC.currentTime;
    o.type='sine'; o.frequency.setValueAtTime(1050, t);
    g.gain.setValueAtTime(0.0001, t);
    g.gain.exponentialRampToValueAtTime(0.35, t+0.01);
    g.gain.exponentialRampToValueAtTime(0.0001, t+0.16);
    o.connect(g); g.connect(AC.destination); o.start(t); o.stop(t+0.17);
  }catch(e){}
}
// Non-blocking toast — a modal alert() PAUSES the <video> on mobile (frozen camera).
function toast(msg){
  let t=document.getElementById('ctoast');
  if(!t){ t=document.createElement('div'); t.id='ctoast';
    t.style.cssText='position:fixed;left:50%;bottom:16px;transform:translateX(-50%);z-index:9999;'
      +'background:#12385f;color:#fff;padding:.7rem 1.1rem;border-radius:10px;font-weight:700;'
      +'box-shadow:0 6px 24px rgba(0,0,0,.35);max-width:92%;text-align:center';
    document.body.appendChild(t); }
  t.textContent=msg; t.style.display='';
  clearTimeout(t._to); t._to=setTimeout(function(){ t.style.display='none'; }, 3500);
}
function esc(s){ return String(s==null?'':s).replace(/[&<>"]/g,function(c){
  return {'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]; }); }
function renderRaces(s){
  const box=document.getElementById('races'); box.style.display='flex';
  box.innerHTML=(s.races||[]).map(function(r){
    const c = r.running?'#28c76f':(r.ended?'#8a97a5':'#33475b');
    const badge = r.running?'🟢 running':(r.ended?'✅ final':'⏸ ready');
    return '<span style="background:'+c+';color:#fff;padding:.35rem .7rem;border-radius:999px;'
      +'font-weight:700;font-size:.9rem">'+esc(r.name)+' · '+badge+' · '+r.count+'</span>';
  }).join('');
}
async function poll(){
  try{ const s=await jget(STATEURL);
    if(MEET){
      RUNNING=!!s.any_running;
      document.getElementById('cstatus').textContent = s.any_running
        ? ('🟢 recording — running: '+(s.running_names||[]).join(', '))
        : '⏸ no race running — start one from the console/phone';
      document.getElementById('cclock').style.display='none';
      let tot=0; (s.races||[]).forEach(function(r){ tot+=r.count; });
      document.getElementById('ccount').textContent=tot;
      renderRaces(s);
    } else {
      CAPMODE=s.capture_mode;
      const open=(s.finishers||[]).filter(function(f){return f.bib==null;}).length;
      // Chute scan keeps filling tapped slots for a started race (even if you've stopped the clock).
      RUNNING = (MODE==='chute') ? s.started : (s.started&&!s.stopped);
      document.getElementById('ccount').textContent=s.finishers.length;
      if(MODE==='chute'){
        const bad = (CAPMODE!=='tap'&&CAPMODE!=='tapselect');
        document.getElementById('cstatus').innerHTML = bad
          ? '<b style="color:#f0b429">⚠ Set this race to “Tap then scan” — chute scan fills tapped places.</b>'
          : ('🚶 chute scan · <b>'+open+'</b> place'+(open===1?'':'s')+' awaiting a bib');
      } else {
        document.getElementById('cstatus').textContent = s.stopped?'🏁 ended':(s.started?'🟢 LIVE':'⏸ not started');
      }
      const el=(s.stop_ms||Date.now())-(s.start_ms||Date.now());
      const sec=Math.max(0,el/1000), h=Math.floor(sec/3600), m=Math.floor(sec%3600/60), ss=Math.floor(sec%60);
      document.getElementById('cclock').textContent=s.start_ms?(h+':'+String(m).padStart(2,'0')+':'+String(ss).padStart(2,'0')):'0:00:00';
    }
  }catch(e){}
  setTimeout(poll,2000);
}
let DCAN=null, DCTX=null;
async function boot(){
  CAN=document.getElementById('c'); CTX=CAN.getContext('2d');
  DCAN=document.createElement('canvas'); DCTX=DCAN.getContext('2d',{willReadFrequently:true});
  DET=new AR.Detector();
  updUi();
  CAN.addEventListener('pointerdown',function(e){ if(MODE!=='line')return;
    const p=toCanvas(e);
    const near = ORIENT==='v' ? Math.abs(p.x-LINEP*CAN.width) : Math.abs(p.y-LINEP*CAN.height);
    if(near<48){ DRAG=true; try{CAN.setPointerCapture(e.pointerId);}catch(_e){} }
    e.preventDefault(); });
  CAN.addEventListener('pointermove',function(e){ if(!DRAG)return;
    const p=toCanvas(e);
    LINEP = ORIENT==='v' ? Math.min(.95,Math.max(.05,p.x/CAN.width))
                         : Math.min(.95,Math.max(.05,p.y/CAN.height));
    e.preventDefault(); });
  CAN.addEventListener('pointerup',function(){ if(DRAG){ DRAG=false; saveCfg(); } });
  VID=document.getElementById('v');
  poll();
  startCamera();          // auto-attempt; if the browser blocks it, the Enable button stays for a retry
}
let LOOPING=false;
async function startCamera(){
  const btn=document.getElementById('enablecam');
  if(btn){ btn.textContent='starting…'; btn.disabled=true; }
  document.getElementById('cstatus').textContent='requesting camera…';
  try{
    if(!navigator.mediaDevices||!navigator.mediaDevices.getUserMedia)
      throw {name:'InsecureContext'};
    const s=await navigator.mediaDevices.getUserMedia({audio:false,
      video:{facingMode:'environment',width:{ideal:1920}}});
    VID.srcObject=s; await VID.play();
    document.getElementById('cstatus').textContent='📷 ready';
    document.getElementById('chint').textContent='';
    if(btn) btn.style.display='none';
    if(!LOOPING){ LOOPING=true; loop(); }
  }catch(e){
    const n=(e&&e.name)||'error';
    document.getElementById('cstatus').textContent='camera unavailable ('+n+')';
    const tip = n==='NotAllowedError'
      ? 'Your browser has this site BLOCKED for camera. Click the camera/lock icon at the LEFT of the address bar → set Camera to Allow → then tap “Enable camera”. (On Mac also check System Settings → Privacy & Security → Camera is on for your browser.)'
      : n==='NotReadableError' ? 'Another app is using the camera — close it, then tap “Enable camera”.'
      : n==='NotFoundError' ? 'No camera found on this device.'
      : n==='InsecureContext' ? 'Camera needs a secure (https) connection — open the site via https://xctimer.com.'
      : 'Tap “Enable camera” to try again.';
    document.getElementById('chint').innerHTML='<b style="color:#f0b429">Camera ('+n+').</b> '+tip
      +' Manual entry below still works.';
    if(btn){ btn.style.display=''; btn.textContent='📷 Enable camera'; btn.disabled=false; }
  }
}
function loop(){
  if(VID && VID.videoWidth>0){
    const w=DW, h=Math.round(VID.videoHeight*w/VID.videoWidth)||Math.round(w*0.75);
    if(CAN.width!==w||CAN.height!==h){ CAN.width=w; CAN.height=h; DCAN.width=w; DCAN.height=h; }
    DCTX.drawImage(VID,0,0,w,h);          // offscreen: feed for the detector only
    let ms=[];
    try{ ms=DET.detect(DCTX.getImageData(0,0,w,h)); }catch(e){}
    CTX.clearRect(0,0,w,h);               // visible canvas: transparent overlay atop the <video>
    const now=performance.now(), vert=(ORIENT==='v'), lp=LINEP*(vert?CAN.width:CAN.height);
    ms.forEach(function(m){
      const cx=(m.corners[0].x+m.corners[1].x+m.corners[2].x+m.corners[3].x)/4;
      const cy=(m.corners[0].y+m.corners[1].y+m.corners[2].y+m.corners[3].y)/4;
      const c=vert?cx:cy;                  // position on the crossing axis
      if(MODE==='frame'||MODE==='chute'){ hit(m.id); }
      else{
        const tr=TRACKS[m.id];
        if(tr && now-tr.t<1500 &&
           ((DIR===1 && tr.c<lp && c>=lp) || (DIR===-1 && tr.c>lp && c<=lp))) hit(m.id);
        TRACKS[m.id]={c:c,t:now};
      }
      const col = SEEN.has(m.id) ? '#28c76f' : (MODE==='line' ? '#f0b429' : '#28c76f');
      CTX.strokeStyle=col; CTX.lineWidth=4; CTX.beginPath();
      m.corners.forEach(function(p,i){ i?CTX.lineTo(p.x,p.y):CTX.moveTo(p.x,p.y); });
      CTX.closePath(); CTX.stroke();
      CTX.fillStyle=col; CTX.font='bold 22px sans-serif';
      CTX.fillText('#'+m.id, m.corners[0].x, m.corners[0].y-8);
    });
    if(MODE==='line'){
      CTX.strokeStyle='#ea6a2d'; CTX.lineWidth=3; CTX.beginPath();
      if(vert){ CTX.moveTo(lp,0); CTX.lineTo(lp,h); } else { CTX.moveTo(0,lp); CTX.lineTo(w,lp); }
      CTX.stroke();
      CTX.fillStyle='#ea6a2d'; CTX.font='bold 26px sans-serif';
      const ar=dirArrow();
      if(vert) CTX.fillText(ar, DIR===1?lp+8:lp-32, 30);
      else CTX.fillText(ar, w/2-8, DIR===1?lp+30:lp-12);
      CTX.beginPath(); CTX.arc(vert?lp:w/2, vert?h/2:lp, 10, 0, 7); CTX.fill();
    }
    for(const id in FLASH){
      if(now-FLASH[id]<600){ CTX.fillStyle='rgba(40,199,111,.16)'; CTX.fillRect(0,0,w,h); }
      else delete FLASH[id];
    }
  }
  setTimeout(loop,66);
}
async function hit(id){
  if(!RUNNING || SEEN.has(id)) return;
  SEEN.add(id);
  try{
    const j=await jpost(RECURL,{bib:id});
    if(j&&j.duplicate){ log('#'+id+' — already recorded'); return; }
    if(j&&j.ok===false){                       // whole-event mode: not registered / race not running
      SEEN.delete(id);                          // let it retry when the race is started
      log('⚠ #'+id+' — '+(j.reason||'skipped'));
      return;
    }
    FLASH[id]=performance.now();
    log((MODE==='chute'?'🚶':'📷')+' <b>#'+id+'</b>'+(j.name?(' '+esc(j.name)):'')+(j.race?(' → '+esc(j.race)):'')
        +' ✓'+(j.remaining!=null&&MODE==='chute'?(' · '+j.remaining+' left'):'')+(j.warn?(' ⚠ '+esc(j.warn)):''));
    beep();                                        // audible confirm — first read only
    try{ navigator.vibrate&&navigator.vibrate(35); }catch(e){}
  }catch(e){ SEEN.delete(id); log('#'+id+' ✕ '+e.message); }
}
async function manual(){
  const el=document.getElementById('mbib'); const v=el.value.trim(); if(!v)return;
  try{
    const j=await jpost(RECURL,{bib:v}); el.value=''; el.focus();
    if(j&&j.duplicate){ log('#'+v+' — already recorded'); return; }
    if(j&&j.ok===false){ log('⚠ #'+v+' — '+(j.reason||'skipped')); toast('⚠ '+(j.reason||'skipped')); return; }
    log('⌨️ <b>#'+v+'</b>'+(j.name?(' '+esc(j.name)):'')+(j.race?(' → '+esc(j.race)):'')
        +' ✓'+(j.warn?(' ⚠ '+esc(j.warn)):''));
    if(j&&j.warn) toast('⚠ '+j.warn);
  }catch(e){ toast(e.message); }
}
const ARUCO_CODES={0:[1,0,0,0,0],1:[1,0,1,1,1],2:[0,1,0,0,1],3:[0,1,1,1,0]};
function drawMarker(ctx,x,y,size,id){
  const cell=size/7; ctx.fillStyle='#000'; ctx.fillRect(x,y,size,size); ctx.fillStyle='#fff';
  for(let r=0;r<5;r++){ const bits=ARUCO_CODES[(id>>(8-2*r))&3];
    for(let col=0;col<5;col++) if(bits[col]) ctx.fillRect(x+(col+1)*cell,y+(r+1)*cell,cell,cell); } }
function selftest(){
  const t=document.createElement('canvas'); t.width=640; t.height=480;
  const x=t.getContext('2d',{willReadFrequently:true});
  x.fillStyle='#fff'; x.fillRect(0,0,640,480);
  drawMarker(x,220,140,200,101);
  const ms=new AR.Detector().detect(x.getImageData(0,0,640,480));
  const pass = ms.length===1 && ms[0].id===101;
  toast(pass ? '✅ Self-test PASS — decoded tag '+ms[0].id
             : '❌ Self-test FAIL — got '+JSON.stringify(ms.map(function(m){return m.id;})));
  document.getElementById('cstatus').textContent = pass ? '📷 ready · self-test OK'
    : document.getElementById('cstatus').textContent;
  if(VID && VID.srcObject){ try{ VID.play(); }catch(e){} }   // resume in case anything paused it
}
boot();
</script>"""


@bp.get("/races/<int:rid>/camera")
@login_required
def race_camera(rid):
    """Camera timing (prototype): auto-record ArUco-tagged bibs as they cross.
    Manual entry stays available right on the page for any missed reads."""
    r, m = _race_or_403(rid, can_record_meet)
    body = (_CAMERA_PAGE
            .replace("__MEET__", "0")
            .replace("__RID__", str(rid))
            .replace("__MID__", str(m["id"]))
            .replace("__MNAME__", str(escape(m["name"])))
            .replace("__RNAME__", str(escape(r["name"]))))
    return shell(g.principal, body, active="meets")


@bp.get("/meets/<int:mid>/camera")
@login_required
def meet_camera(mid):
    """Whole-event camera: one iPad at the line records EVERY race — each detected
    bib is routed to the race that participant is registered for. Community events
    only (participants carry a race); no need to pick a race."""
    m = load_meet(mid)
    if not can_record_meet(m) or not _is_org(m):
        abort(403)
    body = (_CAMERA_PAGE
            .replace("__MEET__", "1")
            .replace("__RID__", str(mid))
            .replace("__MID__", str(mid))
            .replace("__MNAME__", str(escape(m["name"])))
            .replace("__RNAME__", "All races"))
    return shell(g.principal, body, active="events")


@bp.get("/meets/<int:mid>/camera-state")
@login_required
def camera_meet_state(mid):
    m = load_meet(mid)
    if not can_record_meet(m) or not _is_org(m):
        abort(403)
    conn = db.connect()
    races = conn.execute("SELECT * FROM races WHERE meet_id=? ORDER BY id", (mid,)).fetchall()
    counts = {r[0]: r[1] for r in conn.execute(
        "SELECT f.race_id, COUNT(*) FROM finishers f JOIN races r ON r.id=f.race_id "
        "WHERE r.meet_id=? GROUP BY f.race_id", (mid,)).fetchall()}
    conn.close()
    out, running = [], []
    for r in races:
        started, stopped = bool(r["start_time"]), bool(r["stop_time"])
        run = started and not stopped
        if run:
            running.append(r["name"])
        out.append({"id": r["id"], "name": r["name"], "running": run, "ended": stopped,
                    "count": counts.get(r["id"], 0),
                    "start_ms": _ms(_parse(r["start_time"])) if started else None,
                    "stop_ms": _ms(_parse(r["stop_time"])) if stopped else None})
    return jsonify(server_ms=_ms(_now()), any_running=bool(running),
                   running_names=running, races=out)


@bp.post("/meets/<int:mid>/camera-record")
@login_required
def camera_record(mid):
    """Route a detected bib to the race that participant is registered for, and record
    a finisher there (if that race is currently running). Returns ok=False + a reason
    the camera page shows in its log — never a hard error, so timing keeps flowing."""
    m = load_meet(mid)
    if not can_record_meet(m) or not _is_org(m):
        abort(403)
    raw = (request.get_json(silent=True) or {}).get("bib")
    try:
        bib = int(str(raw).strip())
    except (TypeError, ValueError):
        return jsonify(ok=False, reason="bad bib")
    conn = db.connect()
    p = conn.execute("SELECT race_id FROM participants WHERE meet_id=? AND bib=?",
                     (mid, bib)).fetchone()
    if not p:
        conn.close()
        return jsonify(ok=False, bib=bib, reason=f"Bib {bib} not registered")
    rid = p["race_id"]
    r = conn.execute("SELECT * FROM races WHERE id=? AND meet_id=?", (rid, mid)).fetchone() if rid else None
    if not r:
        conn.close()
        return jsonify(ok=False, bib=bib, reason=f"Bib {bib} has no race")
    if not r["start_time"] or r["stop_time"]:
        conn.close()
        return jsonify(ok=False, bib=bib, race=r["name"], reason=f"{r['name']} not running")
    if conn.execute("SELECT 1 FROM finishers WHERE race_id=? AND bib=?", (rid, bib)).fetchone():
        conn.close()
        return jsonify(ok=True, duplicate=True, bib=bib, race=r["name"])
    elapsed = (_now() - _parse(r["start_time"])).total_seconds()
    snap = _snap_for_bib(conn, m, bib)
    seq = (conn.execute("SELECT COALESCE(MAX(seq),0) FROM finishers WHERE race_id=?",
                        (rid,)).fetchone()[0]) + 1
    conn.execute(
        "INSERT INTO finishers (race_id, seq, finish_time, elapsed_seconds, bib, "
        "snap_name, snap_grade, snap_gender, snap_school, snap_age) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (rid, seq, _iso(_now()), elapsed, bib, *snap))
    conn.commit()
    conn.close()
    return jsonify(ok=True, bib=bib, race=r["name"], name=snap[0])

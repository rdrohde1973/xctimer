"""Track & field engine (handoff §8). Reference: ~/track/track_timer.py.

Events (grade x gender), entries (individual + relay), seeded/random heat & lane
draws, marks entry (track times + field attempts, incl. High Jump), per-event
placing + points-table scoring, team totals, heat-sheet PDFs, public results,
and AI vision scan-back of a photographed heat sheet.

Deferred refinements (noted, not blocking a runnable meet): athlete-centric bulk
assignment UI, carry-over from last meet, combined distance races, open-pit HJ,
full bar-by-bar HJ make/miss grid (we record best height cleared).
"""
import io
import json
import random
import time
from datetime import datetime, timezone

from markupsafe import escape
from flask import Blueprint, request, redirect, g, abort, jsonify, Response

from . import db, ai, pdfs, demo
from .auth import login_required
from .tenancy import active_district_id, all_districts
from .ui import shell
from .meets import load_meet, can_view_meet, can_setup_meet, can_record_meet

bp = Blueprint("track", __name__)

DEFAULT_EVENT_LIMIT = 4
DEFAULT_LANES = 8
DEFAULT_SECTION = 16


# ------------------------------- helpers -------------------------------
def _t_now():
    return datetime.now(timezone.utc)


def _t_iso(dt):
    return dt.isoformat()


def _t_parse(s):
    if not s:
        return None
    dt = datetime.fromisoformat(s)
    return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)


def _t_ms(dt):
    return int(dt.timestamp() * 1000)


def _heat_key(heat):
    """Normalize a heat arg to an int; 0 means 'all entries' (no heat filter)."""
    return int(heat) if str(heat).isdigit() else 0


def _districts_for_switcher():
    return all_districts() if g.principal.is_super else None


def parse_time(s):
    """'2:05.4' or '12.34' -> seconds (float), or None."""
    s = (s or "").strip()
    if not s:
        return None
    try:
        if ":" in s:
            m, rest = s.split(":", 1)
            return int(m) * 60 + float(rest)
        return float(s)
    except ValueError:
        return None


def fmt_time(sec):
    if sec is None:
        return ""
    m = int(sec // 60)
    s = sec - 60 * m
    return f"{m}:{s:05.2f}" if m else f"{s:.2f}"


def parse_metric(s):
    s = (s or "").strip()
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def fmt_metric(v):
    return "" if v is None else f"{v:.2f}m"


def lane_order(n):
    """Center-out lane preference, e.g. n=8 -> [4,5,3,6,2,7,1,8]."""
    seq, a, b, toggle = [], (n + 1) // 2, (n + 1) // 2 + 1, True
    while len(seq) < n:
        if toggle and a >= 1:
            seq.append(a); a -= 1
        elif b <= n:
            seq.append(b); b += 1
        elif a >= 1:
            seq.append(a); a -= 1
        toggle = not toggle
    return seq


def points_tables():
    conn = db.connect()
    rows = {r["name"]: json.loads(r["point_values_json"])
            for r in conn.execute("SELECT name, point_values_json FROM points_tables").fetchall()}
    conn.close()
    ind = rows.get("Individual (1-8)", [10, 8, 6, 5, 4, 3, 2, 1])
    rel = rows.get("Relay (1-8)", [20, 16, 12, 10, 8, 6, 4, 2])
    return ind, rel


DEFAULT_PT = "Invitational 10-8-6-4-2-1"


def event_limit(meet):
    try:
        v = meet["event_limit"]
        if v:
            return int(v)
    except (KeyError, IndexError, TypeError, ValueError):
        pass
    try:
        return int(json.loads(meet["settings_json"] or "{}").get("event_limit", DEFAULT_EVENT_LIMIT))
    except (ValueError, TypeError):
        return DEFAULT_EVENT_LIMIT


def _meet_points(meet):
    """(point_values, relay_multiplier) for the meet's selected scoring table."""
    conn = db.connect()
    row = None
    if meet and meet["points_table_id"]:
        row = conn.execute("SELECT * FROM points_tables WHERE id=?",
                           (meet["points_table_id"],)).fetchone()
    if not row:
        row = conn.execute("SELECT * FROM points_tables WHERE name=?", (DEFAULT_PT,)).fetchone()
    conn.close()
    vals = json.loads(row["point_values_json"]) if row else [10, 8, 6, 4, 2, 1]
    try:
        mult = row["relay_multiplier"] or 1.0
    except (KeyError, IndexError, TypeError):
        mult = 1.0
    return vals, mult


def _event_kind(ev):
    return ev["kind"]  # track | field | relay


def _is_hj(ev):
    return ev["ename"] == "High Jump"


# High Jump make/miss grid: bar heights are added one at a time as the bar is raised.
def _parse_ht(s):
    """Feet-inches -> total inches (float). Accepts '5-03', \"5'3\\\"\", '5 03',
    '15-06.5', or a bare number (inches). Utah field events use feet & inches."""
    s = (s or "").strip().replace('"', "").replace("”", "").replace("’", "'")
    if not s:
        return None
    for sep in ("-", "'", " "):
        if sep in s:
            a, _, b = s.partition(sep)
            try:
                return int(a.strip()) * 12 + (float(b.strip()) if b.strip() else 0.0)
            except ValueError:
                return None
    try:
        return float(s)
    except ValueError:
        return None


def _fmt_ht(inches):
    """Total inches -> feet-inches, e.g. 63 -> '5-03', 186.5 -> '15-06.50'."""
    if inches is None:
        return ""
    ft = int(inches // 12)
    ins = inches - ft * 12
    if abs(ins - round(ins)) < 0.01:
        return f"{ft}-{int(round(ins)):02d}"
    return f"{ft}-{ins:05.2f}"


def _hj_best(grid):
    """Best cleared bar (inches) from a {bar: marks} grid — a bar is cleared if its
    marks contain an 'O' (e.g. O, XO, XXO)."""
    best = None
    for bar, mk in (grid or {}).items():
        if mk and "O" in str(mk).upper():
            v = _parse_ht(bar)
            if v is not None and (best is None or v > best):
                best = v
    return best


def _combine_meids(conn, me):
    """Meet-event ids sharing this one's combine group (just itself if not combined)."""
    cid = me["combine_id"] if "combine_id" in me.keys() else None
    if cid:
        return [r[0] for r in conn.execute(
            "SELECT id FROM meet_events WHERE combine_id=?", (cid,)).fetchall()]
    return [me["id"]]


def load_meet_event(meid):
    conn = db.connect()
    row = conn.execute(
        "SELECT me.*, e.name AS ename, e.kind, e.unit, e.laned, e.scoring_order, e.sort "
        "FROM meet_events me JOIN events e ON e.id=me.event_id WHERE me.id=?", (meid,)).fetchone()
    conn.close()
    if not row:
        abort(404)
    return row


# ------------------------------- setup section (meet detail) -------------------------------
def _points_tables():
    conn = db.connect()
    rows = conn.execute("SELECT * FROM points_tables ORDER BY builtin DESC, name").fetchall()
    conn.close()
    return rows


def setup_section(m, setup):
    """Track setup (scoring / event limit / lanes) + events list + batch add-events."""
    conn = db.connect()
    catalog = conn.execute("SELECT * FROM events ORDER BY sort").fetchall()
    mes = conn.execute(
        "SELECT me.*, e.name AS ename, e.kind FROM meet_events me JOIN events e ON e.id=me.event_id "
        "WHERE me.meet_id=? ORDER BY e.sort, me.gender, me.grade", (m["id"],)).fetchall()
    counts = {r[0]: r[1] for r in conn.execute(
        "SELECT me.id, COUNT(en.id) FROM meet_events me LEFT JOIN entries en ON en.meet_event_id=me.id "
        "WHERE me.meet_id=? GROUP BY me.id", (m["id"],)).fetchall()}
    conn.close()

    erows = []
    for me in mes:
        g_ = {"M": "Boys", "F": "Girls"}.get(me["gender"], "Open")
        gr = f' · {me["grade"]}' if me["grade"] else ""
        rm = (f'<form class="inline" method="post" action="/meet-events/{me["id"]}/delete" '
              f'onsubmit="return confirm(\'Remove event?\')"><button class="danger">✕</button></form>'
              if setup else "")
        erows.append(f'<tr><td><b><a href="/meet-events/{me["id"]}">{escape(me["ename"])}</a></b></td>'
                     f'<td>{g_}{gr}</td><td>{counts.get(me["id"], 0)} entries</td>'
                     f'<td style="text-align:right">{rm}</td></tr>')
    etbl = (f'<table><tr><th>Event</th><th>Division</th><th>Entries</th><th></th></tr>'
            f'{"".join(erows)}</table>' if mes else '<p class="muted">none yet — add some below</p>')

    add = ""
    if setup:
        def checks(kind):
            return "".join(
                f'<label style="display:inline-flex;gap:.35rem;align-items:center;margin:0 1rem .4rem 0">'
                f'<input type="checkbox" name="event_ids" value="{e["id"]}" style="width:auto">'
                f'{escape(e["name"])}</label>' for e in catalog if e["kind"] == kind)
        grades = "".join(
            f'<label style="display:inline-flex;gap:.3rem;align-items:center;margin-right:.8rem">'
            f'<input type="checkbox" name="grades" value="{g}" style="width:auto">{g}th</label>'
            for g in (6, 7, 8, 9))
        add = f"""
<h3>Add events</h3>
<p class="muted">Pick events, genders, and grades — every combination is created.</p>
<form method="post" action="/meets/{m['id']}/events">
  <div><b>Track</b><br>{checks('track')}</div>
  <div style="margin-top:.5rem"><b>Relays</b><br>{checks('relay')}</div>
  <div style="margin-top:.5rem"><b>Field</b><br>{checks('field')}</div>
  <div style="margin-top:.9rem">
    <label style="display:inline-flex;gap:.3rem;align-items:center;margin-right:1rem">
      <input type="checkbox" name="genders" value="M" checked style="width:auto">Boys</label>
    <label style="display:inline-flex;gap:.3rem;align-items:center;margin-right:1.4rem">
      <input type="checkbox" name="genders" value="F" checked style="width:auto">Girls</label>
    {grades}
    <button type="submit" style="margin-left:.6rem">+ Add events</button>
  </div>
</form>"""

    return f'<div class="card"><h2>Events at this meet ({len(mes)})</h2>{etbl}{add}</div>'


def _apply_hj_schedule(conn, mid, low_s, high_s, inc):
    """Build the bar ladder low→high by `inc` and assign it to the meet's High Jump
    events: girls get the lowest 10 bars, boys the top 10 (both get all when the
    ladder is 10 bars or fewer)."""
    low, high = _parse_ht(low_s), _parse_ht(high_s)
    if low is None or high is None or not inc or inc <= 0 or high < low:
        return
    bars, v = [], low
    while v <= high + 0.01 and len(bars) < 40:
        bars.append(_fmt_ht(v))
        v += inc
    girls = bars[:10] if len(bars) > 10 else bars
    boys = bars[-10:] if len(bars) > 10 else bars
    for me in conn.execute(
            "SELECT id, gender FROM meet_events WHERE meet_id=? AND "
            "event_id IN (SELECT id FROM events WHERE name='High Jump')", (mid,)).fetchall():
        chosen = girls if me["gender"] == "F" else (boys if me["gender"] == "M" else bars)
        conn.execute("UPDATE meet_events SET bar_heights=? WHERE id=?", (json.dumps(chosen), me["id"]))


def settings_fields(m, setup):
    """Scoring / event-limit / lanes fields (no <form> — hosted by the combined
    Save meet setup form in meet_detail). Returns a read-only summary otherwise."""
    tables = _points_tables()
    default_id = next((t["id"] for t in tables if t["name"] == DEFAULT_PT), None)
    sel = m["points_table_id"] or default_id  # default: Invitational 10-8-6-4-2-1
    st = json.loads(m["settings_json"] or "{}")
    hj_low = escape(str(st.get("hj_low", "3-06")))
    hj_high = escape(str(st.get("hj_high", "5-06")))
    hj_inc = escape(str(st.get("hj_inc", 2)))
    if not setup:
        tname = next((t["name"] for t in tables if t["id"] == sel), "—")
        return (f'<p class="muted">Scoring: {escape(tname)} · event limit {event_limit(m)} '
                f'· {m["lanes"] or 8} lanes · HJ bars {hj_low}–{hj_high} every {hj_inc}"</p>')
    # Hide the legacy seed tables from the picker (unless a meet already uses one).
    legacy = {"Individual (1-8)", "Relay (1-8)"}
    tables = [t for t in tables if t["name"] not in legacy or t["id"] == sel]
    topts = "".join(
        f'<option value="{t["id"]}" {"selected" if t["id"]==sel else ""}>'
        f'{escape(t["name"])}</option>' for t in tables)
    return f"""
<label style="margin-top:.8rem">Scoring <span class="muted">— points table for team scores</span></label>
<select name="points_table_id" style="max-width:340px">{topts}</select>
<div class="row" style="margin-top:.6rem">
  <div style="max-width:150px"><label>Event limit</label>
    <input name="event_limit" type="number" value="{event_limit(m)}"></div>
  <div style="max-width:150px"><label>Track lanes</label>
    <input name="lanes" type="number" value="{m['lanes'] or 8}"></div>
</div>
<label style="margin-top:.8rem">High Jump bar schedule
  <span class="muted">— girls get the lowest 10 bars, boys the top 10</span></label>
<div class="row">
  <div style="max-width:150px"><label>Low bar (ft-in)</label>
    <input name="hj_low" value="{hj_low}" placeholder="3-06"></div>
  <div style="max-width:150px"><label>High bar (ft-in)</label>
    <input name="hj_high" value="{hj_high}" placeholder="5-06"></div>
  <div style="max-width:150px"><label>Increment (in)</label>
    <input name="hj_inc" type="number" value="{hj_inc}" placeholder="2"></div>
</div>"""


@bp.post("/meets/<int:mid>/track-setup")
@login_required
def track_setup(mid):
    """One save for the whole track setup: attending schools + host + scoring/limit/lanes."""
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    valid = {r[0] for r in conn.execute(
        "SELECT id FROM schools WHERE district_id=?", (m["district_id"],)).fetchall()}
    # attending schools
    school_ids = [int(x) for x in request.form.getlist("school_ids") if x.isdigit() and int(x) in valid]
    conn.execute("DELETE FROM meet_schools WHERE meet_id=?", (mid,))
    for s in school_ids:
        conn.execute("INSERT OR IGNORE INTO meet_schools (meet_id, school_id) VALUES (?,?)", (mid, s))
    # host
    h = (request.form.get("host_school_id") or "").strip()
    host = int(h) if h.isdigit() and int(h) in valid else None
    conn.execute("UPDATE meets SET host_school_id=? WHERE id=?", (host, mid))
    # scoring / limit / lanes
    pt = request.form.get("points_table_id")
    if pt and pt.isdigit():
        conn.execute("UPDATE meets SET points_table_id=? WHERE id=?", (int(pt), mid))
    el = request.form.get("event_limit")
    if el and el.isdigit():
        settings = json.loads(m["settings_json"] or "{}")
        settings["event_limit"] = max(1, min(20, int(el)))
        conn.execute("UPDATE meets SET settings_json=?, event_limit=? WHERE id=?",
                     (json.dumps(settings), settings["event_limit"], mid))
    ln = request.form.get("lanes")
    if ln and ln.isdigit():
        conn.execute("UPDATE meets SET lanes=? WHERE id=?", (max(2, min(12, int(ln))), mid))
    # High Jump bar schedule (low/high/increment) -> stored + applied to HJ events
    hj_low = (request.form.get("hj_low") or "").strip()
    hj_high = (request.form.get("hj_high") or "").strip()
    hj_inc_s = (request.form.get("hj_inc") or "").strip()
    if hj_low and hj_high and hj_inc_s.isdigit() \
            and _parse_ht(hj_low) is not None and _parse_ht(hj_high) is not None:
        row = conn.execute("SELECT settings_json FROM meets WHERE id=?", (mid,)).fetchone()
        settings = json.loads(row["settings_json"] or "{}")
        settings.update(hj_low=hj_low, hj_high=hj_high, hj_inc=int(hj_inc_s))
        conn.execute("UPDATE meets SET settings_json=? WHERE id=?", (json.dumps(settings), mid))
        _apply_hj_schedule(conn, mid, hj_low, hj_high, int(hj_inc_s))
    from .meets import assign_meet_bibs
    assign_meet_bibs(conn, mid)     # per-meet bibs for the attending athletes
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}")


@bp.get("/meets/<int:mid>/events")
@login_required
def meet_events(mid):
    # Setup now lives on the meet detail page.
    return redirect(f"/meets/{mid}")


@bp.post("/meets/<int:mid>/events")
@login_required
def add_meet_event(mid):
    """Batch: create every event × gender × grade combination (dupes ignored)."""
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    event_ids = [int(x) for x in request.form.getlist("event_ids") if x.isdigit()]
    genders = [g for g in request.form.getlist("genders") if g in ("M", "F")] or [""]
    grades = [g.strip() for g in request.form.getlist("grades")] or [""]
    if not grades:
        grades = [""]
    conn = db.connect()
    valid = {r[0] for r in conn.execute("SELECT id FROM events").fetchall()}
    for eid in event_ids:
        if eid not in valid:
            continue
        for gen in genders:
            for grd in grades:
                conn.execute(
                    "INSERT OR IGNORE INTO meet_events (meet_id, event_id, gender, grade) "
                    "VALUES (?,?,?,?)", (mid, eid, gen, grd))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}")


@bp.post("/meets/<int:mid>/track-settings")
@login_required
def track_settings(mid):
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    pt = request.form.get("points_table_id")
    if pt and pt.isdigit():
        conn.execute("UPDATE meets SET points_table_id=? WHERE id=?", (int(pt), mid))
    el = request.form.get("event_limit")
    if el and el.isdigit():
        settings = json.loads(m["settings_json"] or "{}")
        settings["event_limit"] = max(1, min(20, int(el)))
        conn.execute("UPDATE meets SET settings_json=?, event_limit=? WHERE id=?",
                     (json.dumps(settings), settings["event_limit"], mid))
    ln = request.form.get("lanes")
    if ln and ln.isdigit():
        conn.execute("UPDATE meets SET lanes=? WHERE id=?", (max(2, min(12, int(ln))), mid))
    hj_low = (request.form.get("hj_low") or "").strip()
    hj_high = (request.form.get("hj_high") or "").strip()
    hj_inc_s = (request.form.get("hj_inc") or "").strip()
    if hj_low and hj_high and hj_inc_s.isdigit() \
            and _parse_ht(hj_low) is not None and _parse_ht(hj_high) is not None:
        row = conn.execute("SELECT settings_json FROM meets WHERE id=?", (mid,)).fetchone()
        settings = json.loads(row["settings_json"] or "{}")
        settings.update(hj_low=hj_low, hj_high=hj_high, hj_inc=int(hj_inc_s))
        conn.execute("UPDATE meets SET settings_json=? WHERE id=?", (json.dumps(settings), mid))
        _apply_hj_schedule(conn, mid, hj_low, hj_high, int(hj_inc_s))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}")


@bp.post("/meet-events/<int:meid>/delete")
@login_required
def delete_meet_event(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    conn.execute("DELETE FROM results WHERE entry_id IN (SELECT id FROM entries WHERE meet_event_id=?)", (meid,))
    conn.execute("DELETE FROM track_taps WHERE meet_event_id=?", (meid,))
    conn.execute("DELETE FROM track_clocks WHERE meet_event_id=?", (meid,))
    conn.execute("DELETE FROM entries WHERE meet_event_id=?", (meid,))
    conn.execute("DELETE FROM meet_events WHERE id=?", (meid,))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{me['meet_id']}/events")


# ------------------------------- entries -------------------------------
def _attending_athletes(conn, mid, gender=None, grade=None):
    q = ("SELECT a.*, s.name AS sname FROM athletes a JOIN schools s ON s.id=a.school_id "
         "JOIN meet_schools ms ON ms.school_id=a.school_id WHERE ms.meet_id=?")
    args = [mid]
    if gender:
        q += " AND a.gender=?"; args.append(gender)
    if grade:
        q += " AND a.grade=?"; args.append(grade)
    q += " ORDER BY s.name, a.name"
    return conn.execute(q, args).fetchall()


def _attending_schools(conn, mid):
    return conn.execute(
        "SELECT s.* FROM schools s JOIN meet_schools ms ON ms.school_id=s.id "
        "WHERE ms.meet_id=? ORDER BY s.name", (mid,)).fetchall()


@bp.post("/meet-events/<int:meid>/entries")
@login_required
def add_entry(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    if me["kind"] == "relay":
        sid = request.form.get("school_id")
        if not (sid or "").isdigit():
            conn.close(); abort(400)
        label = (request.form.get("relay_label") or "A").strip()
        members = [x.strip() for x in (request.form.get("members") or "").split(",") if x.strip()]
        conn.execute(
            "INSERT INTO entries (meet_event_id, school_id, relay_label, members_json) "
            "VALUES (?,?,?,?)", (meid, int(sid), label, json.dumps(members)))
    else:
        # Add an athlete by bib (any attending school) — the way an official works.
        raw = (request.form.get("bib") or "").strip()
        if not raw.isdigit():
            conn.close()
            return redirect(f"/meet-events/{meid}?err=bib")
        a = conn.execute("SELECT a.* FROM meet_bibs mb JOIN athletes a ON a.id=mb.athlete_id "
                         "WHERE mb.meet_id=? AND mb.bib=? LIMIT 1", (me["meet_id"], int(raw))).fetchone()
        if not a:
            conn.close()
            return redirect(f"/meet-events/{meid}?err=nobib")
        if conn.execute("SELECT 1 FROM entries WHERE meet_event_id=? AND runner_id=?",
                        (meid, a["id"])).fetchone():
            conn.close()
            return redirect(f"/meet-events/{meid}")   # already in — no-op
        used = conn.execute(
            "SELECT COUNT(*) FROM entries en JOIN meet_events me ON me.id=en.meet_event_id "
            "WHERE me.meet_id=? AND en.runner_id=?", (me["meet_id"], a["id"])).fetchone()[0]
        # Meet directors approve day-of substitutions past the cap — allow an explicit override.
        if used >= event_limit(m) and not request.form.get("force"):
            conn.close()
            return redirect(f"/meet-events/{meid}?err=limit")
        conn.execute("INSERT INTO entries (meet_event_id, runner_id, school_id) VALUES (?,?,?)",
                     (meid, a["id"], a["school_id"]))
    conn.commit()
    conn.close()
    return redirect(f"/meet-events/{meid}")


@bp.post("/entries/<int:eid>/delete")
@login_required
def delete_entry(eid):
    conn = db.connect()
    e = conn.execute("SELECT * FROM entries WHERE id=?", (eid,)).fetchone()
    conn.close()
    if not e:
        abort(404)
    me = load_meet_event(e["meet_event_id"])
    m = load_meet(me["meet_id"])
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    conn.execute("DELETE FROM results WHERE entry_id=?", (eid,))
    conn.execute("DELETE FROM entries WHERE id=?", (eid,))
    conn.commit()
    conn.close()
    return redirect(f"/meet-events/{e['meet_event_id']}")


# ------------------------------- assign athletes -------------------------------
def _track_tabs(mid, active):
    def tab(href, label, key):
        on = "background:var(--panel2);color:var(--fg)" if active == key else "color:var(--mut)"
        return (f'<a href="{href}" style="padding:.4rem .9rem;border-radius:8px;'
                f'text-decoration:none;{on}">{label}</a>')
    return ('<div style="display:flex;gap:.3rem;margin:.4rem 0 1rem;border-bottom:1px solid var(--line);'
            'padding-bottom:.5rem;flex-wrap:wrap">'
            + tab(f"/meets/{mid}", "⚙️ Setup", "setup")
            + tab(f"/meets/{mid}/assign", "👤 Assign athletes", "assign")
            + tab(f"/meets/{mid}/meet-day", "🏁 Meet day", "meetday")
            + tab(f"/meets/{mid}/track-results", "📊 Results", "results")
            + '</div>')


def _can_assign_school(m, sid):
    p = g.principal
    if not p or p.meet_scope:
        return False
    if p.is_super:
        return True
    if p.district_id != m["district_id"]:
        return False
    if p.role == "district_admin":
        return True
    if p.role == "coach":
        return sid in p.school_ids()
    return False


def _eligible(me, athlete):
    g_ = me["gender"]
    if g_ and g_ != athlete["gender"]:
        return False
    gr = me["grade"]
    if gr not in (None, "") and str(gr) != str(athlete["grade"]):
        return False
    return True


@bp.get("/meets/<int:mid>/assign")
@login_required
def assign_page(mid):
    m = load_meet(mid)
    if not can_view_meet(m) or m["sport"] != "track":
        abort(403)
    conn = db.connect()
    schools = _attending_schools(conn, mid)
    # scope the school picker
    p = g.principal
    pickable = [s for s in schools if _can_assign_school(m, s["id"])]
    if not pickable:
        conn.close()
        abort(403)
    sid = request.args.get("school")
    sid = int(sid) if (sid or "").isdigit() and any(s["id"] == int(sid) for s in pickable) else pickable[0]["id"]
    school = next(s for s in pickable if s["id"] == sid)

    athletes = conn.execute(
        "SELECT * FROM athletes WHERE school_id=? AND does_track=1 AND active=1 "
        "ORDER BY grade, gender, name", (sid,)).fetchall()
    bibmap = {r["athlete_id"]: r["bib"] for r in conn.execute(
        "SELECT athlete_id, bib FROM meet_bibs WHERE meet_id=?", (mid,)).fetchall()}
    mes = conn.execute(
        "SELECT me.*, e.name AS ename, e.kind FROM meet_events me JOIN events e ON e.id=me.event_id "
        "WHERE me.meet_id=? ORDER BY e.sort, me.gender, me.grade", (mid,)).fetchall()
    entered = {}  # athlete_id -> set(meid)
    for r in conn.execute(
        "SELECT en.runner_id, en.meet_event_id FROM entries en JOIN meet_events me ON me.id=en.meet_event_id "
        "WHERE me.meet_id=? AND en.runner_id IS NOT NULL", (mid,)).fetchall():
        entered.setdefault(r["runner_id"], set()).add(r["meet_event_id"])
    relay_squads = {}  # meid -> members_json (this school)
    for r in conn.execute(
        "SELECT meet_event_id, members_json FROM entries WHERE school_id=? AND runner_id IS NULL "
        "AND meet_event_id IN (SELECT id FROM meet_events WHERE meet_id=?)", (sid, mid)).fetchall():
        relay_squads[r["meet_event_id"]] = json.loads(r["members_json"] or "[]")
    conn.close()

    limit = event_limit(m)
    indiv_mes = [me for me in mes if me["kind"] != "relay"]
    relay_mes = [me for me in mes if me["kind"] == "relay"]

    setup = _can_assign_school(m, sid)
    # school picker
    picker = "".join(
        f'<a class="btn {"" if s["id"]==sid else "ghost"}" style="white-space:nowrap" '
        f'href="/meets/{mid}/assign?school={s["id"]}">{escape(s["name"])}</a>' for s in pickable)

    # per-athlete rows — individual/field AND relay events are all just checkboxes
    # which relays (this school) each athlete is already on
    relay_member = {}  # aid -> set(meid)
    for meid, members in relay_squads.items():
        for a in athletes:
            if a["name"] in members:
                relay_member.setdefault(a["id"], set()).add(meid)

    def box(me, aid, checked, relay):
        rel = ' data-relay="1"' if relay else f' onchange="cnt({aid})"'
        tag = "🔗 " if relay else ""
        # No grade/gender tag — eligibility already scopes each athlete to their events.
        return (f'<label style="display:inline-flex;gap:.3rem;align-items:center;margin:0 .8rem .3rem 0">'
                f'<input type="checkbox" name="me_{aid}" value="{me["id"]}" style="width:auto" '
                f'data-ath="{aid}"{rel} {"checked" if checked else ""}>'
                f'{tag}{escape(me["ename"])}</label>')

    rows = []
    for a in athletes:
        cur = entered.get(a["id"], set())
        relcur = relay_member.get(a["id"], set())
        parts = [box(me, a["id"], me["id"] in cur, False) for me in indiv_mes if _eligible(me, a)]
        parts += [box(me, a["id"], me["id"] in relcur, True) for me in relay_mes if _eligible(me, a)]
        boxes = "".join(parts) or '<span class="muted">no eligible events</span>'
        rows.append(
            f'<tr class="arow" data-g="{a["gender"] or ""}" data-gr="{a["grade"] or ""}">'
            f'<td><b>{escape(a["name"])}</b><br><span class="muted">gr {a["grade"] or "?"} '
            f'{a["gender"] or ""}{" · bib "+str(bibmap[a["id"]]) if bibmap.get(a["id"]) else ""}</span></td>'
            f'<td>{boxes}<div class="muted" id="c{a["id"]}"></div></td></tr>')
    hdr = f'Events (limit {limit}; relays 🔗 not counted)' if relay_mes else f'Events (limit {limit})'
    ath_tbl = (f'<table><tr><th>Athlete</th><th>{hdr}</th></tr>{"".join(rows)}</table>'
               if athletes else '<p class="muted">No athletes on this roster.</p>')
    # Filter dropdowns (gender / grade) over the athlete list.
    grades_present = sorted({a["grade"] for a in athletes if a["grade"] is not None})
    grade_opts = '<option value="">All grades</option>' + "".join(
        f'<option value="{g}">{g}th grade</option>' for g in grades_present)
    filter_bar = ("" if not athletes else
                  '<div class="row" style="gap:.6rem;flex-wrap:wrap;margin:.2rem 0 .8rem">'
                  '<div><label>Gender</label>'
                  '<select id="fg" onchange="filterAth()" style="max-width:140px">'
                  '<option value="">All</option><option value="M">Boys</option>'
                  '<option value="F">Girls</option></select></div>'
                  '<div><label>Grade</label>'
                  f'<select id="fgr" onchange="filterAth()" style="max-width:140px">{grade_opts}</select></div>'
                  '</div>')
    relay_block = ""

    body = f"""
<p class="muted"><a href="/meets">← Meets</a></p>
<h1>{escape(m['name'])}</h1>{_track_tabs(mid, 'assign')}
<div class="card" style="display:flex;flex-wrap:wrap;gap:.45rem;align-items:center">
  <b style="margin-right:.2rem">School:</b> {picker}</div>
<div class="card" style="display:flex;flex-wrap:wrap;gap:.45rem;align-items:center">
  <b style="margin-right:.2rem">🖨 {escape(school['name'])} packet:</b>
  <a class="btn ghost" href="/meets/{mid}/school/{sid}/stickers.pdf">Stickers — QR</a>
  <a class="btn ghost" href="/meets/{mid}/school/{sid}/stickers.pdf?code=aruco">Stickers — ArUco</a>
  <a class="btn ghost" href="/meets/{mid}/school/{sid}/biblist.pdf">Bib list + events</a></div>
<div class="card">
  <div class="row" style="justify-content:space-between;align-items:center;flex-wrap:wrap;gap:.5rem">
    <h2 style="margin:0">{escape(school['name'])} — assign athletes</h2>
    <form method="post" action="/meets/{mid}/carryover?school={sid}" style="margin:0">
      <button class="ghost" type="submit" onclick="return confirm('Carry each athlete\\'s events forward from their last meet?')">↩ Carry over last meet</button>
    </form>
  </div>
  {filter_bar}
  <form method="post" action="/meets/{mid}/assign?school={sid}">
    {ath_tbl}
    {relay_block}
    <button type="submit">💾 Save entries</button>
  </form>
</div>"""
    body += f"""
<script>
const LIMIT={limit};
function cnt(aid){{
  const boxes=document.querySelectorAll('input[data-ath="'+aid+'"]:not([data-relay])');
  let n=0; boxes.forEach(b=>{{if(b.checked)n++;}});
  const el=document.getElementById('c'+aid);
  el.textContent=n+' / {limit}';
  el.style.color = n>LIMIT ? 'var(--err)' : 'var(--mut)';
}}
document.querySelectorAll('input[data-ath]:not([data-relay])').forEach(b=>cnt(b.dataset.ath));
function filterAth(){{
  const g=document.getElementById('fg').value, gr=document.getElementById('fgr').value;
  document.querySelectorAll('tr.arow').forEach(function(r){{
    const okG = !g || r.dataset.g===g, okGr = !gr || r.dataset.gr===gr;
    r.style.display = (okG && okGr) ? '' : 'none';
  }});
}}
</script>
"""
    return shell(g.principal, body, active="meets")


@bp.post("/meets/<int:mid>/assign")
@login_required
def save_assign(mid):
    m = load_meet(mid)
    sid = request.args.get("school")
    sid = int(sid) if (sid or "").isdigit() else None
    if sid is None or not _can_assign_school(m, sid):
        abort(403)
    limit = event_limit(m)
    conn = db.connect()
    athletes = conn.execute("SELECT id, name FROM athletes WHERE school_id=? AND does_track=1 AND active=1",
                            (sid,)).fetchall()
    valid_me = {r[0]: r[1] for r in conn.execute(
        "SELECT me.id, e.kind FROM meet_events me JOIN events e ON e.id=me.event_id "
        "WHERE me.meet_id=?", (mid,)).fetchall()}
    # each athlete's checked meet_events (relays included — same checkbox flow)
    sel = {a["id"]: [int(x) for x in request.form.getlist(f"me_{a['id']}")
                     if x.isdigit() and int(x) in valid_me] for a in athletes}

    for a in athletes:
        aid = a["id"]
        indiv = [meid for meid in sel[aid] if valid_me[meid] != "relay"][:limit]  # cap excludes relays
        current = {r[0] for r in conn.execute(
            "SELECT en.meet_event_id FROM entries en JOIN meet_events me ON me.id=en.meet_event_id "
            "WHERE me.meet_id=? AND en.runner_id=?", (mid, aid)).fetchall()}
        for meid in set(indiv) - current:
            conn.execute("INSERT INTO entries (meet_event_id, runner_id, school_id) VALUES (?,?,?)",
                         (meid, aid, sid))
        for meid in current - set(indiv):
            conn.execute("DELETE FROM results WHERE entry_id IN "
                         "(SELECT id FROM entries WHERE meet_event_id=? AND runner_id=?)", (meid, aid))
            conn.execute("DELETE FROM entries WHERE meet_event_id=? AND runner_id=?", (meid, aid))

    # relay squads: members = the athletes (this school) who checked that relay.
    for meid, kind in valid_me.items():
        if kind != "relay":
            continue
        members = [a["name"] for a in athletes if meid in sel[a["id"]]][:4]
        existing = conn.execute(
            "SELECT id, members_json FROM entries "
            "WHERE meet_event_id=? AND school_id=? AND runner_id IS NULL", (meid, sid)).fetchone()
        try:
            existing_members = json.loads(existing["members_json"] or "[]") if existing else []
        except (ValueError, TypeError):
            existing_members = []
        if members == existing_members:
            continue  # unchanged — keep the entry and any recorded relay result
        # Squad changed: clear the old entry (and its result) first to satisfy the
        # results→entries foreign key, then insert the new squad.
        if existing:
            conn.execute("DELETE FROM results WHERE entry_id=?", (existing["id"],))
            conn.execute("DELETE FROM entries WHERE id=?", (existing["id"],))
        if members:
            conn.execute("INSERT INTO entries (meet_event_id, school_id, relay_label, members_json) "
                         "VALUES (?,?,?,?)", (meid, sid, "A", json.dumps(members)))
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}/assign?school={sid}")


@bp.post("/meets/<int:mid>/carryover")
@login_required
def carryover(mid):
    m = load_meet(mid)
    sid = request.args.get("school")
    sid = int(sid) if (sid or "").isdigit() else None
    if sid is None or not _can_assign_school(m, sid):
        abort(403)
    limit = event_limit(m)
    conn = db.connect()
    athletes = conn.execute("SELECT * FROM athletes WHERE school_id=? AND does_track=1 AND active=1",
                            (sid,)).fetchall()
    # this meet's events keyed by (event_id, gender, grade)
    this_mes = {}
    for me in conn.execute(
        "SELECT me.id, me.event_id, me.gender, me.grade, e.kind FROM meet_events me "
        "JOIN events e ON e.id=me.event_id WHERE me.meet_id=?", (mid,)).fetchall():
        this_mes[(me["event_id"], me["gender"] or "", str(me["grade"] or ""))] = me
    for a in athletes:
        # athlete's most recent prior track meet with entries
        prior = conn.execute(
            "SELECT me.event_id, me.gender, me.grade, m.date, m.id AS mid FROM entries en "
            "JOIN meet_events me ON me.id=en.meet_event_id JOIN meets m ON m.id=me.meet_id "
            "WHERE en.runner_id=? AND m.sport='track' AND m.id!=? AND (m.date < ? OR m.date IS NULL) "
            "ORDER BY m.date DESC", (a["id"], mid, m["date"] or "9999")).fetchall()
        if not prior:
            continue
        last_date = prior[0]["date"]
        last_events = [r for r in prior if r["date"] == last_date]
        current = {r[0] for r in conn.execute(
            "SELECT en.meet_event_id FROM entries en JOIN meet_events me ON me.id=en.meet_event_id "
            "WHERE me.meet_id=? AND en.runner_id=?", (mid, a["id"])).fetchall()}
        count = len(current)
        for pe in last_events:
            key = (pe["event_id"], pe["gender"] or "", str(pe["grade"] or ""))
            me = this_mes.get(key)
            if not me or me["kind"] == "relay" or not _eligible(me, a) or count >= limit:
                continue
            if me["id"] in current:
                continue
            conn.execute("INSERT INTO entries (meet_event_id, runner_id, school_id) VALUES (?,?,?)",
                         (me["id"], a["id"], sid))
            current.add(me["id"])
            count += 1
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}/assign?school={sid}")


# ------------------------------- seeding -------------------------------
def _season_seed_value(conn, entry, event_id, exclude_meet_id):
    """Seed value: best (fastest) time in this event at OTHER meets; falls back to
    the entry's manual seed. Lower = faster. None if no prior mark and no seed."""
    if entry["runner_id"]:
        row = conn.execute(
            "SELECT MIN(r.mark_seconds) FROM results r JOIN entries en ON en.id=r.entry_id "
            "JOIN meet_events me ON me.id=en.meet_event_id "
            "WHERE en.runner_id=? AND me.event_id=? AND me.meet_id!=? AND r.dq=0 "
            "AND r.mark_seconds IS NOT NULL", (entry["runner_id"], event_id, exclude_meet_id)).fetchone()
    else:  # relay: school's best time in this event
        row = conn.execute(
            "SELECT MIN(r.mark_seconds) FROM results r JOIN entries en ON en.id=r.entry_id "
            "JOIN meet_events me ON me.id=en.meet_event_id "
            "WHERE en.school_id=? AND en.runner_id IS NULL AND me.event_id=? AND me.meet_id!=? "
            "AND r.dq=0 AND r.mark_seconds IS NOT NULL",
            (entry["school_id"], event_id, exclude_meet_id)).fetchone()
    best = row[0] if row else None
    return best if best is not None else entry["seed"]


def _draw(conn, me, m, mode, entries=None, laned_override=None):
    """Assign heat + lane to a set of entries. Seeded = fastest heat last, center-out lanes."""
    laned = bool(me["laned"]) if laned_override is None else laned_override
    lanes = m["lanes"] or DEFAULT_LANES
    # Laned sprints: one runner per lane. Distance (800/1600/3200) runs a waterfall
    # section that holds ~2x the lanes — so don't split into a 2nd section until there
    # are more than double the lanes (e.g. >16 on an 8-lane track).
    size = lanes if laned else 2 * lanes
    if entries is None:
        entries = conn.execute("SELECT * FROM entries WHERE meet_event_id=?", (me["id"],)).fetchall()
    entries = list(entries)
    if mode == "random":
        random.shuffle(entries)
    else:  # seeded by season best (fastest first)
        vals = {e["id"]: _season_seed_value(conn, e, me["event_id"], m["id"]) for e in entries}
        seeded = sorted([e for e in entries if vals[e["id"]] is not None], key=lambda e: vals[e["id"]])
        unseeded = [e for e in entries if vals[e["id"]] is None]
        random.shuffle(unseeded)
        entries = seeded + unseeded
    n = len(entries)
    groups = [entries[i:i + size] for i in range(0, n, size)] or []
    H = len(groups)
    for gi, grp in enumerate(groups):
        heat_no = H - gi  # fastest group runs last
        # Laned: center-out lanes. Distance/section: running order (pole = 1).
        lanes = lane_order(size)[:len(grp)] if laned else list(range(1, len(grp) + 1))
        for idx, e in enumerate(grp):
            conn.execute("UPDATE entries SET heat=?, lane=? WHERE id=?",
                         (heat_no, lanes[idx], e["id"]))


@bp.post("/meet-events/<int:meid>/clear-draw")
@login_required
def clear_draw(meid):
    """Undo the heat/lane draw (entries stay) — for events re-declared day-of."""
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    meids = _combine_meids(conn, me)
    qm = ",".join("?" * len(meids))
    conn.execute(f"UPDATE entries SET heat=NULL, lane=NULL WHERE meet_event_id IN ({qm})",
                 tuple(meids))
    conn.commit()
    conn.close()
    return redirect(f"/meet-events/{meid}")


@bp.post("/meet-events/<int:meid>/seed")
@login_required
def seed_event(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_setup_meet(m):
        abort(403)
    conn = db.connect()
    _draw(conn, me, m, request.form.get("mode", "seeded"))
    conn.commit()
    conn.close()
    return redirect(f"/meet-events/{meid}")


# ------------------------------- open pit console -------------------------------
@bp.get("/meets/<int:mid>/pit")
@login_required
def pit_console(mid):
    """Open-pit field entry: the pit official records ANY athlete by bib — the mark
    is auto-filed into that athlete's own gender×grade division (created if needed).
    This is how a real jr-high pit runs: whoever steps up, one console."""
    m = load_meet(mid)
    if not can_record_meet(m) or m["sport"] != "track":
        abort(403)
    conn = db.connect()
    evs = conn.execute(
        "SELECT DISTINCT e.id, e.name FROM events e JOIN meet_events me ON me.event_id=e.id "
        "WHERE me.meet_id=? AND e.kind='field' AND e.name!='High Jump' ORDER BY e.sort",
        (mid,)).fetchall()
    conn.close()
    opts = "".join(f'<option value="{e["id"]}">{escape(e["name"])}</option>' for e in evs)
    if not opts:
        opts = '<option value="">— no field events at this meet —</option>'
    body = f"""
<p class="muted"><a href="/meets">← Meets</a></p>
<h1>{escape(m['name'])}</h1>{_track_tabs(mid, 'meetday')}
<h2>🏖 Open pit</h2>
<p class="sub">Record whoever steps up, by bib. The mark files into the athlete's own
division automatically. Enter feet-inches — just type <b>10 6</b> for 10'6" (a dash or
apostrophe works too); <b>F</b> = foul.
<b>On a phone?</b> Use the 🏖 Pit tab in the Track Timer app for a touch-friendly version.</p>
<div class="card">
  <div class="row" style="flex-wrap:wrap;gap:.6rem">
    <div style="max-width:200px"><label>Event</label><select id="pev">{opts}</select></div>
    <div style="max-width:120px"><label>Bib</label>
      <input id="pbib" inputmode="numeric" autocomplete="off"
        onkeydown="if(event.key==='Enter')document.getElementById('pa0').focus()"></div>
    <div style="max-width:110px"><label>Att 1</label><input id="pa0"></div>
    <div style="max-width:110px"><label>Att 2</label><input id="pa1"></div>
    <div style="max-width:110px"><label>Att 3</label><input id="pa2"
      onkeydown="if(event.key==='Enter')pitPost()"></div>
    <div style="display:flex;align-items:flex-end"><button onclick="pitPost()">✔ Record</button></div>
  </div>
  <div id="pmsg" style="margin-top:.5rem"></div>
</div>
<div class="card"><h2>Recorded this session</h2><table id="plog">
  <tr><th>Bib</th><th>Athlete</th><th>Event</th><th>Division</th><th>Attempts</th><th>Best</th></tr></table></div>
<script>
async function pitPost(){{
  const ev=document.getElementById('pev').value, bib=document.getElementById('pbib').value.trim();
  const atts=[0,1,2].map(k=>document.getElementById('pa'+k).value.trim());
  const box=document.getElementById('pmsg');
  if(!ev){{box.innerHTML='<p class="msg err">No field events at this meet.</p>';return;}}
  if(!bib){{box.innerHTML='<p class="msg err">Enter a bib.</p>';return;}}
  if(!atts.some(a=>a)){{box.innerHTML='<p class="msg err">Enter at least one attempt.</p>';return;}}
  try{{
    const j=await jpost('/meets/{mid}/pit/post',{{event_id:ev,bib:bib,attempts:atts}});
    box.innerHTML='<p class="msg ok">✔ '+esc(j.name)+' — <b>'+esc(j.event)+'</b> — '+esc(j.division)+' — best '+esc(j.best||'—')+'</p>';
    const t=document.getElementById('plog');
    t.insertAdjacentHTML('afterbegin','<tr><td>'+esc(bib)+'</td><td>'+esc(j.name)+'</td><td><b>'+esc(j.event)+'</b></td><td>'+esc(j.division)
      +'</td><td>'+esc(atts.filter(a=>a).join(', '))+'</td><td><b>'+esc(j.best||'')+'</b></td></tr>');
    ['pbib','pa0','pa1','pa2'].forEach(k=>document.getElementById(k).value='');
    document.getElementById('pbib').focus();
  }}catch(e){{ box.innerHTML='<p class="msg err">'+esc(e.message)+'</p>'; }}
}}
</script>"""
    return shell(g.principal, body, active="meets")


@bp.get("/meets/<int:mid>/pit/lookup")
@login_required
def pit_lookup(mid):
    """Live bib -> athlete lookup for the open-pit console (confirm-as-you-type)."""
    m = load_meet(mid)
    if not can_record_meet(m) or m["sport"] != "track":
        abort(403)
    try:
        bib = int((request.args.get("bib") or "").strip())
    except (TypeError, ValueError):
        return jsonify(found=False)
    conn = db.connect()
    a = conn.execute(
        "SELECT a.id, a.name, a.grade, a.gender, s.name AS sname FROM meet_bibs mb "
        "JOIN athletes a ON a.id=mb.athlete_id JOIN schools s ON s.id=a.school_id "
        "WHERE mb.meet_id=? AND mb.bib=? AND a.active=1", (mid, bib)).fetchone()
    if not a:
        conn.close()
        return jsonify(found=False)
    gword = {"M": "Boys", "F": "Girls"}.get(a["gender"], "Open")
    division = gword + (f" {a['grade']}th" if a["grade"] else "")
    # Already-recorded attempts for THIS event (so the pit doesn't double-enter).
    attempts, best = None, None
    ev = request.args.get("event")
    if ev and str(ev).isdigit():
        me = conn.execute(
            "SELECT id FROM meet_events WHERE meet_id=? AND event_id=? AND gender=? AND grade=?",
            (mid, int(ev), a["gender"] or "", a["grade"])).fetchone()
        if me:
            r = conn.execute(
                "SELECT r.mark_metric, r.mark_seconds, r.attempts_json FROM results r "
                "JOIN entries en ON en.id=r.entry_id WHERE en.meet_event_id=? AND en.runner_id=?",
                (me["id"], a["id"])).fetchone()
            if r:
                if r["attempts_json"]:
                    try:
                        attempts = list(json.loads(r["attempts_json"]))
                    except Exception:  # noqa: BLE001
                        attempts = None
                if r["mark_metric"] is not None:
                    best = _fmt_ht(r["mark_metric"])
                elif r["mark_seconds"] is not None:
                    best = fmt_time(r["mark_seconds"])
    conn.close()
    return jsonify(found=True, name=a["name"], school=a["sname"], division=division,
                   attempts=attempts, best=best)


@bp.post("/meets/<int:mid>/pit/post")
@login_required
def pit_post(mid):
    m = load_meet(mid)
    if not can_record_meet(m) or m["sport"] != "track":
        abort(403)
    d = request.get_json(silent=True) or {}
    try:
        event_id, bib = int(d.get("event_id")), int(str(d.get("bib")).strip())
    except (TypeError, ValueError):
        return jsonify(error="Pick an event and enter a bib number"), 400
    atts = [str(a or "").strip() for a in (d.get("attempts") or [])][:3]
    atts += [""] * (3 - len(atts))
    legal = [v for v in (_parse_ht(a) for a in atts) if v is not None]
    if not any(atts):
        return jsonify(error="Enter at least one attempt"), 400
    conn = db.connect()
    a = conn.execute(
        "SELECT a.*, s.name AS sname FROM meet_bibs mb JOIN athletes a ON a.id=mb.athlete_id "
        "JOIN schools s ON s.id=a.school_id WHERE mb.meet_id=? AND mb.bib=? AND a.active=1",
        (mid, bib)).fetchone()
    if not a:
        conn.close()
        return jsonify(error=f"No athlete with bib {bib} at this meet"), 400
    # The athlete's own division event — created on the fly if it wasn't set up.
    me = conn.execute(
        "SELECT me.id FROM meet_events me WHERE me.meet_id=? AND me.event_id=? "
        "AND me.gender=? AND me.grade=?", (mid, event_id, a["gender"] or "", a["grade"])).fetchone()
    if me:
        meid = me["id"]
    else:
        meid = conn.execute("INSERT INTO meet_events (meet_id, event_id, gender, grade) "
                            "VALUES (?,?,?,?)",
                            (mid, event_id, a["gender"] or "", a["grade"])).lastrowid
    en = conn.execute("SELECT id FROM entries WHERE meet_event_id=? AND runner_id=?",
                      (meid, a["id"])).fetchone()
    eid = en["id"] if en else conn.execute(
        "INSERT INTO entries (meet_event_id, runner_id, school_id) VALUES (?,?,?)",
        (meid, a["id"], a["school_id"])).lastrowid
    best = max(legal) if legal else None
    conn.execute("DELETE FROM results WHERE entry_id=?", (eid,))
    conn.execute("INSERT INTO results (entry_id, mark_metric, attempts_json, dq, snap_name, "
                 "snap_bib, snap_school) VALUES (?,?,?,0,?,?,?)",
                 (eid, best, json.dumps(atts), a["name"], bib, a["sname"]))
    _recompute_places(conn, load_meet_event(meid))
    conn.commit()
    conn.close()
    gword = {"M": "Boys", "F": "Girls"}.get(a["gender"], "Open")
    division = f"{gword}" + (f" {a['grade']}th" if a["grade"] else "")
    evname = ""
    try:
        c2 = db.connect()
        er = c2.execute("SELECT name FROM events WHERE id=?", (event_id,)).fetchone()
        c2.close()
        evname = er["name"] if er else ""
    except Exception:  # noqa: BLE001
        evname = ""
    return jsonify(ok=True, name=a["name"], division=division, event=evname,
                   best=_fmt_ht(best) if best is not None else None)


@bp.get("/meets/<int:mid>/meet-day")
@login_required
def meet_day_page(mid):
    m = load_meet(mid)
    if not can_view_meet(m) or m["sport"] != "track":
        abort(403)
    setup = can_setup_meet(m)
    conn = db.connect()
    mes = conn.execute(
        "SELECT me.*, e.name AS ename, e.kind, e.laned FROM meet_events me JOIN events e ON e.id=me.event_id "
        "WHERE me.meet_id=? ORDER BY e.sort, me.gender, me.grade", (mid,)).fetchall()
    counts = {r[0]: r[1] for r in conn.execute(
        "SELECT me.id, COUNT(en.id) FROM meet_events me LEFT JOIN entries en ON en.meet_event_id=me.id "
        "WHERE me.meet_id=? GROUP BY me.id", (mid,)).fetchall()}
    drawn_counts = {r[0]: r[1] for r in conn.execute(
        "SELECT me.id, COUNT(en.id) FROM meet_events me "
        "LEFT JOIN entries en ON en.meet_event_id=me.id AND en.heat IS NOT NULL "
        "WHERE me.meet_id=? GROUP BY me.id", (mid,)).fetchall()}
    conn.close()

    def div(me):
        return {"M": "Boys", "F": "Girls"}.get(me["gender"], "Open") + (f" {me['grade']}" if me["grade"] else "")

    draw = ""
    if setup:
        draw = f"""
<div class="card"><h2>Draw heats &amp; lanes — whole meet</h2>
<p class="muted">Assigns heats/lanes for every track &amp; relay event at once.</p>
<form method="post" action="/meets/{mid}/draw-all" style="display:inline">
  <input type="hidden" name="mode" value="seeded"><button type="submit">Seed by season best</button></form>
<form method="post" action="/meets/{mid}/draw-all" style="display:inline">
  <input type="hidden" name="mode" value="random"><button class="ghost" type="submit">Random draw</button></form>
</div>"""

    any_entries = any(counts.get(me["id"]) for me in mes)
    packet = ""
    if mes:
        pk = (f'<a class="btn" href="/meets/{mid}/heatsheets.pdf" target="_blank">📄 Download meet packet</a>'
              if any_entries else '<span class="muted">Assign athletes (and draw heats) to generate sheets.</span>')
        packet = (f'<div class="card"><h2>📄 Heat sheets</h2>'
                  f'<p class="muted">One packet of every event with entries, or grab a single '
                  f'event\'s sheet from the list below.</p>{pk}</div>')

    combinable = [me for me in mes if me["kind"] == "relay" or (me["kind"] == "track" and not me["laned"])]
    combine = ""
    if setup and combinable:
        by_cid = {}
        for me in combinable:
            cid = me["combine_id"] if "combine_id" in me.keys() else None
            if cid:
                by_cid.setdefault(cid, []).append(me)
        existing = ""
        for cid, grp in by_cid.items():
            names = ", ".join(f'{g["ename"]} {div(g)}' for g in grp)
            existing += (f'<div style="margin:.3rem 0">🔗 {escape(names)} '
                         f'<form class="inline" method="post" action="/meets/{mid}/uncombine">'
                         f'<input type="hidden" name="combine_id" value="{cid}">'
                         f'<button class="ghost">split</button></form></div>')
        opts = "".join(f'<label style="display:block"><input type="checkbox" name="meids" value="{me["id"]}" '
                       f'style="width:auto"> {escape(me["ename"])} {div(me)}</label>' for me in combinable)
        combine = f"""
<details class="card">
  <summary style="cursor:pointer;font-weight:700;font-size:1.05rem">🔗 Combine races</summary>
  <p class="muted" style="margin-top:.6rem">Run several grade × gender groups of the SAME
  distance/relay event as one physical race. Results still score by grade.</p>{existing}
  <form method="post" action="/meets/{mid}/combine">
    <div class="card" style="background:var(--panel2)">{opts}</div>
    <button type="submit" style="margin-top:.5rem">🔗 Combine selected</button>
  </form>
</details>"""

    # Combined events share a lettered, color-coded chip (🔗 A, 🔗 B…) so you can
    # SEE which ones run together, not just that they're combined with something.
    _cpal = ["#3f8cff", "#3fbf7f", "#e0a83f", "#c968d8", "#f0625b", "#4fc3c8"]
    _cgroups = []
    for me in mes:
        cid = me["combine_id"] if "combine_id" in me.keys() else None
        if cid and cid not in _cgroups:
            _cgroups.append(cid)

    def _chip(cid):
        if not cid:
            return ""
        k = _cgroups.index(cid)
        col = _cpal[k % len(_cpal)]
        return (f' <span style="background:{col}22;color:{col};border:1px solid {col};'
                f'border-radius:999px;padding:.05rem .5rem;font-size:.75rem;font-weight:700;'
                f'white-space:nowrap">🔗 {chr(65 + k)}</span>')

    rows = []
    for i, me in enumerate(mes):
        n = counts.get(me["id"], 0)
        cid = me["combine_id"] if "combine_id" in me.keys() else None
        status = "drawn" if drawn_counts.get(me["id"]) else ("entered" if n else "")
        hs = (f'<a class="btn ghost" href="/meet-events/{me["id"]}/heatsheet.pdf" target="_blank">'
              f'Heat sheet</a>' if n else '<span class="muted">—</span>')
        rows.append(f'<tr data-order="{i}" data-gender="{me["gender"] or ""}" data-grade="{me["grade"] or 0}">'
                    f'<td><a href="/meet-events/{me["id"]}"><b>{escape(me["ename"])}</b></a> '
                    f'<span class="muted">{div(me)}</span>{_chip(cid)}</td>'
                    f'<td>{n} entries</td>'
                    f'<td>{status}</td>'
                    f'<td style="text-align:right">{hs}</td></tr>')
    grades = sorted({me["grade"] for me in mes if me["grade"] is not None})
    grade_opts = '<option value="">All</option>' + "".join(f'<option value="{gd}">{gd}</option>' for gd in grades)
    filterbar = ('<div style="display:flex;justify-content:flex-end;align-items:center;gap:.5rem;'
                 'margin-bottom:.5rem;flex-wrap:wrap"><span class="muted">Filter</span>'
                 '<select id="fgender" onchange="filterEv()" style="max-width:110px">'
                 '<option value="">All</option><option value="M">Boys</option>'
                 '<option value="F">Girls</option></select>'
                 f'<select id="fgrade" onchange="filterEv()" style="max-width:110px">{grade_opts}</select></div>')
    ev_js = ('<script>function filterEv(){var g=document.getElementById("fgender").value,'
             'gr=document.getElementById("fgrade").value,tb=document.getElementById("evbody");'
             '[].forEach.call(tb.querySelectorAll("tr"),function(r){'
             'var ok=(!g||r.dataset.gender===g)&&(!gr||r.dataset.grade===gr);'
             'r.style.display=ok?"":"none";});}</script>')
    ev_tbl = (f'<div class="card"><h2>Events</h2>{filterbar}'
              f'<table><thead><tr><th>Event</th><th>Entries</th><th>Status</th>'
              f'<th style="text-align:right">Heat sheet</th></tr></thead>'
              f'<tbody id="evbody">{"".join(rows)}</tbody></table>{ev_js}</div>'
              if mes else '<div class="card muted">No events yet.</div>')

    pit = (f'<div class="card"><b>🏖 Field pits:</b> '
           f'<a class="btn" href="/meets/{mid}/pit">Open-pit console</a> '
           f'<span class="muted">— record any athlete by bib at the LJ/SP pit; marks file '
           f'into their own division automatically.</span></div>'
           if can_record_meet(m) else "")
    body = (f'<p class="muted"><a href="/meets">← Meets</a></p><h1>{escape(m["name"])}</h1>'
            f'{_track_tabs(mid, "meetday")}{pit}{draw}{packet}{ev_tbl}{combine}')
    return shell(g.principal, body, active="meets")


@bp.post("/meets/<int:mid>/combine")
@login_required
def combine_races(mid):
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    meids = [int(x) for x in request.form.getlist("meids") if x.isdigit()]
    if len(meids) < 2:
        return redirect(f"/meets/{mid}/meet-day")
    conn = db.connect()
    rows = conn.execute(
        f"SELECT me.id, me.event_id, e.kind, e.laned FROM meet_events me JOIN events e ON e.id=me.event_id "
        f"WHERE me.meet_id=? AND me.id IN ({','.join('?' * len(meids))})", (mid, *meids)).fetchall()
    if rows and len({r["event_id"] for r in rows}) == 1 and all(
            r["kind"] == "relay" or not r["laned"] for r in rows):
        nxt = (conn.execute("SELECT COALESCE(MAX(combine_id), 0) FROM meet_events").fetchone()[0]) + 1
        for r in rows:
            conn.execute("UPDATE meet_events SET combine_id=? WHERE id=?", (nxt, r["id"]))
        conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}/meet-day")


@bp.post("/meets/<int:mid>/uncombine")
@login_required
def uncombine_races(mid):
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    cid = request.form.get("combine_id")
    if (cid or "").isdigit():
        conn = db.connect()
        conn.execute("UPDATE meet_events SET combine_id=NULL WHERE meet_id=? AND combine_id=?",
                     (mid, int(cid)))
        conn.commit()
        conn.close()
    return redirect(f"/meets/{mid}/meet-day")


@bp.post("/meets/<int:mid>/draw-all")
@login_required
def draw_all(mid):
    m = load_meet(mid)
    if not can_setup_meet(m):
        abort(403)
    mode = request.form.get("mode", "seeded")
    conn = db.connect()
    mes = conn.execute(
        "SELECT me.*, e.name AS ename, e.kind, e.laned, e.scoring_order, e.sort "
        "FROM meet_events me JOIN events e ON e.id=me.event_id "
        "WHERE me.meet_id=? AND e.kind != 'field'", (mid,)).fetchall()
    done_combines = set()
    for me in mes:
        cid = me["combine_id"] if "combine_id" in me.keys() else None
        if cid:
            if cid in done_combines:
                continue
            done_combines.add(cid)
            ent = conn.execute(
                "SELECT en.* FROM entries en JOIN meet_events me2 ON me2.id=en.meet_event_id "
                "WHERE me2.combine_id=?", (cid,)).fetchall()
            _draw(conn, me, m, mode, entries=ent, laned_override=False)
        else:
            _draw(conn, me, m, mode)
    conn.commit()
    conn.close()
    return redirect(f"/meets/{mid}/meet-day")


# ------------------------------- marks + placing -------------------------------
def _entry_label(conn, e):
    if e["runner_id"]:
        a = conn.execute(
            "SELECT a.name, mb.bib AS bib, s.name AS sname FROM athletes a "
            "JOIN schools s ON s.id=a.school_id "
            "JOIN meet_events me ON me.id=? "
            "LEFT JOIN meet_bibs mb ON mb.athlete_id=a.id AND mb.meet_id=me.meet_id "
            "WHERE a.id=?", (e["meet_event_id"], e["runner_id"])).fetchone()
        if a:
            return a["name"], a["bib"], a["sname"]
        return "—", None, None
    s = conn.execute("SELECT name FROM schools WHERE id=?", (e["school_id"],)).fetchone()
    return f'{s["name"] if s else "?"} {e["relay_label"] or ""}'.strip(), None, s["name"] if s else None


def _recompute_places(conn, me):
    """Rank results for a meet_event and write place numbers."""
    rows = conn.execute(
        "SELECT r.*, e.meet_event_id FROM results r JOIN entries e ON e.id=r.entry_id "
        "WHERE e.meet_event_id=?", (me["id"],)).fetchall()
    scored = []
    for r in rows:
        if r["dq"]:
            continue
        mark = r["mark_seconds"] if me["unit"] == "seconds" else r["mark_metric"]
        if mark is None:
            continue
        scored.append((mark, r["id"]))
    reverse = me["scoring_order"] == "desc"
    scored.sort(key=lambda x: x[0], reverse=reverse)
    # Standard competition ranking with ties: equal marks share the lower place
    # (e.g. 1, 1, 3). Points are split across the tied positions in build_results.
    i = 0
    n = len(scored)
    while i < n:
        j = i
        while j < n and scored[j][0] == scored[i][0]:
            j += 1
        for k in range(i, j):
            conn.execute("UPDATE results SET place=? WHERE id=?", (i + 1, scored[k][1]))
        i = j
    keep = {rid for _, rid in scored}
    for r in rows:
        if r["id"] not in keep:
            conn.execute("UPDATE results SET place=NULL WHERE id=?", (r["id"],))


@bp.post("/meet-events/<int:meid>/marks")
@login_required
def save_marks(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    conn = db.connect()
    meids = _combine_meids(conn, me)
    q = ",".join("?" * len(meids))
    entries = conn.execute(f"SELECT * FROM entries WHERE meet_event_id IN ({q})", tuple(meids)).fetchall()
    hj = _is_hj(me)
    for e in entries:
        eid = e["id"]
        dq = 1 if request.form.get(f"dq_{eid}") else 0
        mark_seconds = mark_metric = None
        attempts = None
        if me["unit"] == "seconds":
            mark_seconds = parse_time(request.form.get(f"mark_{eid}"))
        elif hj:
            mark_metric = parse_metric(request.form.get(f"mark_{eid}"))
        else:  # LJ / SP: three attempts kept verbatim (incl 'F' fouls), best = max legal
            attempts = [(request.form.get(f"a{n}_{eid}") or "").strip() for n in (1, 2, 3)]
            legal = [v for v in (_parse_ht(x) for x in attempts) if v is not None]
            mark_metric = max(legal) if legal else None
        name, bib, school = _entry_label(conn, e)
        conn.execute("DELETE FROM results WHERE entry_id=?", (eid,))
        conn.execute(
            "INSERT INTO results (entry_id, mark_seconds, mark_metric, attempts_json, dq, "
            "snap_name, snap_bib, snap_school) VALUES (?,?,?,?,?,?,?,?)",
            (eid, mark_seconds, mark_metric, json.dumps(attempts) if (attempts and any(attempts)) else None,
             dq, name, bib, school))
    # Rank within each meet-event so combined races still score by grade/gender.
    for m2 in meids:
        _recompute_places(conn, load_meet_event(m2))
    conn.commit()
    conn.close()
    return redirect(f"/meet-events/{meid}")


# ------------------------------- High Jump grid -------------------------------
def _hj_grid_html(meid, me, entries, res, labels, setup, record):
    bars = []
    if me["bar_heights"]:
        try:
            bars = json.loads(me["bar_heights"]) or []
        except (ValueError, TypeError):
            bars = []
    bars = sorted({b for b in (bars or []) if _parse_ht(b) is not None}, key=_parse_ht)
    rows_data = []
    for e in entries:
        name, bib, school = labels[e["id"]]
        r = res.get(e["id"])
        grid = {}
        if r and r["attempts_json"]:
            try:
                g_ = json.loads(r["attempts_json"])
                grid = g_ if isinstance(g_, dict) else {}
            except (ValueError, TypeError):
                grid = {}
        rows_data.append({"eid": e["id"], "name": name, "bib": bib, "school": school or "",
                          "marks": grid, "dq": bool(r and r["dq"]),
                          "best": (_fmt_ht(r["mark_metric"]) if r and r["mark_metric"] is not None else ""),
                          "place": (r["place"] if r and r["place"] else "")})
    add_h = ("" if not record else
             '<div class="hjrow"><b>Bar schedule:</b>'
             '<input id="hjstart" placeholder="Start height (ft-in, e.g. 4-00)" style="max-width:190px">'
             '<input id="hjinc" type="number" min="1" step="1" placeholder="increment (in)" style="max-width:130px">'
             '<button type="button" onclick="genSchedule()">Generate</button></div>'
             '<div class="hjrow"><input id="newht" placeholder="…or add one height (ft-in)" '
             'style="max-width:220px"><button type="button" class="ghost" onclick="addHeight()">+ Add height</button></div>'
             '<p class="muted" style="margin:.2rem 0 .6rem">The meet director sets the opening height and '
             'increment in advance (jr-high openers are usually low-4-foot for boys, high-3/low-4 for girls; '
             'bars often rise 2&Prime; early, then 1&Prime; near the top). Per height: '
             '<b>O</b>=clear, <b>X</b>=miss (XO, XXO), <b>P</b>=pass, <b>XXX</b>=out.</p>')
    add_bib = ("" if not record else
               '<div class="hjrow"><b>Last-minute add by bib:</b> '
               '<input id="addbib" placeholder="Bib #" inputmode="numeric" style="max-width:120px">'
               '<button type="button" onclick="addBib()">+ Add</button></div>')
    note = (f'<p class="muted">➕ Add athletes on the '
            f'<a href="/meets/{me["meet_id"]}/assign">Assign athletes</a> screen (per athlete). '
            f'This grid is for entering results.</p>')
    return f"""
<div class="card" style="overflow-x:auto">
  {note}{add_bib}
  <div class="hjrow"><b>Heights so far</b> <span class="muted">(low → high):</span> <span id="barchips"></span></div>
  {add_h}
  <div id="hjtable"></div>
</div>
<style>
.hjrow{{display:flex;gap:.5rem;align-items:center;flex-wrap:wrap;margin:.5rem 0}}
.barchip{{display:inline-block;background:var(--panel2);border:1px solid var(--line);border-radius:999px;padding:.1rem .55rem;font-size:.8rem;margin:.1rem}}
table.hjgrid{{white-space:nowrap;font-size:.9rem}}
table.hjgrid th,table.hjgrid td{{padding:.35rem .45rem;text-align:center;border-bottom:1px solid var(--line)}}
table.hjgrid td:nth-child(2),table.hjgrid th:nth-child(2){{text-align:left}}
.hjc{{width:54px;text-align:center;text-transform:uppercase}}
.hjgrid .pl{{font-weight:800;color:var(--acc)}}
.hjgrid .best{{font-weight:700;font-variant-numeric:tabular-nums}}
</style>
<script>
const MEID={meid}, CANREC={str(bool(record)).lower()}, SETUP={str(bool(setup)).lower()};
let BARS={json.dumps(bars)};
let ROWS={json.dumps(rows_data)};
function inches(b){{ if(!b)return null; if((''+b).indexOf('-')>=0){{const p=(''+b).split('-'),f=parseInt(p[0]),i=parseInt(p[1]);return (isNaN(f)||isNaN(i))?null:f*12+i;}} const n=parseFloat(b); return isNaN(n)?null:Math.round(n); }}
function fmtht(n){{ if(n==null)return''; n=Math.round(n); return Math.floor(n/12)+'-'+String(n%12).padStart(2,'0'); }}
function bestOf(marks){{ let best=null; BARS.forEach(function(b){{ const mk=(marks[b]||''); if(mk.toUpperCase().indexOf('O')>=0){{const v=inches(b); if(v!=null&&(best==null||v>best))best=v;}} }}); return best; }}
function render(){{
  BARS=[...new Set(BARS)].filter(b=>inches(b)!=null).sort((a,b)=>inches(a)-inches(b));
  document.getElementById('barchips').innerHTML = BARS.map(b=>'<span class="barchip">'+esc(b)+'</span>').join(' ') || '<span class="muted">none yet</span>';
  let h='<table class="hjgrid"><tr><th>PL</th><th>Athlete</th>'+BARS.map(b=>'<th>'+esc(b)+'</th>').join('')+'<th>Best</th><th>DQ</th><th></th></tr>';
  ROWS.forEach(function(r,ri){{
    h+='<tr><td class="pl" id="pl'+ri+'">'+(r.place||'')+'</td>'
      +'<td><b>'+esc(r.name)+'</b>'+(r.bib?' <span class=muted>#'+r.bib+'</span>':'')+(r.school?'<br><span class=muted>'+esc(r.school)+'</span>':'')+'</td>'
      +BARS.map(function(b){{ return '<td><input class="hjc" '+(CANREC?'':'disabled')+' value="'+esc(r.marks[b]||'')+'" data-ri="'+ri+'" data-bar="'+esc(b)+'" onchange="edit(this)"></td>'; }}).join('')
      +'<td class="best" id="best'+ri+'">'+(bestOf(r.marks)!=null?fmtht(bestOf(r.marks)):esc(r.best||''))+'</td>'
      +'<td><input type="checkbox" '+(r.dq?'checked':'')+' '+(CANREC?'':'disabled')+' onchange="dqc('+ri+',this)"></td>'
      +'<td>'+(SETUP?'<button class="danger" onclick="delRow('+r.eid+')">✕</button>':'')+'</td></tr>';
  }});
  h+='</table>';
  document.getElementById('hjtable').innerHTML=h;
}}
function edit(el){{ const ri=+el.dataset.ri,b=el.dataset.bar; el.value=el.value.trim().toUpperCase(); ROWS[ri].marks[b]=el.value;
  document.getElementById('best'+ri).textContent=fmtht(bestOf(ROWS[ri].marks)); save(); }}
function dqc(ri,el){{ ROWS[ri].dq=el.checked; save(); }}
async function save(){{ try{{ const j=await jpost('/meet-events/'+MEID+'/hj-save',{{bars:BARS,rows:ROWS}});
  (j.results||[]).forEach(function(x){{ const r=ROWS.find(rr=>rr.eid==x.eid); if(r)r.place=x.place; }});
  ROWS.forEach(function(r,ri){{ const c=document.getElementById('pl'+ri); if(c)c.textContent=r.place||''; }});
 }}catch(e){{}} }}
function addHeight(){{ const el=document.getElementById('newht'),v=el.value.trim(); if(inches(v)==null){{alert('Use ft-in, e.g. 4-02');return;}} BARS.push(v); el.value=''; render(); save(); }}
function genSchedule(){{
  const s=inches(document.getElementById('hjstart').value.trim());
  const inc=parseFloat(document.getElementById('hjinc').value);
  if(s==null){{alert('Enter a start height in ft-in, e.g. 4-00');return;}}
  if(isNaN(inc)||inc<=0){{alert('Enter an increment in inches, e.g. 2');return;}}
  const n=Math.max(8, Math.min(12, Math.round(18/inc)+1));   // ~1.5 ft of bars above the opener
  BARS=[]; for(let k=0;k<n;k++) BARS.push(fmtht(s+Math.round(k*inc)));
  render(); save();
}}
async function addBib(){{ const el=document.getElementById('addbib'),v=el.value.trim(); if(!v)return;
  try{{ await jpost('/meet-events/'+MEID+'/add-bib',{{bib:v}}); location.reload(); }}catch(e){{ alert(e.message); }} }}
async function delRow(eid){{ if(!confirm('Remove this athlete from the event?'))return;
  await fetch('/entries/'+eid+'/delete',{{method:'POST'}}); location.reload(); }}
render();
</script>"""


@bp.post("/meet-events/<int:meid>/hj-save")
@login_required
def hj_save(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    data = request.get_json(silent=True) or {}
    bars = sorted({b for b in (data.get("bars") or []) if _parse_ht(b) is not None}, key=_parse_ht)
    conn = db.connect()
    conn.execute("UPDATE meet_events SET bar_heights=? WHERE id=?", (json.dumps(bars), meid))
    valid = {e["id"]: e for e in conn.execute("SELECT * FROM entries WHERE meet_event_id=?",
                                               (meid,)).fetchall()}
    for row in (data.get("rows") or []):
        try:
            eid = int(row.get("eid"))
        except (TypeError, ValueError):
            continue
        e = valid.get(eid)
        if not e:
            continue
        marks = row.get("marks") or {}
        grid = {b: str(marks.get(b, "")).strip().upper() for b in bars if str(marks.get(b, "")).strip()}
        name, bib, school = _entry_label(conn, e)
        dq = 1 if row.get("dq") else 0
        if grid:                          # grid cells present -> best comes from the grid
            conn.execute("DELETE FROM results WHERE entry_id=?", (eid,))
            conn.execute("INSERT INTO results (entry_id, mark_metric, attempts_json, dq, snap_name, "
                         "snap_bib, snap_school) VALUES (?,?,?,?,?,?,?)",
                         (eid, _hj_best(grid), json.dumps(grid), dq, name, bib, school))
        else:                             # no grid marks -> keep any scanned result, sync DQ only
            existing = conn.execute("SELECT id FROM results WHERE entry_id=?", (eid,)).fetchone()
            if existing:
                conn.execute("UPDATE results SET dq=? WHERE entry_id=?", (dq, eid))
            elif dq:
                conn.execute("INSERT INTO results (entry_id, mark_metric, attempts_json, dq, "
                             "snap_name, snap_bib, snap_school) VALUES (?,NULL,?,1,?,?,?)",
                             (eid, json.dumps({}), name, bib, school))
    _recompute_places(conn, load_meet_event(meid))
    out = [{"eid": r["entry_id"], "place": r["place"], "best": _fmt_ht(r["mark_metric"])}
           for r in conn.execute("SELECT r.entry_id, r.place, r.mark_metric FROM results r "
                                 "JOIN entries e ON e.id=r.entry_id WHERE e.meet_event_id=?",
                                 (meid,)).fetchall()]
    conn.commit()
    conn.close()
    return jsonify(ok=True, results=out)


@bp.post("/meet-events/<int:meid>/add-bib")
@login_required
def add_bib(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    raw = (request.get_json(silent=True) or {}).get("bib")
    try:
        bib = int(str(raw).strip())
    except (TypeError, ValueError):
        return jsonify(error="Enter a bib number"), 400
    conn = db.connect()
    a = conn.execute("SELECT a.id, a.school_id, a.name FROM meet_bibs mb "
                     "JOIN athletes a ON a.id=mb.athlete_id "
                     "WHERE mb.meet_id=? AND mb.bib=? LIMIT 1", (me["meet_id"], bib)).fetchone()
    if not a:
        conn.close()
        return jsonify(error=f"No athlete with bib {bib} at this meet"), 400
    exists = conn.execute("SELECT 1 FROM entries WHERE meet_event_id=? AND runner_id=?",
                          (meid, a["id"])).fetchone()
    if not exists:
        conn.execute("INSERT INTO entries (meet_event_id, runner_id, school_id) VALUES (?,?,?)",
                     (meid, a["id"], a["school_id"]))
        conn.commit()
    conn.close()
    return jsonify(ok=True, name=a["name"])


# ------------------------------- event page -------------------------------
@bp.get("/meet-events/<int:meid>")
@login_required
def event_page(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_view_meet(m):
        abort(403)
    record = can_record_meet(m)
    setup = can_setup_meet(m)
    hj = _is_hj(me)
    conn = db.connect()
    meids = _combine_meids(conn, me)
    combined = len(meids) > 1
    qm = ",".join("?" * len(meids))
    entries = conn.execute(
        f"SELECT * FROM entries WHERE meet_event_id IN ({qm}) ORDER BY heat, lane, id",
        tuple(meids)).fetchall()
    res = {r["entry_id"]: r for r in conn.execute(
        f"SELECT r.* FROM results r JOIN entries e ON e.id=r.entry_id WHERE e.meet_event_id IN ({qm})",
        tuple(meids)).fetchall()}
    labels = {e["id"]: _entry_label(conn, e) for e in entries}
    # division label per meet-event (for combined display)
    me_div = {r["id"]: ({"M": "B", "F": "G"}.get(r["gender"], "") + (str(r["grade"]) if r["grade"] else ""))
              for r in conn.execute(f"SELECT id, gender, grade FROM meet_events WHERE id IN ({qm})",
                                    tuple(meids)).fetchall()}
    athletes = _attending_athletes(conn, me["meet_id"], me["gender"], me["grade"]) \
        if me["kind"] != "relay" else []
    schools = _attending_schools(conn, me["meet_id"]) if me["kind"] == "relay" else []
    conn.close()

    div = {"M": "Boys", "F": "Girls"}.get(me["gender"], "Open")
    ename = f'{me["ename"]} — {div}' + (f' {me["grade"]}' if me["grade"] else "")
    if combined:
        ename += " · 🔗 combined"
    _errs = {"limit": "That athlete has hit the meet event limit.",
             "nobib": "No athlete with that bib is at this meet.",
             "bib": "Enter a bib number to add an athlete."}
    err = (f'<div class="msg err">{_errs[request.args["err"]]}</div>'
           if request.args.get("err") in _errs else "")

    # Entries + marks form
    def mark_cell(e):
        r = res.get(e["id"])
        if me["unit"] == "seconds":
            v = fmt_time(r["mark_seconds"]) if r and r["mark_seconds"] is not None else ""
            return f'<input name="mark_{e["id"]}" value="{v}" placeholder="mm:ss.t / s.t" style="width:110px">'
        if hj:
            v = _fmt_ht(r["mark_metric"]) if r and r["mark_metric"] is not None else ""
            return f'<input name="mark_{e["id"]}" value="{v}" placeholder="5-02" style="width:90px">'
        atts = json.loads(r["attempts_json"]) if r and r["attempts_json"] else ["", "", ""]

        def _av(x):
            return _fmt_ht(x) if isinstance(x, (int, float)) else (x or "")   # old data was numeric
        cells = "".join(
            f'<input name="a{n}_{e["id"]}" value="{escape(_av(atts[n-1]) if n-1 < len(atts) else "")}" '
            f'placeholder="A{n} · 15-06" style="width:82px">' for n in (1, 2, 3))
        return cells

    rows = []
    for e in entries:
        name, bib, school = labels[e["id"]]
        r = res.get(e["id"])
        place = r["place"] if r and r["place"] else ""
        dqc = f'<input type="checkbox" name="dq_{e["id"]}" style="width:auto" {"checked" if r and r["dq"] else ""}>'
        hl = f'{e["heat"] or ""}' + (f'/{e["lane"]}' if e["lane"] else "")
        dtag = (f' <span class="pill">{me_div.get(e["meet_event_id"], "")}</span>'
                if combined else "")
        delc = (f'<form class="inline" method="post" action="/entries/{e["id"]}/delete">'
                f'<button class="danger">✕</button></form>' if setup else "")
        rows.append(
            f'<tr><td>{place}</td><td class="muted">{hl}</td>'
            f'<td><b>{escape(name)}</b>{f" #{bib}" if bib else ""}{dtag}<br>'
            f'<span class="muted">{escape(school or "")}</span></td>'
            f'<td>{mark_cell(e)}</td><td>{dqc}</td><td>{delc}</td></tr>')
    marks_form = ""
    if entries:
        inner = (f'<table><tr><th>Pl</th><th>Ht/Ln</th><th>Competitor</th>'
                 f'<th>{"Attempts (best counts)" if me["kind"]!="relay" and me["unit"]=="metric" and not hj else ("Height" if hj else "Mark/Time")}</th>'
                 f'<th>DQ</th><th></th></tr>{"".join(rows)}</table>')
        if record:
            marks_form = (f'<form method="post" action="/meet-events/{meid}/marks">'
                          f'<div class="card">{inner}'
                          f'<button type="submit" style="margin-top:1rem">Save marks &amp; re-rank</button>'
                          f'</div></form>')
        else:
            marks_form = f'<div class="card">{inner}</div>'
    else:
        marks_form = '<div class="card muted">No entries yet.</div>'

    if hj:   # interactive make/miss grid replaces the single-mark table
        marks_form = _hj_grid_html(meid, me, entries, res, labels, setup, record)

    # Add-entry form
    add = ""
    if setup:
        if me["kind"] == "relay":
            sopts = "".join(f'<option value="{s["id"]}">{escape(s["name"])}</option>' for s in schools)
            add = f"""
<div class="card"><h2>Add relay squad</h2>
<form method="post" action="/meet-events/{meid}/entries" class="row">
  <div><label>School</label><select name="school_id">{sopts}</select></div>
  <div style="max-width:90px"><label>Squad</label><input name="relay_label" value="A"></div>
  <div><label>Members (comma-sep)</label><input name="members" placeholder="Lee, Cho, Ray, Doe"></div>
  <div style="display:flex;align-items:flex-end"><button type="submit">Add</button></div>
</form></div>"""
        else:
            add = f"""
<div class="card"><h2>Add athlete</h2>
<form method="post" action="/meet-events/{meid}/entries" class="row" style="flex-wrap:wrap">
  <div style="max-width:180px"><label>Bib number</label>
    <input name="bib" type="number" inputmode="numeric" autofocus></div>
  <div style="display:flex;align-items:flex-end"><button type="submit">Add</button></div>
  <label style="display:flex;gap:.4rem;align-items:center;font-size:.9rem;margin-top:1.4rem">
    <input type="checkbox" name="force" style="width:auto"> Meet-day sub — override event limit</label>
</form>
<p class="muted">Enter the bib to add that athlete ({div}) to this event.</p></div>"""

    # Seed / sheets (setup) + camera scan-back (record)
    seed_part = ""
    if setup:
        seed_part = f"""
<div class="row">
  <form class="inline" method="post" action="/meet-events/{meid}/seed">
    <input type="hidden" name="mode" value="seeded"><button type="submit">Seed by season best</button></form>
  <form class="inline" method="post" action="/meet-events/{meid}/seed">
    <input type="hidden" name="mode" value="random"><button class="ghost" type="submit">Random draw</button></form>
  <form class="inline" method="post" action="/meet-events/{meid}/clear-draw"
    onsubmit="return confirm('Clear heats/lanes for this event (entries stay)?')">
    <button class="ghost" type="submit">Clear draw</button></form>
  <a class="btn ghost" href="/meet-events/{meid}/heatsheet.pdf">Heat sheet (PDF)</a>
</div>"""
    scan_part = ""
    if record:
        scan_part = f"""
<div style="margin-top:1rem"><label>📷 Scan a filled sheet — reads the handwritten marks with AI</label>
  <input type="file" id="scanf" accept="image/*" capture="environment">
  <button type="button" onclick="scan()" style="margin-top:.5rem">Read marks</button>
  <div id="scanout"></div>
</div>"""
    tools = ""
    if seed_part or scan_part:
        tools = f"""
<div class="card"><h2>Seed, sheets &amp; scan</h2>{seed_part}{scan_part}</div>
<script>
async function scan(){{
  const f=document.getElementById('scanf').files[0];
  if(!f){{alert('Take or choose a photo');return;}}
  document.getElementById('scanout').innerHTML='<p class="muted">Reading…</p>';
  const fd=new FormData(); fd.append('image',f);
  const r=await fetch('/meet-events/{meid}/scan',{{method:'POST',body:fd}});
  const j=await r.json();
  if(!r.ok){{document.getElementById('scanout').innerHTML='<p class="msg err">'+esc(j.error||'Failed')+'</p>';return;}}
  if(!j.marks||!j.marks.length){{document.getElementById('scanout').innerHTML='<p class="msg err">No marks read.</p>';return;}}
  window.SCANFIELD=!!j.field; window.SCANHJ=!!j.hj;
  let h='<p class="muted">Review &amp; edit, then post. Matches to athletes by bib; unknown bibs at this meet are entered automatically.</p><table>';
  const dqtd=i=>'<td><input type="checkbox" id="sd'+i+'" style="width:auto"></td>';
  if(SCANHJ){{
    h+='<tr><th>Bib</th><th>Best height</th><th>Misses</th><th>DQ</th></tr>';
    j.marks.forEach((m,i)=>{{ h+='<tr><td><input id="sb'+i+'" value="'+esc(m.bib==null?'':m.bib)+'" style="width:64px"></td>'
      +'<td><input id="sh'+i+'" value="'+esc(m.height==null?'':m.height)+'" placeholder="4-08" style="width:100px"></td>'
      +'<td><input id="sx'+i+'" value="'+esc(m.misses==null?'':m.misses)+'" style="width:56px"></td>'+dqtd(i)+'</tr>'; }});
  }} else if(SCANFIELD){{
    h+='<tr><th>Bib</th><th>A1</th><th>A2</th><th>A3</th><th>DQ</th></tr>';
    j.marks.forEach((m,i)=>{{ const a=m.attempts||['','',''];
      h+='<tr><td><input id="sb'+i+'" value="'+esc(m.bib==null?'':m.bib)+'" style="width:64px"></td>'
        +[0,1,2].map(k=>'<td><input id="sa'+i+'_'+k+'" value="'+esc(a[k]||'')+'" style="width:74px"></td>').join('')+dqtd(i)+'</tr>'; }});
  }} else {{
    h+='<tr><th>Bib</th><th>Mark</th><th>DQ</th></tr>';
    j.marks.forEach((m,i)=>{{ h+='<tr><td><input id="sb'+i+'" value="'+esc(m.bib==null?'':m.bib)+'" style="width:70px"></td>'
      +'<td><input id="sm'+i+'" value="'+esc(m.mark==null?'':m.mark)+'" style="width:120px"></td>'+dqtd(i)+'</tr>'; }});
  }}
  h+='</table><button type="button" onclick="postScan('+j.marks.length+')" style="margin-top:.6rem">Post marks</button>';
  document.getElementById('scanout').innerHTML=h;
}}
async function postScan(n){{
  const marks=[];
  const dqv=i=>{{const el=document.getElementById('sd'+i);return !!(el&&el.checked);}};
  for(let i=0;i<n;i++){{ const b=document.getElementById('sb'+i).value.trim(); if(!b)continue;
    if(SCANHJ){{ const ht=document.getElementById('sh'+i).value.trim(); const mi=document.getElementById('sx'+i).value.trim();
      if(ht) marks.push({{bib:b,height:ht,misses:mi?parseInt(mi):null,dq:dqv(i)}}); }}
    else if(SCANFIELD){{ const a=[0,1,2].map(k=>document.getElementById('sa'+i+'_'+k).value.trim());
      if(a.some(x=>x)) marks.push({{bib:b,attempts:a,dq:dqv(i)}}); }}
    else {{ const mk=document.getElementById('sm'+i).value.trim(); if(mk) marks.push({{bib:b,mark:mk,dq:dqv(i)}}); }}
  }}
  try{{ const j=await jpost('/meet-events/{meid}/scan/post',{{marks}});
    let msg='Posted '+j.applied+' marks';
    if(j.added&&j.added.length) msg+='; added last-minute: '+j.added.join(', ');
    if(j.unmatched&&j.unmatched.length) msg+='; unknown bibs: '+j.unmatched.join(', ');
    alert(msg); location.reload(); }}
  catch(e){{ alert(e.message); }}
}}
</script>"""

    field_note = ""
    if me["unit"] == "metric":
        ex = ('the best height cleared like <code>4-08</code> or <code>5&#39;2&quot;</code>' if hj
              else 'marks like <code>15-06</code>, <code>5-03</code>, or <code>5&#39;3&quot;</code>')
        field_note = f'<p class="muted">Enter <b>feet-inches</b> — {ex}.</p>'
    body = (f'<p class="muted"><a href="/meets/{me["meet_id"]}/meet-day">← Meet day</a></p>'
            f'<h1>{escape(ename)}</h1>{err}{field_note}{marks_form}{add}{tools}')
    return shell(g.principal, body, active="meets")


# ------------------------------- vision scan-back -------------------------------
def _resolve_scanned_event(conn, scanned_bibs, code_meid):
    """Which meet_event is this sheet for? Prefer the event whose entrants actually
    match the scanned (PRINTED) bibs — deterministic and immune to OCR misreads of the
    printed code — and fall back to the code only when the bibs are inconclusive.
    Returns (meid, how) or (None, None)."""
    best = None  # (meid, match_count)
    if scanned_bibs:
        q = ",".join("?" * len(scanned_bibs))
        rows = conn.execute(
            f"SELECT me.id AS meid, me.meet_id, COUNT(DISTINCT a.bib) AS n "
            f"FROM entries en JOIN athletes a ON a.id=en.runner_id "
            f"JOIN meet_events me ON me.id=en.meet_event_id "
            f"WHERE a.bib IN ({q}) GROUP BY me.id ORDER BY n DESC", tuple(scanned_bibs)).fetchall()
        for r in rows:
            if can_record_meet(load_meet(r["meet_id"])):
                best = (r["meid"], r["n"])
                break
    # A strong bib match wins outright (survives a misread sheet code).
    if best and best[1] >= max(2, (len(scanned_bibs) + 1) // 2):
        return best[0], "bibs"
    # Otherwise trust the printed code if it points to a recordable event.
    if code_meid:
        r = conn.execute("SELECT meet_id FROM meet_events WHERE id=?", (code_meid,)).fetchone()
        if r and can_record_meet(load_meet(r["meet_id"])):
            return code_meid, "code"
    if best:                       # last resort: a weak bib match beats nothing
        return best[0], "bibs"
    return None, None


@bp.post("/track/scan")
@login_required
def scan_auto():
    """Scan a heat sheet WITHOUT choosing an event — resolve which event it is from the
    printed bibs (falling back to the sheet code) and return its marks."""
    import re
    f = request.files.get("image")
    if not f:
        return jsonify(error="No image"), 400
    media = f.mimetype or "image/jpeg"
    data = f.read()
    try:
        res = ai.vision_read_sheet(data, media_type=media)
    except Exception as e:  # noqa: BLE001
        return jsonify(error=f"Vision read failed: {e}"), 400
    code = res.get("sheet_code") or ""
    mt = re.search(r"E\s*0*(\d+)", code.upper())
    code_meid = int(mt.group(1)) if mt else None
    scanned_bibs = set()
    for mk in res.get("marks", []):
        try:
            scanned_bibs.add(int(str(mk.get("bib")).strip()))
        except (TypeError, ValueError):
            pass
    conn = db.connect()
    meid, _how = _resolve_scanned_event(conn, scanned_bibs, code_meid)
    me = conn.execute("SELECT * FROM meet_events WHERE id=?", (meid,)).fetchone() if meid else None
    conn.close()
    if not me:
        return jsonify(error="Couldn't tell which event this sheet is for — the code didn't "
                             "read and no bib numbers matched a meet you're recording. Retake "
                             "the photo with the whole sheet (and its top-right code) in frame."), 400
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    me = load_meet_event(meid)
    gword = {"M": "Boys", "F": "Girls"}.get(me["gender"], "")
    gr = f'{me["grade"]}th Grade ' if me["grade"] else ""
    label = f'{gr}{gword + " " if gword else ""}{me["ename"]}'.strip()
    hj = _is_hj(me)
    field = me["unit"] == "metric" and not hj           # LJ / SP -> read all 3 attempts
    try:
        if hj:                                          # High Jump -> best height per athlete
            rows = ai.vision_read_hj(data, media_type=media)["rows"]
            marks = [{"bib": r["bib"], "name": r["name"], "height": r["height"],
                      "misses": r["misses"]} for r in rows]
        elif field:
            rows = ai.vision_read_field(data, media_type=media)["rows"]
            marks = [{"bib": r["bib"], "name": r["name"], "attempts": r["attempts"]} for r in rows]
        else:
            marks = res.get("marks", [])
    except Exception as e:  # noqa: BLE001
        return jsonify(error=f"Vision read failed: {e}"), 400
    return jsonify(meid=meid, label=label, meet=m["name"], marks=marks, field=field, hj=hj)


@bp.post("/meet-events/<int:meid>/scan")
@login_required
def scan_back(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    f = request.files.get("image")
    if not f:
        return jsonify(error="No image"), 400
    media = f.mimetype or "image/jpeg"
    hj = _is_hj(me)
    field = me["unit"] == "metric" and not hj           # LJ / SP -> 3 attempts each
    data = f.read()
    try:
        if hj:
            rows = ai.vision_read_hj(data, media_type=media)["rows"]
            marks = [{"bib": r["bib"], "name": r["name"], "height": r["height"],
                      "misses": r["misses"]} for r in rows]
        elif field:
            rows = ai.vision_read_field(data, media_type=media)["rows"]
            marks = [{"bib": r["bib"], "name": r["name"], "attempts": r["attempts"]} for r in rows]
        else:
            marks = ai.vision_read_marks(data, media_type=media)
    except Exception as e:  # noqa: BLE001
        return jsonify(error=f"Vision read failed: {e}"), 400
    return jsonify(marks=marks, field=field, hj=hj)


@bp.post("/meet-events/<int:meid>/scan/post")
@login_required
def scan_post(meid):
    """Apply reviewed scan marks: match each bib to its entry, write the mark, re-rank."""
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    marks = (request.get_json(silent=True) or {}).get("marks", [])
    conn = db.connect()
    meids = _combine_meids(conn, me)
    qm = ",".join("?" * len(meids))
    # bib -> entry (individual entries in this event/group)
    entrymap = {}
    for r in conn.execute(
        f"SELECT en.id AS eid, a.bib, a.name, s.name AS sname FROM entries en "
        f"JOIN athletes a ON a.id=en.runner_id JOIN schools s ON s.id=a.school_id "
        f"WHERE en.meet_event_id IN ({qm}) AND a.bib IS NOT NULL", tuple(meids)).fetchall():
        entrymap[r["bib"]] = r
    applied, unmatched, added = 0, [], []
    for mk in marks:
        try:
            bib = int(str(mk.get("bib")).strip())
        except (TypeError, ValueError):
            continue
        row = entrymap.get(bib)
        if not row:
            # Last-minute athlete on the sheet but not entered: if that bib belongs to
            # an athlete at this meet, enter them on the spot (a real sheet is truth).
            a = conn.execute(
                "SELECT a.id, a.name, a.bib, s.name AS sname, a.school_id FROM athletes a "
                "JOIN schools s ON s.id=a.school_id JOIN meet_schools ms ON ms.school_id=s.id "
                "WHERE ms.meet_id=? AND a.bib=? AND a.active=1", (m["id"], bib)).fetchone()
            if not a:
                unmatched.append(bib)
                continue
            eid = conn.execute("INSERT INTO entries (meet_event_id, runner_id, school_id) "
                               "VALUES (?,?,?)", (me["id"], a["id"], a["school_id"])).lastrowid
            row = {"eid": eid, "bib": bib, "name": a["name"], "sname": a["sname"]}
            entrymap[bib] = row
            added.append(bib)
        dq = 1 if mk.get("dq") else 0
        if "height" in mk:   # High Jump: single best cleared height (+ optional misses)
            met = _parse_ht(str(mk.get("height", "")).strip())
            if met is None:
                continue
            miss = mk.get("misses")
            try:
                miss = int(miss) if (miss is not None and str(miss).strip() != "") else None
            except (TypeError, ValueError):
                miss = None
            aj = json.dumps({"misses": miss}) if miss is not None else None
            conn.execute("DELETE FROM results WHERE entry_id=?", (row["eid"],))
            conn.execute(
                "INSERT INTO results (entry_id, mark_metric, attempts_json, dq, "
                "snap_name, snap_bib, snap_school) VALUES (?,?,?,?,?,?,?)",
                (row["eid"], met, aj, dq, row["name"], bib, row["sname"]))
            applied += 1
            continue
        if "attempts" in mk and me["unit"] != "seconds":   # LJ / SP: keep all 3, verbatim
            atts = [str(a or "").strip() for a in (mk.get("attempts") or [])][:3]
            atts += [""] * (3 - len(atts))
            legal = [v for v in (_parse_ht(a) for a in atts) if v is not None]
            if not any(atts):
                continue
            best = max(legal) if legal else None
            conn.execute("DELETE FROM results WHERE entry_id=?", (row["eid"],))
            conn.execute(
                "INSERT INTO results (entry_id, mark_metric, attempts_json, dq, "
                "snap_name, snap_bib, snap_school) VALUES (?,?,?,?,?,?,?)",
                (row["eid"], best, json.dumps(atts), dq, row["name"], bib, row["sname"]))
            applied += 1
            continue
        raw = str(mk.get("mark", "")).strip()
        if me["unit"] == "seconds":
            sec, met = parse_time(raw), None
        else:
            sec, met = None, _parse_ht(raw)   # field, single mark
        if sec is None and met is None:
            continue
        conn.execute("DELETE FROM results WHERE entry_id=?", (row["eid"],))
        conn.execute(
            "INSERT INTO results (entry_id, mark_seconds, mark_metric, dq, "
            "snap_name, snap_bib, snap_school) VALUES (?,?,?,?,?,?,?)",
            (row["eid"], sec, met, dq, row["name"], bib, row["sname"]))
        applied += 1
    for m2 in meids:
        _recompute_places(conn, load_meet_event(m2))
    conn.commit()
    conn.close()
    return jsonify(applied=applied, unmatched=unmatched, added=added)


# ------------------------------- live tap timer (distance / relay) -------------------------------
# Server-backed so a heat can be timed on one device and assigned on another,
# exactly like the XC race console. State lives in track_clocks + track_taps,
# keyed by (meet_event, heat). Assigning a tap writes the result immediately so
# every device — and the results page — stays in sync.
def _tap_entries(conn, meids, hk):
    """Assignable runners/squads for this heat (all entries when hk==0)."""
    qm = ",".join("?" * len(meids))
    if hk:
        rows = conn.execute(f"SELECT * FROM entries WHERE meet_event_id IN ({qm}) AND heat=? "
                            f"ORDER BY lane, id", (*meids, hk)).fetchall()
    else:
        rows = conn.execute(f"SELECT * FROM entries WHERE meet_event_id IN ({qm}) "
                            f"ORDER BY heat, lane, id", tuple(meids)).fetchall()
    out = []
    for e in rows:
        name, bib, school = _entry_label(conn, e)
        lbl = name + (f" #{bib}" if bib else "") + (f" · {school}" if school and not bib else "")
        out.append({"id": e["id"], "label": lbl})
    return out


@bp.get("/meet-events/<int:meid>/time")
@login_required
def time_console(meid):
    """Multi-device tap timer for a distance/relay heat: start the clock, tap each
    finisher, assign each to a runner/squad. State is shared across devices."""
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    heat = request.args.get("heat", "")
    hk = _heat_key(heat)
    div = {"M": "Boys", "F": "Girls"}.get(me["gender"], "Open")
    title = f'{me["ename"]} — {div}' + (f' {me["grade"]}' if me["grade"] else "")
    sub = (f"Heat {heat}" if heat else "All entries") + " · Start, then tap each finisher"
    allow_bib = "true" if me["kind"] != "relay" else "false"   # bibs = individuals only
    from .xc import CONSOLE_CSS  # share the XC timing-console look
    body = f"""
<style>{CONSOLE_CSS}</style>
<p class="muted"><a href="/phone/meet/{me['meet_id']}">← Track Timer</a></p>
<h1>{escape(title)} <span class="muted" style="font-weight:400">· tap timer</span></h1>
<p class="sub">{escape(sub)} <span class="muted">· shared across devices</span></p>
<div class="card">
  <div id="clock" class="tc-clock">0:00:00.000</div>
  <div class="tc-btns">
    <button id="btn-start" onclick="startRace()">🚦 Start</button>
    <button id="btn-stop" onclick="stopRace()">⏹ Stop</button>
    <button class="ghost" onclick="resetRace()">🔄 Reset</button>
  </div>
  <div id="status" class="tc-status wait">Loading…</div>
  <button id="tapbtn" onclick="tap()" disabled
    style="font-size:1.7rem;padding:1.1rem;width:100%;max-width:440px;margin:.7rem auto 0;display:block">TAP finisher</button>
</div>
<div class="card"><h2>Finishers (<span id="cnt">0</span>)</h2>
  <table id="rows"></table>
  <button onclick="doneRace()" style="margin-top:1.1rem;width:100%">✅ Done — back to heats</button>
</div>
<script>
const RID={meid}, HEAT={hk}, ALLOW_BIB={allow_bib};
let OFFSET=0, START=null, STOPMS=null, STARTED=false, STOPPED=false, TAPS=[], ENTS=[], BUSY=false, PICKING=false;
function nowms(){{ return Date.now()+OFFSET; }}
function fmt(sec){{ if(sec==null)return''; sec=Math.max(0,sec);
  const h=Math.floor(sec/3600), m=Math.floor((sec%3600)/60), s=sec-3600*h-60*m;
  return h+':'+String(m).padStart(2,'0')+':'+s.toFixed(3).padStart(6,'0'); }}
function tick(){{ const c=document.getElementById('clock');
  if(!STARTED){{ c.textContent='0:00:00.000'; c.classList.remove('stopped'); return; }}
  const end=STOPMS||nowms(); c.textContent=fmt((end-START)/1000); c.classList.toggle('stopped',STOPPED); }}
function syncUI(){{
  document.getElementById('btn-start').disabled = STARTED && !STOPPED;
  document.getElementById('btn-stop').disabled = !STARTED || STOPPED;
  document.getElementById('tapbtn').disabled = !STARTED || STOPPED;
  const st=document.getElementById('status');
  if(!STARTED){{ st.className='tc-status wait'; st.textContent='Not started.'; }}
  else if(STOPPED){{ st.className='tc-status end'; st.textContent='🏁 Race ended — assign each finisher.'; }}
  else {{ st.className='tc-status run'; st.textContent='🟢 Running — tap each runner as they cross.'; }}
}}
function esc2(s){{ return String(s).replace(/[&<>"]/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}}[c])); }}
function render(){{
  document.getElementById('cnt').textContent=TAPS.length;
  const taken={{}}; TAPS.forEach(t=>{{ if(t.entry_id) taken[t.entry_id]=t.id; }});
  const rows=document.getElementById('rows');
  if(!TAPS.length){{ rows.innerHTML='<tr><td class="muted">No finishers yet — tap as runners cross.</td></tr>'; return; }}
  let h='<tr><th>#</th><th>Time</th><th>Runner</th><th></th></tr>';
  TAPS.forEach(function(t,i){{
    let opts='<option value="">— pick —</option>';
    ENTS.forEach(function(e){{
      const mine=t.entry_id==e.id, used=(taken[e.id]&&taken[e.id]!=t.id);
      if(used) return;                       // already assigned elsewhere -> drop from this picker
      opts+='<option value="'+e.id+'"'+(mine?' selected':'')+'>'+esc2(e.label)+'</option>';
    }});
    if(ALLOW_BIB) opts+='<option value="__bib__">➕ Bib not listed…</option>';
    h+='<tr><td>'+(i+1)+'</td><td style="font-variant-numeric:tabular-nums">'+fmt(t.elapsed)+'</td>'
      +'<td><select onfocus="PICKING=true" onblur="setTimeout(function(){{PICKING=false}},350)" onchange="assign('+t.id+',this.value)">'+opts+'</select></td>'
      +'<td style="text-align:right"><button class="danger" onclick="delTap('+t.id+')">\\u2715</button></td></tr>';
  }});
  rows.innerHTML=h;
}}
async function load(){{
  if(BUSY||PICKING)return;                 // never rebuild rows while a picker is open
  try{{
    const s=await jget('/meet-events/'+RID+'/time/state?heat='+HEAT);
    OFFSET=s.server_ms-Date.now(); START=s.start_ms; STOPMS=s.stop_ms;
    STARTED=s.started; STOPPED=s.stopped; TAPS=s.taps; ENTS=s.entries;
    syncUI(); if(!PICKING) render();
  }}catch(e){{}}
}}
async function act(url, body){{
  BUSY=true;
  try{{ await jpost(url, body||{{}}); }}
  catch(e){{ if(e&&e.message) alert(e.message); }}
  finally{{ BUSY=false; }}
  await load();
}}
async function startRace(){{
  if(STARTED&&!STOPPED)return;
  const body={{}};
  if(STOPPED&&TAPS.length){{ if(!confirm('This heat already has an ended race with '+TAPS.length+' tap(s). Start over and CLEAR them?'))return; body.clear=true; }}
  act('/meet-events/'+RID+'/time/start?heat='+HEAT, body);
}}
function stopRace(){{ if(STARTED&&!STOPPED) act('/meet-events/'+RID+'/time/stop?heat='+HEAT); }}
function resetRace(){{ if(confirm('Reset clears the clock, all taps, and any results for this heat. Continue?'))
  act('/meet-events/'+RID+'/time/reset?heat='+HEAT); }}
function buzz(){{ try{{ navigator.vibrate && navigator.vibrate(35); }}catch(e){{}} }}
function tap(){{ if(!STARTED||STOPPED)return; buzz(); act('/meet-events/'+RID+'/time/tap?heat='+HEAT); }}
function assign(tid,v){{
  PICKING=false;
  if(v==='__bib__'){{
    const b=prompt('Runner not in the list — enter their bib number:');
    if(!b||!b.trim()){{ load(); return; }}   // cancelled -> restore the row
    act('/track-taps/'+tid+'/assign-bib', {{bib:b.trim()}});
    return;
  }}
  act('/track-taps/'+tid+'/assign', {{entry_id:v}});
}}
function delTap(tid){{ if(confirm('Remove this finish?')) act('/track-taps/'+tid+'/delete'); }}
async function doneRace(){{
  if(STARTED&&!STOPPED){{ if(!confirm('The clock is still running. Stop the race and return to the heats menu?'))return;
    try{{ await jpost('/meet-events/'+RID+'/time/stop?heat='+HEAT, {{}}); }}catch(e){{}} }}
  location.href='/phone/meet/{me['meet_id']}';
}}
let WL=null; async function wlock(){{ try{{ WL=await navigator.wakeLock.request('screen'); }}catch(e){{}} }}
document.addEventListener('visibilitychange',()=>{{ if(document.visibilityState==='visible'){{ wlock(); load(); }} }});
wlock();
setInterval(tick,75);
setInterval(()=>{{ if(!document.hidden) load(); }},2000);
load();
</script>"""
    return shell(g.principal, body, active="phone", bare=True)


@bp.get("/meet-events/<int:meid>/time/state")
@login_required
def time_state(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    hk = _heat_key(request.args.get("heat", ""))
    conn = db.connect()
    meids = _combine_meids(conn, me)
    clk = conn.execute("SELECT * FROM track_clocks WHERE meet_event_id=? AND heat=?", (meid, hk)).fetchone()
    taps = conn.execute("SELECT * FROM track_taps WHERE meet_event_id=? AND heat=? ORDER BY seq",
                        (meid, hk)).fetchall()
    ents = _tap_entries(conn, meids, hk)
    conn.close()
    start = _t_parse(clk["start_time"]) if clk else None
    stop = _t_parse(clk["stop_time"]) if clk else None
    return jsonify(
        server_ms=_t_ms(_t_now()),
        start_ms=_t_ms(start) if start else None,
        stop_ms=_t_ms(stop) if stop else None,
        started=bool(start), stopped=bool(stop),
        taps=[{"id": t["id"], "seq": t["seq"], "elapsed": t["elapsed_seconds"],
               "entry_id": t["entry_id"]} for t in taps],
        entries=ents)


@bp.post("/meet-events/<int:meid>/time/start")
@login_required
def time_start(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    hk = _heat_key(request.args.get("heat", ""))
    clear = bool((request.get_json(silent=True) or {}).get("clear"))
    conn = db.connect()
    clk = conn.execute("SELECT * FROM track_clocks WHERE meet_event_id=? AND heat=?", (meid, hk)).fetchone()
    n = conn.execute("SELECT COUNT(*) FROM track_taps WHERE meet_event_id=? AND heat=?", (meid, hk)).fetchone()[0]
    if clk and clk["start_time"] and clk["stop_time"] and n and not clear:
        conn.close()
        return jsonify(error=f"This heat already has an ended race with {n} tap(s).",
                       needs_clear=True), 409
    if clear:
        conn.execute("DELETE FROM track_taps WHERE meet_event_id=? AND heat=?", (meid, hk))
    conn.execute("INSERT INTO track_clocks (meet_event_id, heat, start_time, stop_time) VALUES (?,?,?,NULL) "
                 "ON CONFLICT(meet_event_id, heat) DO UPDATE SET start_time=excluded.start_time, stop_time=NULL",
                 (meid, hk, _t_iso(_t_now())))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/meet-events/<int:meid>/time/stop")
@login_required
def time_stop(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    hk = _heat_key(request.args.get("heat", ""))
    conn = db.connect()
    conn.execute("UPDATE track_clocks SET stop_time=? WHERE meet_event_id=? AND heat=? AND stop_time IS NULL",
                 (_t_iso(_t_now()), meid, hk))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/meet-events/<int:meid>/time/reset")
@login_required
def time_reset(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    hk = _heat_key(request.args.get("heat", ""))
    conn = db.connect()
    meids = _combine_meids(conn, me)
    # drop the results this heat produced (assigned taps), then the taps + clock
    eids = [t["entry_id"] for t in conn.execute(
        "SELECT entry_id FROM track_taps WHERE meet_event_id=? AND heat=? AND entry_id IS NOT NULL",
        (meid, hk)).fetchall()]
    for eid in eids:
        conn.execute("DELETE FROM results WHERE entry_id=?", (eid,))
    conn.execute("DELETE FROM track_taps WHERE meet_event_id=? AND heat=?", (meid, hk))
    conn.execute("DELETE FROM track_clocks WHERE meet_event_id=? AND heat=?", (meid, hk))
    for m2 in meids:
        _recompute_places(conn, load_meet_event(m2))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/meet-events/<int:meid>/time/tap")
@login_required
def time_tap(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        abort(403)
    hk = _heat_key(request.args.get("heat", ""))
    conn = db.connect()
    clk = conn.execute("SELECT * FROM track_clocks WHERE meet_event_id=? AND heat=?", (meid, hk)).fetchone()
    if not clk or not clk["start_time"] or clk["stop_time"]:
        conn.close()
        return jsonify(error="Race is not running."), 409
    elapsed = (_t_now() - _t_parse(clk["start_time"])).total_seconds()
    seq = (conn.execute("SELECT COALESCE(MAX(seq),0) FROM track_taps WHERE meet_event_id=? AND heat=?",
                        (meid, hk)).fetchone()[0]) + 1
    conn.execute("INSERT INTO track_taps (meet_event_id, heat, seq, elapsed_seconds) VALUES (?,?,?,?)",
                 (meid, hk, seq, elapsed))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/track-taps/<int:tid>/assign")
@login_required
def tap_assign(tid):
    conn = db.connect()
    t = conn.execute("SELECT * FROM track_taps WHERE id=?", (tid,)).fetchone()
    if not t:
        conn.close()
        abort(404)
    me = load_meet_event(t["meet_event_id"])
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        conn.close()
        abort(403)
    raw = (request.get_json(silent=True) or {}).get("entry_id")
    # unassigning (or reassigning) frees the entry the tap previously held
    if t["entry_id"]:
        conn.execute("DELETE FROM results WHERE entry_id=?", (t["entry_id"],))
    if raw in (None, "", 0, "0"):
        conn.execute("UPDATE track_taps SET entry_id=NULL WHERE id=?", (tid,))
    else:
        try:
            eid = int(raw)
        except (TypeError, ValueError):
            conn.close()
            return jsonify(error="Bad runner"), 400
        if conn.execute("SELECT 1 FROM track_taps WHERE entry_id=? AND id!=?", (eid, tid)).fetchone():
            conn.close()
            return jsonify(error="That runner is already assigned to another finish."), 400
        e = conn.execute("SELECT * FROM entries WHERE id=?", (eid,)).fetchone()
        if not e:
            conn.close()
            return jsonify(error="Unknown runner"), 400
        name, bib, school = _entry_label(conn, e)
        conn.execute("UPDATE track_taps SET entry_id=? WHERE id=?", (eid, tid))
        conn.execute("DELETE FROM results WHERE entry_id=?", (eid,))
        conn.execute("INSERT INTO results (entry_id, mark_seconds, dq, snap_name, snap_bib, snap_school) "
                     "VALUES (?,?,0,?,?,?)", (eid, t["elapsed_seconds"], name, bib, school))
    for m2 in _combine_meids(conn, me):
        _recompute_places(conn, load_meet_event(m2))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/track-taps/<int:tid>/assign-bib")
@login_required
def tap_assign_bib(tid):
    """Assign a finish to a runner who wasn't entered in this event: look the bib up
    among athletes at this meet, enter them on the spot, then assign as usual."""
    conn = db.connect()
    t = conn.execute("SELECT * FROM track_taps WHERE id=?", (tid,)).fetchone()
    if not t:
        conn.close()
        abort(404)
    me = load_meet_event(t["meet_event_id"])
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        conn.close()
        abort(403)
    try:
        bib = int(str((request.get_json(silent=True) or {}).get("bib", "")).strip())
    except (TypeError, ValueError):
        conn.close()
        return jsonify(error="Enter a valid bib number."), 400
    meids = _combine_meids(conn, me)
    qm = ",".join("?" * len(meids))
    # reuse an existing entry for this bib in the event (group), else enter them now
    e = conn.execute(
        f"SELECT en.* FROM entries en JOIN athletes a ON a.id=en.runner_id "
        f"WHERE en.meet_event_id IN ({qm}) AND a.bib=?", (*meids, bib)).fetchone()
    if not e:
        a = conn.execute(
            "SELECT a.id, a.school_id FROM athletes a "
            "JOIN schools s ON s.id=a.school_id JOIN meet_schools ms ON ms.school_id=s.id "
            "WHERE ms.meet_id=? AND a.bib=? AND a.active=1", (m["id"], bib)).fetchone()
        if not a:
            conn.close()
            return jsonify(error=f"No athlete with bib #{bib} at this meet."), 400
        eid = conn.execute("INSERT INTO entries (meet_event_id, runner_id, school_id, heat) "
                           "VALUES (?,?,?,?)", (me["id"], a["id"], a["school_id"], t["heat"] or None)).lastrowid
        e = conn.execute("SELECT * FROM entries WHERE id=?", (eid,)).fetchone()
    if conn.execute("SELECT 1 FROM track_taps WHERE entry_id=? AND id!=?", (e["id"], tid)).fetchone():
        conn.close()
        return jsonify(error="That runner is already assigned to another finish."), 400
    name, ebib, school = _entry_label(conn, e)
    if t["entry_id"]:
        conn.execute("DELETE FROM results WHERE entry_id=?", (t["entry_id"],))
    conn.execute("UPDATE track_taps SET entry_id=? WHERE id=?", (e["id"], tid))
    conn.execute("DELETE FROM results WHERE entry_id=?", (e["id"],))
    conn.execute("INSERT INTO results (entry_id, mark_seconds, dq, snap_name, snap_bib, snap_school) "
                 "VALUES (?,?,0,?,?,?)", (e["id"], t["elapsed_seconds"], name, ebib, school))
    for m2 in meids:
        _recompute_places(conn, load_meet_event(m2))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


@bp.post("/track-taps/<int:tid>/delete")
@login_required
def tap_delete(tid):
    conn = db.connect()
    t = conn.execute("SELECT * FROM track_taps WHERE id=?", (tid,)).fetchone()
    if not t:
        conn.close()
        abort(404)
    me = load_meet_event(t["meet_event_id"])
    m = load_meet(me["meet_id"])
    if not can_record_meet(m):
        conn.close()
        abort(403)
    if t["entry_id"]:
        conn.execute("DELETE FROM results WHERE entry_id=?", (t["entry_id"],))
    conn.execute("DELETE FROM track_taps WHERE id=?", (tid,))
    for m2 in _combine_meids(conn, me):
        _recompute_places(conn, load_meet_event(m2))
    conn.commit()
    conn.close()
    return jsonify(ok=True)


# ------------------------------- heat sheet PDF -------------------------------
@bp.get("/meet-events/<int:meid>/heatsheet.pdf")
@login_required
def heatsheet_pdf(meid):
    me = load_meet_event(meid)
    m = load_meet(me["meet_id"])
    if not can_view_meet(m):
        abort(403)
    conn = db.connect()
    entries = conn.execute(
        "SELECT * FROM entries WHERE meet_event_id=? ORDER BY heat, lane, id", (meid,)).fetchall()
    rows = []
    for e in entries:
        name, bib, school = _entry_label(conn, e)
        rows.append({"heat": e["heat"], "lane": e["lane"], "bib": bib,
                     "name": name, "school": school})
    conn.close()
    div = {"M": "Boys", "F": "Girls"}.get(me["gender"], "Open")
    title = f'{m["name"]} — {me["ename"]} ({div})'
    kind = ("hj" if _is_hj(me) else "field") if me["kind"] == "field" else "track"
    bars = None
    if _is_hj(me) and me["bar_heights"]:
        try:
            bars = sorted({b for b in json.loads(me["bar_heights"]) if _parse_ht(b) is not None},
                          key=_parse_ht)
        except (ValueError, TypeError):
            bars = None
    pdf = pdfs.heat_sheet_pdf(title, rows, laned=bool(me["laned"]),
                              token=f"XCTSHEET E{meid}", kind=kind, bars=bars)
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": f'inline; filename="heatsheet.pdf"'})


@bp.get("/meets/<int:mid>/heatsheets.pdf")
@login_required
def meet_heatsheets(mid):
    """Meet-wide heat sheets — every event with drawn heats, one packet."""
    m = load_meet(mid)
    if not can_view_meet(m):
        abort(403)
    conn = db.connect()
    mes = conn.execute(
        "SELECT me.id, me.gender, me.grade, me.bar_heights, e.name AS ename, e.laned, e.kind "
        "FROM meet_events me JOIN events e ON e.id=me.event_id WHERE me.meet_id=? "
        "ORDER BY e.sort, me.gender", (mid,)).fetchall()
    sections = []
    for me in mes:
        entries = conn.execute(
            "SELECT * FROM entries WHERE meet_event_id=? ORDER BY heat, lane, id", (me["id"],)).fetchall()
        rows = []
        for e in entries:
            name, bib, school = _entry_label(conn, e)
            rows.append({"heat": e["heat"], "lane": e["lane"], "bib": bib,
                         "name": name, "school": school})
        div = {"M": "Boys", "F": "Girls"}.get(me["gender"], "Open")
        title = f'{me["ename"]} — {div}' + (f' {me["grade"]}' if me["grade"] else "")
        kind = ("hj" if me["ename"] == "High Jump" else "field") if me["kind"] == "field" else "track"
        bars = None
        if kind == "hj" and me["bar_heights"]:
            try:
                bars = sorted({b for b in json.loads(me["bar_heights"]) if _parse_ht(b) is not None},
                              key=_parse_ht)
            except (ValueError, TypeError):
                bars = None
        sections.append((title, rows, bool(me["laned"]), f"XCTSHEET E{me['id']}", kind, bars))
    conn.close()
    pdf = pdfs.multi_heat_sheet_pdf(sections)
    return Response(pdf, mimetype="application/pdf",
                    headers={"Content-Disposition": 'inline; filename="meet-heatsheets.pdf"'})


# ------------------------------- results + scoring -------------------------------
def build_results(mid):
    """Per-event results + team point totals for a track meet."""
    conn = db.connect()
    meet = conn.execute("SELECT * FROM meets WHERE id=?", (mid,)).fetchone()
    vals, mult = _meet_points(meet)
    rel_vals = [round(v * mult) for v in vals]
    meet_events = conn.execute(
        "SELECT me.*, e.name AS ename, e.kind, e.unit, e.scoring_order, e.sort "
        "FROM meet_events me JOIN events e ON e.id=me.event_id WHERE me.meet_id=? "
        "ORDER BY e.sort, me.gender", (mid,)).fetchall()
    events_out = []
    team_pts = {}  # (school, gender) -> points
    for me in meet_events:
        table = rel_vals if me["kind"] == "relay" else vals
        rows = conn.execute(
            "SELECT r.*, en.school_id FROM results r JOIN entries en ON en.id=r.entry_id "
            "WHERE en.meet_event_id=? AND r.place IS NOT NULL ORDER BY r.place", (me["id"],)).fetchall()
        # Tie-aware points: athletes sharing a place split the sum of the positions
        # they occupy (e.g. two tied for 1st split points for places 1+2).
        place_counts = {}
        for r in rows:
            place_counts[r["place"]] = place_counts.get(r["place"], 0) + 1

        def pts_for(place):
            k = place_counts[place]
            total = sum(table[p - 1] if p - 1 < len(table) else 0 for p in range(place, place + k))
            return total / k

        items = []
        for r in rows:
            pts = pts_for(r["place"])
            # Track = time; field (LJ/HJ/SP) = feet-inches. Field shows all attempts.
            atts = ""
            if me["unit"] == "seconds":
                mark = fmt_time(r["mark_seconds"])
            else:
                mark = _fmt_ht(r["mark_metric"])
                if me["ename"] != "High Jump" and r["attempts_json"]:
                    try:
                        arr = json.loads(r["attempts_json"])
                        atts = ", ".join((_fmt_ht(x) if isinstance(x, (int, float)) else str(x))
                                         for x in arr if x not in (None, ""))
                    except (ValueError, TypeError):
                        atts = ""
            items.append({"place": r["place"], "mark": mark, "attempts": atts,
                          "name": r["snap_name"], "bib": r["snap_bib"],
                          "school": r["snap_school"], "points": pts})
            if r["snap_school"]:
                key = (r["snap_school"], me["gender"] or "U", me["grade"])
                team_pts[key] = team_pts.get(key, 0) + pts
        # DQ'd athletes stay visible at the bottom of the event (no place, no points)
        # — an official expects to see the DQ on the results, not a disappearance.
        for r in conn.execute(
                "SELECT r.* FROM results r JOIN entries en ON en.id=r.entry_id "
                "WHERE en.meet_event_id=? AND r.dq=1", (me["id"],)).fetchall():
            mark = fmt_time(r["mark_seconds"]) if me["unit"] == "seconds" else _fmt_ht(r["mark_metric"])
            items.append({"place": "DQ", "mark": mark or "", "attempts": "",
                          "name": r["snap_name"], "bib": r["snap_bib"],
                          "school": r["snap_school"], "points": 0})
        events_out.append({"name": f'{me["ename"]} — {_div_grade(me["gender"], me["grade"])}',
                           "items": items, "gkey": me["gender"] or "U"})
    conn.close()
    # Team totals broken down by gender × grade, plus a combined whole-meet total
    # (what gets announced at a dual: one school ranking across every division).
    totals, overall = {}, {}
    for (school, gender, grade), pts in team_pts.items():
        d = totals.setdefault((gender, grade), {})
        d[school] = d.get(school, 0) + pts
        overall[school] = overall.get(school, 0) + pts
    return {"events": events_out, "totals": totals, "overall": overall}


def _div_grade(gender, grade):
    """'F', 8 -> 'Girls 8th Grade'; gender-only when no grade."""
    g = {"M": "Boys", "F": "Girls"}.get(gender, "Open")
    return f"{g} {grade}th Grade" if grade else g


def _grp_sort(k):
    """Order team-score / event groups by grade then gender (Girls, Boys, Open)."""
    gender, grade = k
    return (grade if grade is not None else 999, {"F": 0, "M": 1}.get(gender, 2))


def _fmt_pts(x):
    if x in (None, 0):
        return ""
    return str(int(x)) if float(x).is_integer() else f"{x:.1f}"


_LIVE_CACHE = {}       # (mid, name_mode) -> (expires_monotonic, heats_list)
_LIVE_TTL = 1.0        # recompute the DB snapshot at most once/sec no matter the crowd
LIVE_HOLD = 30.0       # keep an ended heat on the public board this many seconds


def public_live(mid, name_mode=None):
    """Public 'live now' feed. server_ms is always fresh (spectator clocks stay
    accurate), but the heats snapshot is micro-cached ~1s so a crowd of spectators
    polling can't hammer the DB — 1 read/sec instead of one read per request."""
    key = (mid, name_mode)
    nowm = time.monotonic()
    hit = _LIVE_CACHE.get(key)
    if hit and hit[0] > nowm:
        heats = hit[1]
    else:
        heats = _live_heats(mid, name_mode)
        _LIVE_CACHE[key] = (nowm + _LIVE_TTL, heats)
    return {"server_ms": _t_ms(_t_now()), "heats": heats}


def _live_heats(mid, name_mode=None):
    """Track heats currently running, plus ones that ended within LIVE_HOLD seconds
    (kept on the board briefly so spectators can read the final result)."""
    conn = db.connect()
    rows = conn.execute(
        "SELECT tc.meet_event_id AS meid, tc.heat AS heat, tc.start_time, tc.stop_time "
        "FROM track_clocks tc JOIN meet_events me ON me.id=tc.meet_event_id "
        "WHERE me.meet_id=? AND tc.start_time IS NOT NULL "
        "ORDER BY tc.start_time", (mid,)).fetchall()
    now = _t_now()
    heats = []
    for r in rows:
        stop = _t_parse(r["stop_time"])
        if stop and (now - stop).total_seconds() > LIVE_HOLD:
            continue                       # ended more than 30s ago -> off the board
        me = load_meet_event(r["meid"])
        label = _div_grade(me["gender"], me["grade"])
        name = f'{me["ename"]} · {label}' + (f' · Heat {r["heat"]}' if r["heat"] else "")
        taps = conn.execute("SELECT * FROM track_taps WHERE meet_event_id=? AND heat=? ORDER BY seq",
                            (r["meid"], r["heat"])).fetchall()
        fin = []
        for i, t in enumerate(taps):
            who = school = None
            if t["entry_id"]:
                e = conn.execute("SELECT * FROM entries WHERE id=?", (t["entry_id"],)).fetchone()
                if e:
                    nm, _bib, sch = _entry_label(conn, e)
                    who, school = demo.public_ident(nm, _bib, name_mode), sch
            fin.append({"n": i + 1, "elapsed": t["elapsed_seconds"], "who": who, "school": school})
        heats.append({"name": name, "start_ms": _t_ms(_t_parse(r["start_time"])),
                      "stop_ms": _t_ms(stop) if stop else None, "ended": bool(stop),
                      "finishers": fin})
    conn.close()
    return heats


def results_inner(mid, name_mode=None):
    data = build_results(mid)
    if not data["events"]:
        return '<div class="card muted">No events yet.</div>'
    html = ['<div class="card">'
            '<input id="rsearch" placeholder="Search athlete / school / event…" oninput="rfilter()" '
            'style="max-width:320px"> '
            '<select id="rgender" onchange="rfilter()" style="max-width:120px;margin-left:.4rem">'
            '<option value="">All</option><option value="M">Boys</option>'
            '<option value="F">Girls</option></select></div>']
    # Combined whole-meet team score first (announced at duals) when >1 division scored.
    if data.get("overall") and len(data["totals"]) > 1:
        ranked = sorted(data["overall"].items(), key=lambda x: -x[1])
        trs = "".join(f'<tr data-text="{escape((s or "").lower())}"><td>{i+1}</td>'
                      f'<td>{escape(s)}</td><td><b>{_fmt_pts(p)}</b></td></tr>'
                      for i, (s, p) in enumerate(ranked))
        html.append(f'<div class="card rcard" data-gender="U" data-title="overall team scores">'
                    f'<h2>🏆 Overall — Team scores (all divisions)</h2>'
                    f'<table><tr><th>Rank</th><th>School</th><th>Points</th></tr>{trs}</table></div>')
    for (gender, grade) in sorted(data["totals"], key=_grp_sort):
        t = data["totals"][(gender, grade)]
        if not t:
            continue
        label = _div_grade(gender, grade)
        ranked = sorted(t.items(), key=lambda x: -x[1])
        trs = "".join(f'<tr data-text="{escape((s or "").lower())}"><td>{i+1}</td>'
                      f'<td>{escape(s)}</td><td><b>{_fmt_pts(p)}</b></td></tr>'
                      for i, (s, p) in enumerate(ranked))
        html.append(f'<div class="card rcard" data-gender="{gender or "U"}" '
                    f'data-title="{escape((label + " team scores").lower())}">'
                    f'<h2>{escape(label)} — Team scores</h2><table><tr><th>Rank</th><th>School</th>'
                    f'<th>Points</th></tr>{trs}</table></div>')
    for ev in data["events"]:
        if not ev["items"]:
            continue
        any_att = any(i.get("attempts") for i in ev["items"])
        trs = ""
        for i in ev["items"]:
            nm = demo.public_ident(i["name"] or "", i.get("bib"), name_mode)
            txt = f'{nm} {i["school"] or ""}'.lower()
            att_td = (f'<td class="muted">{escape(i.get("attempts") or "")}</td>' if any_att else "")
            trs += (f'<tr data-text="{escape(txt)}"><td>{i["place"]}</td>'
                    f'<td>{escape(nm)}</td><td>{escape(i["school"] or "")}</td>'
                    f'<td><b>{escape(i["mark"])}</b></td>{att_td}'
                    f'<td>{_fmt_pts(i["points"])}</td></tr>')
        att_th = "<th>Attempts</th>" if any_att else ""
        html.append(f'<div class="card rcard" data-gender="{ev["gkey"]}" '
                    f'data-title="{escape(ev["name"].lower())}"><h2>{escape(ev["name"])}</h2>'
                    f'<table><tr><th>Pl</th><th>Competitor</th><th>School</th>'
                    f'<th>Best</th>{att_th}<th>Pts</th></tr>{trs}</table></div>')
    html.append("""<script>
function _v(id){var e=document.getElementById(id);return e?e.value:'';}
function rfilter(){var q=_v('rsearch').toLowerCase(),g=_v('rgender');
 document.querySelectorAll('.rcard').forEach(function(c){
  var gok=!g||c.dataset.gender===g||c.dataset.gender==='U';
  var titleHit=q&&(c.dataset.title||'').indexOf(q)>=0,any=false;
  c.querySelectorAll('tr[data-text]').forEach(function(tr){
   var show=gok&&(!q||titleHit||tr.dataset.text.indexOf(q)>=0);
   tr.style.display=show?'':'none';if(show)any=true;});
  c.style.display=(gok&&any)?'':'none';});}
</script>""")
    return "".join(html)


@bp.get("/meets/<int:mid>/track-results")
@login_required
def results_page(mid):
    m = load_meet(mid)
    if not can_view_meet(m):
        abort(403)
    import os
    import base64
    import qrcode
    base = os.environ.get("XC_PUBLIC_URL", request.host_url.rstrip("/"))
    url = f"{base}/r/{m['public_token']}"
    b = io.BytesIO()
    qrcode.make(url).save(b, format="PNG")
    qr_uri = "data:image/png;base64," + base64.b64encode(b.getvalue()).decode()
    qr_card = (f'<div class="card" style="display:flex;gap:1rem;align-items:center;flex-wrap:wrap">'
               f'<img src="{qr_uri}" width="128" height="128" '
               f'style="background:#fff;padding:8px;border-radius:10px">'
               f'<div><b>Public results</b><br>'
               f'<span class="muted">Scan to open the live public results page — share on the big screen '
               f'or a flyer.</span><br><a href="{url}" target="_blank">{escape(url)}</a></div></div>')
    body = (f'<p class="muted"><a href="/meets/{mid}">← {escape(m["name"])}</a></p>'
            f'<h1>{escape(m["name"])} — Results</h1>{_track_tabs(mid, "results")}'
            f'<div class="row"><a class="btn ghost" href="/r/{m["public_token"]}" target="_blank">'
            f'Public page ↗</a> <a class="btn ghost" href="/meets/{mid}/track-results.xlsx">Export xlsx</a></div>'
            f'{qr_card}{results_inner(mid, name_mode=demo.mode_for(g.principal))}')
    return shell(g.principal, body, active="meets")


@bp.get("/meets/<int:mid>/track-results.xlsx")
@login_required
def results_xlsx(mid):
    m = load_meet(mid)
    if not can_view_meet(m):
        abort(403)
    fname = (m["name"] or "results").replace(" ", "_")
    return Response(track_workbook(mid),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})


def track_workbook(mid, name_mode=None):
    """Track results workbook bytes (Team Scores + Results). name_mode masks names
    for public downloads (district mask_public)."""
    import io as _io
    import openpyxl
    data = build_results(mid)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Team Scores")
    if data.get("overall") and len(data["totals"]) > 1:
        ws.append(["Overall — Team Scores (all divisions)"])
        ws.append(["Rank", "School", "Points"])
        for i, (s, p) in enumerate(sorted(data["overall"].items(), key=lambda x: -x[1])):
            ws.append([i + 1, s, _fmt_pts(p)])
        ws.append([])
    for (gender, grade) in sorted(data["totals"], key=_grp_sort):
        t = data["totals"][(gender, grade)]
        if not t:
            continue
        ws.append([f"{_div_grade(gender, grade)} — Team Scores"])
        ws.append(["Rank", "School", "Points"])
        for i, (s, p) in enumerate(sorted(t.items(), key=lambda x: -x[1])):
            ws.append([i + 1, s, _fmt_pts(p)])
        ws.append([])
    ws2 = wb.create_sheet("Results")
    for ev in data["events"]:
        if not ev["items"]:
            continue
        ws2.append([ev["name"]])
        ws2.append(["Place", "Competitor", "School", "Best", "Attempts", "Points"])
        for i in ev["items"]:
            ws2.append([i["place"], demo.public_ident(i["name"] or "", i.get("bib"), name_mode), i["school"],
                        i["mark"], i.get("attempts") or "", _fmt_pts(i["points"])])
        ws2.append([])
    if not wb.sheetnames:
        wb.create_sheet("Results").append(["No results yet"])
    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
